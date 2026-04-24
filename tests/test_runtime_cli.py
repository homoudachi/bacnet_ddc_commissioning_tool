import csv
import json
import pathlib
import shutil
import zipfile
import socket
import subprocess
import sys
import threading
import time
import unittest


ROOT = pathlib.Path(__file__).resolve().parents[1]
RUNTIME_CLI = ROOT / "tools" / "runtime" / "app.py"
FIXTURES = ROOT / "tests" / "fixtures"


def _run_runtime(*args: str) -> subprocess.CompletedProcess[str]:
    cmd = [sys.executable, str(RUNTIME_CLI), *args]
    return subprocess.run(cmd, capture_output=True, text=True, check=False)


def _build_i_am_packet(device_instance: int) -> bytes:
    object_identifier = (8 << 22) | (device_instance & 0x3FFFFF)
    apdu = b"\x10\x00\xc4" + object_identifier.to_bytes(4, "big") + b"\x22\x00\x91\x00"
    npdu = b"\x01\x00"
    payload = npdu + apdu
    bvlc = b"\x81\x0a" + (len(payload) + 4).to_bytes(2, "big")
    return bvlc + payload


def _bvlc_original_unicast(npdu_and_apdu: bytes) -> bytes:
    total = 4 + len(npdu_and_apdu)
    return b"\x81\x0a" + total.to_bytes(2, "big") + npdu_and_apdu


class _FakeBipUdpServer:
    """Minimal BACnet/IP UDP peer for loopback tests.

    Responds to Who-Is with a standards-shaped I-Am (via bacpypes3 when available),
    ReadProperty (presentValue) with encoded Complex ACKs, and WriteProperty with
    Simple ACK. Falls back to a fixed I-Am-only pattern when bacpypes3 is missing.
    """

    def __init__(
        self,
        device_instance: int,
        *,
        analog_input_present: float = 21.5,
        msv_present: int = 1,
        av_tacho_present: float = 0.0,
        av_supply_fan_present: float = 0.0,
        av_heat_present: float = 0.0,
        ao_valve_present: float = 0.0,
    ) -> None:
        self.device_instance = device_instance
        self._ai_present = float(analog_input_present)
        self._msv_present = int(msv_present)
        self._av_tacho_present = float(av_tacho_present)
        self._av_supply_fan_present = float(av_supply_fan_present)
        self._av_heat_present = float(av_heat_present)
        self._ao_valve_present = float(ao_valve_present)
        self._lock = threading.Lock()
        self._stop = threading.Event()
        self._ready = threading.Event()
        self._thread = threading.Thread(target=self._run, daemon=True)
        self.port = 0

    def start(self) -> None:
        self._thread.start()
        if not self._ready.wait(timeout=2):
            raise RuntimeError("fake bip server failed to start")

    def stop(self) -> None:
        self._stop.set()
        self._thread.join(timeout=2)

    def _iam_only(self, sock: socket.socket, addr: tuple[str, int]) -> None:
        sock.sendto(_build_i_am_packet(self.device_instance), addr)

    def _handle_bacpypes(self, sock: socket.socket, addr: tuple[str, int], apdu_bytes: bytes) -> None:
        from bacpypes3.apdu import (
            APDU,
            ConfirmedRequestPDU,
            ConfirmedServiceChoice,
            IAmRequest,
            ReadPropertyACK,
            ReadPropertyRequest,
            SimpleAckPDU,
            UnconfirmedRequestPDU,
            UnconfirmedServiceChoice,
            WritePropertyRequest,
        )
        from bacpypes3.basetypes import PropertyIdentifier, Segmentation
        from bacpypes3.constructeddata import Any
        from bacpypes3.primitivedata import ObjectIdentifier, Real, Unsigned
        from bacpypes3.pdu import IPv4Address, PDU

        src = IPv4Address(f"{addr[0]}:{addr[1]}")
        inc = APDU.decode(PDU(apdu_bytes))
        inc.pduSource = src

        if isinstance(inc, UnconfirmedRequestPDU):
            if int(inc.apduService) == int(UnconfirmedServiceChoice.whoIs):
                iam = IAmRequest(
                    iAmDeviceIdentifier=ObjectIdentifier(("device", self.device_instance)),
                    maxAPDULengthAccepted=Unsigned(1476),
                    segmentationSupported=Segmentation("noSegmentation"),
                    vendorID=Unsigned(0),
                    destination=src,
                )
                apdu_wire = iam.encode().encode().pduData
                sock.sendto(_bvlc_original_unicast(b"\x01\x00" + apdu_wire), addr)
            return

        if not isinstance(inc, ConfirmedRequestPDU):
            return

        svc = int(inc.apduService)
        if svc == int(ConfirmedServiceChoice.readProperty):
            req = ReadPropertyRequest.decode(inc)
            if str(req.propertyIdentifier) != "present-value":
                return
            ot = int(req.objectIdentifier[0])
            oi = int(req.objectIdentifier[1])
            with self._lock:
                ai_val = self._ai_present
                msv_val = self._msv_present
                av_tacho = self._av_tacho_present
                av_fan = self._av_supply_fan_present
                av_heat = self._av_heat_present
                ao_valve = self._ao_valve_present
            if ot == 0 and oi == 2:
                payload = Any(Real(ai_val))
            elif ot == 19 and oi == 50:
                payload = Any(Unsigned(msv_val))
            elif ot == 2 and oi == 1:
                payload = Any(Real(av_tacho))
            elif ot == 2 and oi == 3:
                payload = Any(Real(av_fan))
            elif ot == 2 and oi == 4:
                payload = Any(Real(av_heat))
            elif ot == 1 and oi == 5:
                payload = Any(Real(ao_valve))
            else:
                return
            ack = ReadPropertyACK(
                objectIdentifier=req.objectIdentifier,
                propertyIdentifier=req.propertyIdentifier,
                propertyValue=payload,
            )
            inner = ack.encode()
            inner.apduInvokeID = inc.apduInvokeID
            inner.apduSeg = 0
            inner.apduMor = 0
            inner.set_context(inc)
            wire = inner.encode().pduData
            sock.sendto(_bvlc_original_unicast(b"\x01\x00" + wire), addr)
            return

        if svc == int(ConfirmedServiceChoice.writeProperty):
            wreq = WritePropertyRequest.decode(inc)
            if str(wreq.propertyIdentifier) != "present-value":
                return
            ot_w = int(wreq.objectIdentifier[0])
            oi_w = int(wreq.objectIdentifier[1])
            if ot_w == 19 and oi_w == 50:
                new_val = int(wreq.propertyValue.cast_out(Unsigned))
                with self._lock:
                    self._msv_present = new_val
            elif ot_w == 2 and oi_w == 3:
                new_val = float(wreq.propertyValue.cast_out(Real))
                with self._lock:
                    self._av_supply_fan_present = new_val
            elif ot_w == 2 and oi_w == 4:
                new_val = float(wreq.propertyValue.cast_out(Real))
                with self._lock:
                    self._av_heat_present = new_val
            elif ot_w == 1 and oi_w == 5:
                new_val = float(wreq.propertyValue.cast_out(Real))
                with self._lock:
                    self._ao_valve_present = new_val
            else:
                return
            sack = SimpleAckPDU(service_choice=ConfirmedServiceChoice.writeProperty, context=inc)
            wire = sack.encode().pduData
            sock.sendto(_bvlc_original_unicast(b"\x01\x00" + wire), addr)

    def _extract_apdu(self, data: bytes) -> bytes | None:
        if len(data) < 4 or data[0] != 0x81 or data[1] not in (0x0A, 0x0B):
            return None
        length = int.from_bytes(data[2:4], "big")
        if length > len(data) or length < 4 + 2:
            return None
        inner = data[4:length]
        if len(inner) < 2:
            return None
        return inner[2:]

    def _run(self) -> None:
        sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        try:
            sock.bind(("127.0.0.1", 0))
            sock.settimeout(0.1)
            self.port = sock.getsockname()[1]
            self._ready.set()
            while not self._stop.is_set():
                try:
                    data, addr = sock.recvfrom(2048)
                except socket.timeout:
                    continue
                apdu_bytes = self._extract_apdu(data)
                if apdu_bytes is None:
                    continue
                try:
                    self._handle_bacpypes(sock, addr, apdu_bytes)
                except ModuleNotFoundError:
                    self._iam_only(sock, addr)
                except Exception:
                    if apdu_bytes and apdu_bytes[0] == 0x10:
                        self._iam_only(sock, addr)
        finally:
            sock.close()


class RuntimeCliTests(unittest.TestCase):
    def setUp(self) -> None:
        FIXTURES.mkdir(parents=True, exist_ok=True)
        self.run_dir = FIXTURES / "runtime-run"

    def tearDown(self) -> None:
        if self.run_dir.exists():
            shutil.rmtree(self.run_dir)

    def test_init_run_creates_layout_config_and_log(self) -> None:
        result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-001",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )

        self.assertEqual(0, result.returncode)
        self.assertTrue((self.run_dir / "config" / "runtime-config.json").exists())
        self.assertTrue((self.run_dir / "logs" / "events.jsonl").exists())
        self.assertTrue((self.run_dir / "state").exists())
        self.assertTrue((self.run_dir / "artifacts").exists())

        config = json.loads(
            (self.run_dir / "config" / "runtime-config.json").read_text(encoding="utf-8")
        )
        self.assertEqual("job-001", config["job_id"])
        self.assertIn("controllers_csv", config)

        lines = (
            self.run_dir / "logs" / "events.jsonl"
        ).read_text(encoding="utf-8").strip().splitlines()
        self.assertGreaterEqual(len(lines), 1)
        first_event = json.loads(lines[0])
        self.assertEqual("run_initialized", first_event["event"])

        logo = self.run_dir / "artifacts" / "branding" / "logo.png"
        self.assertTrue(logo.is_file())
        ref = ROOT / "docs" / "examples" / "branding" / "commissioning-logo-placeholder.png"
        self.assertEqual(ref.stat().st_size, logo.stat().st_size)

    def test_compile_import_uses_run_config_and_writes_state_outputs(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-compile",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)

        result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))

        self.assertEqual(0, result.returncode)
        runtime_job = self.run_dir / "state" / "runtime-job.json"
        report = self.run_dir / "state" / "import-report.json"
        self.assertTrue(runtime_job.exists())
        self.assertTrue(report.exists())

        report_obj = json.loads(report.read_text(encoding="utf-8"))
        self.assertTrue(report_obj["compile_ok"])

        lines = (
            self.run_dir / "logs" / "events.jsonl"
        ).read_text(encoding="utf-8").strip().splitlines()
        events = [json.loads(line)["event"] for line in lines]
        self.assertIn("import_compiled", events)

    def test_validate_import_writes_separate_artifacts(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-validate-import",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)
        first_mtime = (self.run_dir / "state" / "runtime-job.json").stat().st_mtime_ns

        val_dir = self.run_dir / "artifacts" / "import-validation-custom"
        result = _run_runtime(
            "validate-import",
            "--run-dir",
            str(self.run_dir),
            "--output-dir",
            str(val_dir),
        )
        self.assertEqual(0, result.returncode)
        self.assertTrue((val_dir / "runtime-job.json").exists())
        self.assertTrue((val_dir / "import-report.json").exists())
        second_mtime = (self.run_dir / "state" / "runtime-job.json").stat().st_mtime_ns
        self.assertEqual(first_mtime, second_mtime)

        lines = (
            self.run_dir / "logs" / "events.jsonl"
        ).read_text(encoding="utf-8").strip().splitlines()
        events = [json.loads(line)["event"] for line in lines]
        self.assertIn("import_validated", events)

    def test_print_job_graph_after_compile(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-print-graph",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)

        result = _run_runtime("print-job-graph", "--run-dir", str(self.run_dir))
        self.assertEqual(0, result.returncode)
        self.assertIn("job-print-graph", result.stdout)
        self.assertIn("FCU-01A", result.stdout)
        self.assertIn("read_allowlist=", result.stdout)
        self.assertIn("point_checkout=", result.stdout)
        self.assertRegex(result.stdout, r"FCU-01A.*point_checkout=2")
        self.assertRegex(result.stdout, r"skip_gated_steps=1")
        self.assertRegex(result.stdout, r"modulation_action_steps=2")

        lines = (
            self.run_dir / "logs" / "events.jsonl"
        ).read_text(encoding="utf-8").strip().splitlines()
        events = [json.loads(line)["event"] for line in lines]
        self.assertIn("job_graph_printed", events)

    def test_bacnet_point_checkout_errors_when_no_point_checkout(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-point-checkout-missing",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)
        init_flow_result = _run_runtime(
            "init-flow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "HRV-01",
        )
        self.assertEqual(0, init_flow_result.returncode)

        job_path = self.run_dir / "state" / "runtime-job.json"
        job = json.loads(job_path.read_text(encoding="utf-8"))
        for c in job["controllers"]:
            if c.get("controller_label") == "HRV-01":
                c["point_checkout"] = []
                break
        job_path.write_text(json.dumps(job, indent=2), encoding="utf-8")

        result = _run_runtime(
            "bacnet-point-checkout",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "HRV-01",
        )
        self.assertEqual(2, result.returncode)
        self.assertIn("no point_checkout", result.stdout)

    def test_bacnet_point_checkout_writes_artifact(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-point-checkout-artifact",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)
        init_flow_result = _run_runtime(
            "init-flow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
        )
        self.assertEqual(0, init_flow_result.returncode)

        result = _run_runtime(
            "bacnet-point-checkout",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
            "--timeout-seconds",
            "0.1",
            "--retries",
            "1",
        )
        self.assertEqual(2, result.returncode)
        artifact = self.run_dir / "artifacts" / "bacnet_point_checkout" / "FCU-01A.json"
        self.assertTrue(artifact.exists())
        payload = json.loads(artifact.read_text(encoding="utf-8"))
        self.assertGreaterEqual(payload["point_count"], 1)
        self.assertFalse(payload["all_read_ok"])

    def test_bacnet_read_rejects_object_not_on_read_allowlist(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-bacnet-read-deny",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)

        result = _run_runtime(
            "bacnet-read",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
            "--object-id",
            "av_supply_fan_command",
        )
        self.assertEqual(2, result.returncode)
        self.assertIn("commissioning_read_allowlist", result.stdout)

    def test_bacnet_read_ok_with_fake_bacnet_server(self) -> None:
        server = _FakeBipUdpServer(device_instance=21001, analog_input_present=21.5)
        server.start()
        try:
            time.sleep(0.05)
            self.run_dir.mkdir(parents=True, exist_ok=True)
            csv_path = self.run_dir / "controllers-local.csv"
            with csv_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(
                    handle,
                    fieldnames=[
                        "controller_label",
                        "profile_id",
                        "bacnet_device_instance",
                        "bacnet_ip",
                        "bacnet_port",
                        "building_floor",
                        "notes",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "controller_label": "FCU-LOCAL",
                        "profile_id": "fcu_2pipe_chw_electric_heat_v1",
                        "bacnet_device_instance": "21001",
                        "bacnet_ip": "127.0.0.1",
                        "bacnet_port": str(server.port),
                        "building_floor": "L01",
                        "notes": "test",
                    }
                )

            init_result = _run_runtime(
                "init-run",
                "--run-dir",
                str(self.run_dir),
                "--job-id",
                "job-bacnet-read-fake",
                "--controllers-csv",
                str(csv_path),
                "--profiles-dir",
                str(ROOT / "docs" / "examples"),
                "--scenarios-dir",
                str(ROOT / "docs" / "examples" / "simulator-scenarios"),
            )
            self.assertEqual(0, init_result.returncode)
            compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
            self.assertEqual(0, compile_result.returncode)

            result = _run_runtime(
                "bacnet-read",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-LOCAL",
                "--object-id",
                "ai_sat",
                "--timeout-seconds",
                "0.5",
                "--retries",
                "1",
            )
        finally:
            server.stop()

        self.assertEqual(0, result.returncode)
        payload = json.loads(result.stdout)
        self.assertEqual("read_ok", payload["status"])
        self.assertIn("21.5", payload.get("read", {}).get("value_str", ""))
        timeouts = payload.get("bacnet_timeouts", {})
        self.assertEqual(3.0, timeouts.get("who_is_timeout_seconds"))
        self.assertEqual(8.0, timeouts.get("apdu_timeout_seconds"))

    def test_bacnet_read_rejects_invalid_apdu_timeout(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-bacnet-apdu-invalid",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)

        result = _run_runtime(
            "bacnet-read",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
            "--object-id",
            "ai_sat",
            "--apdu-timeout",
            "0",
        )
        self.assertEqual(2, result.returncode)
        self.assertIn("apdu-timeout", result.stdout)

    def test_bacnet_read_custom_apdu_timeout_with_fake_server(self) -> None:
        server = _FakeBipUdpServer(device_instance=21001, analog_input_present=21.5)
        server.start()
        try:
            time.sleep(0.05)
            self.run_dir.mkdir(parents=True, exist_ok=True)
            csv_path = self.run_dir / "controllers-local.csv"
            with csv_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(
                    handle,
                    fieldnames=[
                        "controller_label",
                        "profile_id",
                        "bacnet_device_instance",
                        "bacnet_ip",
                        "bacnet_port",
                        "building_floor",
                        "notes",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "controller_label": "FCU-LOCAL",
                        "profile_id": "fcu_2pipe_chw_electric_heat_v1",
                        "bacnet_device_instance": "21001",
                        "bacnet_ip": "127.0.0.1",
                        "bacnet_port": str(server.port),
                        "building_floor": "L01",
                        "notes": "test",
                    }
                )

            init_result = _run_runtime(
                "init-run",
                "--run-dir",
                str(self.run_dir),
                "--job-id",
                "job-bacnet-read-apdu",
                "--controllers-csv",
                str(csv_path),
                "--profiles-dir",
                str(ROOT / "docs" / "examples"),
                "--scenarios-dir",
                str(ROOT / "docs" / "examples" / "simulator-scenarios"),
            )
            self.assertEqual(0, init_result.returncode)
            compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
            self.assertEqual(0, compile_result.returncode)

            result = _run_runtime(
                "bacnet-read",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-LOCAL",
                "--object-id",
                "ai_sat",
                "--timeout-seconds",
                "0.5",
                "--retries",
                "1",
                "--apdu-timeout",
                "15",
            )
        finally:
            server.stop()

        self.assertEqual(0, result.returncode)
        payload = json.loads(result.stdout)
        self.assertEqual("read_ok", payload["status"])
        self.assertEqual(15.0, payload.get("bacnet_timeouts", {}).get("apdu_timeout_seconds"))
        artifact = self.run_dir / "artifacts" / "bacnet_reads" / "FCU-LOCAL-ai_sat.json"
        self.assertTrue(artifact.exists())
        saved = json.loads(artifact.read_text(encoding="utf-8"))
        self.assertEqual(15.0, saved.get("bacnet_timeouts", {}).get("apdu_timeout_seconds"))

    def test_dry_run_bacnet_write_execute_ok_with_fake_bacnet_server(self) -> None:
        server = _FakeBipUdpServer(device_instance=21001, msv_present=1)
        server.start()
        try:
            time.sleep(0.05)
            self.run_dir.mkdir(parents=True, exist_ok=True)
            csv_path = self.run_dir / "controllers-local.csv"
            with csv_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(
                    handle,
                    fieldnames=[
                        "controller_label",
                        "profile_id",
                        "bacnet_device_instance",
                        "bacnet_ip",
                        "bacnet_port",
                        "building_floor",
                        "notes",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "controller_label": "FCU-LOCAL",
                        "profile_id": "fcu_2pipe_chw_electric_heat_v1",
                        "bacnet_device_instance": "21001",
                        "bacnet_ip": "127.0.0.1",
                        "bacnet_port": str(server.port),
                        "building_floor": "L01",
                        "notes": "test",
                    }
                )

            init_result = _run_runtime(
                "init-run",
                "--run-dir",
                str(self.run_dir),
                "--job-id",
                "job-bacnet-write-exec",
                "--controllers-csv",
                str(csv_path),
                "--profiles-dir",
                str(ROOT / "docs" / "examples"),
                "--scenarios-dir",
                str(ROOT / "docs" / "examples" / "simulator-scenarios"),
            )
            self.assertEqual(0, init_result.returncode)
            compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
            self.assertEqual(0, compile_result.returncode)

            write_result = _run_runtime(
                "dry-run-bacnet-write",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-LOCAL",
                "--object-id",
                "msv_test_mode",
                "--value",
                "3",
                "--technician-name",
                "Alex Tech",
                "--note",
                "Execute against fake device",
                "--execute",
                "--timeout-seconds",
                "0.5",
                "--retries",
                "1",
                "--apdu-timeout",
                "12",
            )
            self.assertEqual(0, write_result.returncode)
            write_payload = json.loads(write_result.stdout)
            self.assertEqual("write_ok", write_payload["status"])
            self.assertEqual(
                {"who_is_timeout_seconds": 3.0, "apdu_timeout_seconds": 12.0},
                write_payload.get("bacnet_timeouts"),
            )
            plan_path = (
                self.run_dir / "artifacts" / "bacnet_write_plans" / "FCU-LOCAL-msv_test_mode.json"
            )
            self.assertTrue(plan_path.exists())
            plan_saved = json.loads(plan_path.read_text(encoding="utf-8"))
            self.assertEqual(12.0, plan_saved.get("bacnet_timeouts", {}).get("apdu_timeout_seconds"))

            read_result = _run_runtime(
                "bacnet-read",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-LOCAL",
                "--object-id",
                "msv_test_mode",
                "--timeout-seconds",
                "0.5",
                "--retries",
                "1",
            )
        finally:
            server.stop()

        self.assertEqual(0, read_result.returncode)
        read_payload = json.loads(read_result.stdout)
        self.assertEqual("read_ok", read_payload["status"])
        self.assertIn("3", read_payload.get("read", {}).get("value_str", ""))

    def test_bacnet_point_checkout_rejects_invalid_apdu_timeout(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-point-apdu-invalid",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)

        result = _run_runtime(
            "bacnet-point-checkout",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
            "--apdu-timeout",
            "-1",
        )
        self.assertEqual(2, result.returncode)
        self.assertIn("apdu-timeout", result.stdout)

    def test_bacnet_point_checkout_all_ok_with_fake_bacnet_server(self) -> None:
        server = _FakeBipUdpServer(device_instance=21001, analog_input_present=22.0, msv_present=2)
        server.start()
        try:
            time.sleep(0.05)
            self.run_dir.mkdir(parents=True, exist_ok=True)
            csv_path = self.run_dir / "controllers-local.csv"
            with csv_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(
                    handle,
                    fieldnames=[
                        "controller_label",
                        "profile_id",
                        "bacnet_device_instance",
                        "bacnet_ip",
                        "bacnet_port",
                        "building_floor",
                        "notes",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "controller_label": "FCU-LOCAL",
                        "profile_id": "fcu_2pipe_chw_electric_heat_v1",
                        "bacnet_device_instance": "21001",
                        "bacnet_ip": "127.0.0.1",
                        "bacnet_port": str(server.port),
                        "building_floor": "L01",
                        "notes": "test",
                    }
                )

            init_result = _run_runtime(
                "init-run",
                "--run-dir",
                str(self.run_dir),
                "--job-id",
                "job-point-checkout-fake",
                "--controllers-csv",
                str(csv_path),
                "--profiles-dir",
                str(ROOT / "docs" / "examples"),
                "--scenarios-dir",
                str(ROOT / "docs" / "examples" / "simulator-scenarios"),
            )
            self.assertEqual(0, init_result.returncode)
            compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
            self.assertEqual(0, compile_result.returncode)

            result = _run_runtime(
                "bacnet-point-checkout",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-LOCAL",
                "--timeout-seconds",
                "0.5",
                "--retries",
                "1",
                "--apdu-timeout",
                "9",
            )
        finally:
            server.stop()

        self.assertEqual(0, result.returncode)
        payload = json.loads(result.stdout)
        self.assertTrue(payload["all_read_ok"])
        self.assertEqual(2, payload["point_count"])
        statuses = [r.get("status") for r in payload.get("reads", [])]
        self.assertEqual(["read_ok", "read_ok"], statuses)
        for row in payload["reads"]:
            self.assertEqual(9.0, row.get("bacnet_timeouts", {}).get("apdu_timeout_seconds"))
        self.assertIn("22.0", payload["reads"][0].get("read", {}).get("value_str", ""))
        self.assertIn("2", payload["reads"][1].get("read", {}).get("value_str", ""))

    def test_export_run_summary_requires_runtime_job(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-export-no-compile",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        out_path = self.run_dir / "artifacts" / "custom-summary.json"
        result = _run_runtime(
            "export-run-summary",
            "--run-dir",
            str(self.run_dir),
            "--output-json",
            str(out_path),
        )
        self.assertEqual(2, result.returncode)
        self.assertIn("compile-import", result.stdout)

    def test_export_run_summary_after_compile_and_init_flow(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-export-summary",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)

        result = _run_runtime("export-run-summary", "--run-dir", str(self.run_dir))
        self.assertEqual(0, result.returncode)
        default_path = self.run_dir / "artifacts" / "run-summary.json"
        self.assertTrue(default_path.exists())
        summary = json.loads(default_path.read_text(encoding="utf-8"))
        self.assertEqual("0.1-run-summary", summary["schema_version"])
        self.assertEqual("job-export-summary", summary["job_id"])
        self.assertEqual(3, len(summary["controllers"]))
        for row in summary["controllers"]:
            self.assertFalse(row["flow_initialized"])
            self.assertIsNone(row["next_open_step"])

        init_flow = _run_runtime(
            "init-flow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
        )
        self.assertEqual(0, init_flow.returncode)
        record = _run_runtime(
            "record-step",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
            "--step-id",
            "half_design_airflow_auto",
            "--status",
            "passed",
            "--technician-name",
            "Alex Tech",
            "--note",
            "ok",
        )
        self.assertEqual(0, record.returncode)

        out2 = self.run_dir / "artifacts" / "summary-after-step.json"
        result2 = _run_runtime(
            "export-run-summary",
            "--run-dir",
            str(self.run_dir),
            "--output-json",
            str(out2),
        )
        self.assertEqual(0, result2.returncode)
        summary2 = json.loads(out2.read_text(encoding="utf-8"))
        fcu = [r for r in summary2["controllers"] if r["controller_label"] == "FCU-01A"][0]
        self.assertTrue(fcu["flow_initialized"])
        self.assertEqual("confirm_tachometer_reference_half_flow", fcu["next_open_step"]["step_id"])
        self.assertEqual("pending", fcu["next_open_step"]["status"])

        log_lines = (
            self.run_dir / "logs" / "events.jsonl"
        ).read_text(encoding="utf-8").strip().splitlines()
        events = [json.loads(line)["event"] for line in log_lines]
        self.assertIn("run_summary_exported", events)

    def test_export_run_summary_writes_csv_when_requested(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-export-csv",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)

        csv_path = self.run_dir / "artifacts" / "run-summary.csv"
        result = _run_runtime(
            "export-run-summary",
            "--run-dir",
            str(self.run_dir),
            "--output-csv",
            str(csv_path),
        )
        self.assertEqual(0, result.returncode)
        self.assertTrue(csv_path.exists())
        text = csv_path.read_text(encoding="utf-8")
        self.assertIn("controller_label", text)
        self.assertIn("FCU-01A", text)

    def test_export_run_summary_embed_import_and_bip_blobs(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-export-embed",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)
        bip_result = _run_runtime(
            "verify-bip-list",
            "--run-dir",
            str(self.run_dir),
            "--timeout-seconds",
            "0.1",
            "--retries",
            "1",
        )
        self.assertEqual(2, bip_result.returncode)

        out_path = self.run_dir / "artifacts" / "summary-embedded.json"
        result = _run_runtime(
            "export-run-summary",
            "--run-dir",
            str(self.run_dir),
            "--output-json",
            str(out_path),
            "--embed-import-report",
            "--embed-bip-list-summary",
        )
        self.assertEqual(0, result.returncode)
        summary = json.loads(out_path.read_text(encoding="utf-8"))
        self.assertIn("import_report", summary)
        self.assertTrue(summary["import_report"]["compile_ok"])
        self.assertIn("bip_list_summary", summary)
        self.assertEqual(3, summary["bip_list_summary"]["total"])

    def test_list_flows_empty_when_no_flow_state(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-list-flows-empty",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)

        result = _run_runtime("list-flows", "--run-dir", str(self.run_dir))
        self.assertEqual(0, result.returncode)
        payload = json.loads(result.stdout)
        self.assertEqual(0, payload["flow_count"])
        self.assertEqual([], payload["flows"])

    def test_list_flows_and_show_flow_after_init_flow(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-list-flows",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)
        init_flow_result = _run_runtime(
            "init-flow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
        )
        self.assertEqual(0, init_flow_result.returncode)

        list_result = _run_runtime("list-flows", "--run-dir", str(self.run_dir))
        self.assertEqual(0, list_result.returncode)
        listed = json.loads(list_result.stdout)
        self.assertEqual(1, listed["flow_count"])
        self.assertEqual("FCU-01A", listed["flows"][0]["controller_label"])
        self.assertEqual(
            "fcu_2pipe_chw_electric_heat_v1", listed["flows"][0]["profile_id"]
        )
        self.assertGreater(listed["flows"][0]["step_count"], 0)
        self.assertEqual(
            listed["flows"][0]["step_count"],
            listed["flows"][0]["status_counts"].get("pending", 0),
        )

        show_result = _run_runtime(
            "show-flow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
        )
        self.assertEqual(0, show_result.returncode)
        detail = json.loads(show_result.stdout)
        self.assertEqual("FCU-01A", detail["controller_label"])
        self.assertIn("steps", detail)
        self.assertGreater(len(detail["steps"]), 0)
        self.assertEqual("pending", detail["steps"][0]["status"])

        log_lines = (
            self.run_dir / "logs" / "events.jsonl"
        ).read_text(encoding="utf-8").strip().splitlines()
        events = [json.loads(line)["event"] for line in log_lines]
        self.assertIn("flows_listed", events)
        self.assertIn("flow_viewed", events)

    def test_show_flow_errors_when_controller_has_no_flow_state(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-show-flow-missing",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)

        result = _run_runtime(
            "show-flow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "HRV-01",
        )
        self.assertEqual(2, result.returncode)
        self.assertIn("flow state not found", result.stdout)

    def test_set_session_value_requires_init_flow(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-session-no-flow",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)

        result = _run_runtime(
            "set-session-value",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
            "--key",
            "rat_degC",
            "--value",
            "22.5",
            "--technician-name",
            "Alex Tech",
            "--note",
            "Manual RAT",
        )
        self.assertEqual(2, result.returncode)
        self.assertIn("init-flow first", result.stdout)

    def test_set_session_value_and_show_session_round_trip(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-session-roundtrip",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)
        init_flow_result = _run_runtime(
            "init-flow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
        )
        self.assertEqual(0, init_flow_result.returncode)

        set_result = _run_runtime(
            "set-session-value",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
            "--key",
            "rat_degC",
            "--value",
            "22.5",
            "--technician-name",
            "Alex Tech",
            "--note",
            "Manual RAT for heat-rise",
        )
        self.assertEqual(0, set_result.returncode)
        self.assertIn("session_value_set=true", set_result.stdout)

        session_path = self.run_dir / "state" / "sessions" / "FCU-01A.json"
        self.assertTrue(session_path.exists())
        stored = json.loads(session_path.read_text(encoding="utf-8"))
        self.assertEqual("FCU-01A", stored["controller_label"])
        self.assertEqual("22.5", stored["values"]["rat_degC"]["value"])
        self.assertEqual("Alex Tech", stored["values"]["rat_degC"]["technician_name"])

        show_result = _run_runtime(
            "show-session",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
        )
        self.assertEqual(0, show_result.returncode)
        shown = json.loads(show_result.stdout)
        self.assertEqual("22.5", shown["values"]["rat_degC"]["value"])

        log_lines = (
            self.run_dir / "logs" / "events.jsonl"
        ).read_text(encoding="utf-8").strip().splitlines()
        events = [json.loads(line)["event"] for line in log_lines]
        self.assertIn("session_value_set", events)
        self.assertIn("session_viewed", events)

    def test_show_session_errors_when_missing(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-show-session-missing",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)

        result = _run_runtime(
            "show-session",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
        )
        self.assertEqual(2, result.returncode)
        self.assertIn("session state not found", result.stdout)

    def test_init_flow_rejects_second_init_without_force(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-init-twice",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)
        first = _run_runtime(
            "init-flow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
        )
        self.assertEqual(0, first.returncode)
        second = _run_runtime(
            "init-flow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
        )
        self.assertEqual(2, second.returncode)
        self.assertIn("already exists", second.stdout)

    def test_init_flow_force_replaces_state_and_backups_prior(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-init-force",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)
        first = _run_runtime(
            "init-flow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
        )
        self.assertEqual(0, first.returncode)
        record = _run_runtime(
            "record-step",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
            "--step-id",
            "half_design_airflow_auto",
            "--status",
            "passed",
            "--technician-name",
            "Alex Tech",
            "--note",
            "first run",
        )
        self.assertEqual(0, record.returncode)

        second = _run_runtime(
            "init-flow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
            "--force",
            "--reset-technician-name",
            "Lead Tech",
            "--reset-reason",
            "Wrong controller row; restarting commissioning",
        )
        self.assertEqual(0, second.returncode)

        flow_state = json.loads(
            (self.run_dir / "state" / "flows" / "FCU-01A.json").read_text(encoding="utf-8")
        )
        self.assertEqual("pending", flow_state["steps"][0]["status"])
        backups = list((self.run_dir / "state" / "flow_backups").glob("FCU-01A-*.json"))
        self.assertEqual(1, len(backups))
        prior = json.loads(backups[0].read_text(encoding="utf-8"))
        self.assertEqual("passed", prior["steps"][0]["status"])

        log_lines = (
            self.run_dir / "logs" / "events.jsonl"
        ).read_text(encoding="utf-8").strip().splitlines()
        events = [json.loads(line)["event"] for line in log_lines]
        self.assertIn("flow_reinitialized", events)

    def test_init_flow_force_requires_audit_fields(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-init-force-audit",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)
        first = _run_runtime(
            "init-flow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
        )
        self.assertEqual(0, first.returncode)

        bad = _run_runtime(
            "init-flow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
            "--force",
            "--reset-technician-name",
            "",
            "--reset-reason",
            "",
        )
        self.assertEqual(2, bad.returncode)
        self.assertIn("--reset-technician-name", bad.stdout)

    def test_dry_run_bacnet_write_rejects_non_allowlisted_object(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-drywrite-deny",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)

        result = _run_runtime(
            "dry-run-bacnet-write",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
            "--object-id",
            "av_supply_fan_command",
            "--value",
            "50",
            "--technician-name",
            "Alex Tech",
            "--note",
            "Should be blocked",
        )
        self.assertEqual(2, result.returncode)
        self.assertIn("commissioning_write_allowlist", result.stdout)

    def test_dry_run_bacnet_write_planned_with_localhost_udp_server(self) -> None:
        server = _FakeBipUdpServer(device_instance=21001)
        server.start()
        try:
            time.sleep(0.05)
            self.run_dir.mkdir(parents=True, exist_ok=True)
            csv_path = self.run_dir / "controllers-local.csv"
            with csv_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(
                    handle,
                    fieldnames=[
                        "controller_label",
                        "profile_id",
                        "bacnet_device_instance",
                        "bacnet_ip",
                        "bacnet_port",
                        "building_floor",
                        "notes",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "controller_label": "FCU-LOCAL",
                        "profile_id": "fcu_2pipe_chw_electric_heat_v1",
                        "bacnet_device_instance": "21001",
                        "bacnet_ip": "127.0.0.1",
                        "bacnet_port": str(server.port),
                        "building_floor": "L01",
                        "notes": "test",
                    }
                )

            init_result = _run_runtime(
                "init-run",
                "--run-dir",
                str(self.run_dir),
                "--job-id",
                "job-drywrite-ok",
                "--controllers-csv",
                str(csv_path),
                "--profiles-dir",
                str(ROOT / "docs" / "examples"),
                "--scenarios-dir",
                str(ROOT / "docs" / "examples" / "simulator-scenarios"),
            )
            self.assertEqual(0, init_result.returncode)
            compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
            self.assertEqual(0, compile_result.returncode)

            result = _run_runtime(
                "dry-run-bacnet-write",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-LOCAL",
                "--object-id",
                "msv_test_mode",
                "--value",
                "3",
                "--technician-name",
                "Alex Tech",
                "--note",
                "Arm airflow verify mode",
                "--timeout-seconds",
                "0.5",
                "--retries",
                "1",
            )
        finally:
            server.stop()

        self.assertEqual(0, result.returncode)
        payload = json.loads(result.stdout)
        self.assertEqual("dry_run_allowed", payload["status"])
        self.assertEqual(19, payload["target"]["object_type"])
        self.assertEqual(50, payload["target"]["object_instance"])
        artifact = (
            self.run_dir / "artifacts" / "bacnet_write_plans" / "FCU-LOCAL-msv_test_mode.json"
        )
        self.assertTrue(artifact.exists())

        log_lines = (
            self.run_dir / "logs" / "events.jsonl"
        ).read_text(encoding="utf-8").strip().splitlines()
        events = [json.loads(line)["event"] for line in log_lines]
        self.assertIn("bacnet_write_planned", events)

    def test_verify_simulator_writes_artifact_and_logs_event(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-verify",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)

        result = _run_runtime(
            "verify-simulator",
            "--run-dir",
            str(self.run_dir),
            "--profile",
            "ci",
            "--scenario",
            "happy-path",
            "--strict",
        )

        self.assertEqual(0, result.returncode)
        artifact = self.run_dir / "artifacts" / "simulator" / "ci-happy-path.json"
        self.assertTrue(artifact.exists())
        summary = json.loads(artifact.read_text(encoding="utf-8"))
        self.assertTrue(summary["strict_pass"])
        self.assertEqual("ci", summary["profile"])

        lines = (
            self.run_dir / "logs" / "events.jsonl"
        ).read_text(encoding="utf-8").strip().splitlines()
        events = [json.loads(line)["event"] for line in lines]
        self.assertIn("simulator_verified", events)

    def test_probe_bip_writes_artifact_and_logs_event(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-bip-probe",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)

        result = _run_runtime(
            "probe-bip",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
            "--timeout-seconds",
            "0.1",
            "--retries",
            "1",
        )

        self.assertIn(result.returncode, (0, 2))
        artifact = self.run_dir / "artifacts" / "bip" / "FCU-01A.json"
        self.assertTrue(artifact.exists())
        summary = json.loads(artifact.read_text(encoding="utf-8"))
        self.assertEqual("FCU-01A", summary["controller_label"])
        self.assertIn(summary["status"], {"reachable_verified", "identity_mismatch", "unreachable_timeout"})

        lines = (
            self.run_dir / "logs" / "events.jsonl"
        ).read_text(encoding="utf-8").strip().splitlines()
        events = [json.loads(line)["event"] for line in lines]
        self.assertIn("bip_probed", events)

    def test_verify_bip_list_writes_summary_artifact_and_logs_event(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-bip-list",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)

        result = _run_runtime(
            "verify-bip-list",
            "--run-dir",
            str(self.run_dir),
            "--timeout-seconds",
            "0.1",
            "--retries",
            "1",
            "--strict",
        )

        self.assertEqual(2, result.returncode)
        summary_path = self.run_dir / "artifacts" / "bip" / "list-summary.json"
        self.assertTrue(summary_path.exists())
        summary = json.loads(summary_path.read_text(encoding="utf-8"))
        self.assertEqual(3, summary["total"])
        self.assertIn("status_counts", summary)
        self.assertFalse(summary["strict_pass"])

        lines = (
            self.run_dir / "logs" / "events.jsonl"
        ).read_text(encoding="utf-8").strip().splitlines()
        events = [json.loads(line)["event"] for line in lines]
        self.assertIn("bip_list_verified", events)

    def test_verify_bip_list_non_strict_allows_known_unavailable(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-bip-list-nonstrict",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)

        overrides_path = self.run_dir / "config" / "bip-known-unavailable.json"
        overrides_path.write_text(
            json.dumps(
                {
                    "controller_labels": ["FCU-01A", "FCU-01B", "HRV-01"],
                    "allow_known_unavailable": True,
                },
                indent=2,
            ),
            encoding="utf-8",
        )

        result = _run_runtime(
            "verify-bip-list",
            "--run-dir",
            str(self.run_dir),
            "--timeout-seconds",
            "0.1",
            "--retries",
            "1",
            "--known-unavailable-file",
            str(overrides_path),
        )

        self.assertEqual(0, result.returncode)
        summary = json.loads(
            (self.run_dir / "artifacts" / "bip" / "list-summary.json").read_text(
                encoding="utf-8"
            )
        )
        self.assertTrue(summary["strict_pass"])
        self.assertEqual(3, summary["status_counts"].get("known_unavailable", 0))

    def test_init_flow_creates_controller_flow_state(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-flow-init",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)

        result = _run_runtime(
            "init-flow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
        )

        self.assertEqual(0, result.returncode)
        flow_state_path = self.run_dir / "state" / "flows" / "FCU-01A.json"
        self.assertTrue(flow_state_path.exists())
        flow_state = json.loads(flow_state_path.read_text(encoding="utf-8"))
        self.assertEqual("FCU-01A", flow_state["controller_label"])
        self.assertGreater(len(flow_state["steps"]), 0)
        self.assertEqual("pending", flow_state["steps"][0]["status"])

    def test_init_flow_persists_step_policy_and_history_fields(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-flow-init-policy",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)

        result = _run_runtime(
            "init-flow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
        )
        self.assertEqual(0, result.returncode)

        flow_state_path = self.run_dir / "state" / "flows" / "FCU-01A.json"
        flow_state = json.loads(flow_state_path.read_text(encoding="utf-8"))
        target_step = [s for s in flow_state["steps"] if s["step_id"] == "half_design_airflow_auto"][
            0
        ]
        self.assertIn("skippable", target_step)
        self.assertIsInstance(target_step["skippable"], bool)
        self.assertIn("requires_step_ids", target_step)
        self.assertIsInstance(target_step["requires_step_ids"], list)
        self.assertIn("history", target_step)
        self.assertEqual([], target_step["history"])

    def test_record_step_updates_status_and_captures_technician_signoff(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-flow-record",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)
        init_flow_result = _run_runtime(
            "init-flow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
        )
        self.assertEqual(0, init_flow_result.returncode)

        result = _run_runtime(
            "record-step",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
            "--step-id",
            "half_design_airflow_auto",
            "--status",
            "passed",
            "--technician-name",
            "Alex Tech",
            "--note",
            "Reached target airflow in tolerance",
        )

        self.assertEqual(0, result.returncode)
        flow_state = json.loads(
            (self.run_dir / "state" / "flows" / "FCU-01A.json").read_text(
                encoding="utf-8"
            )
        )
        step = [s for s in flow_state["steps"] if s["step_id"] == "half_design_airflow_auto"][
            0
        ]
        self.assertEqual("passed", step["status"])
        self.assertEqual("Alex Tech", step["technician_name"])
        self.assertIn("Reached target airflow", step["note"])
        self.assertIn("history", step)
        self.assertGreaterEqual(len(step["history"]), 1)
        self.assertEqual("pending", step["history"][-1]["previous_status"])
        self.assertEqual("passed", step["history"][-1]["new_status"])
        self.assertEqual("status_update", step["history"][-1]["reason_code"])

        lines = (
            self.run_dir / "logs" / "events.jsonl"
        ).read_text(encoding="utf-8").strip().splitlines()
        parsed_events = [json.loads(line) for line in lines]
        events = [entry["event"] for entry in parsed_events]
        self.assertIn("flow_initialized", events)
        self.assertIn("flow_step_recorded", events)
        flow_step_event = [entry for entry in parsed_events if entry["event"] == "flow_step_recorded"][-1]
        self.assertEqual("pending", flow_step_event["previous_status"])
        self.assertEqual("passed", flow_step_event["new_status"])
        self.assertEqual("status_update", flow_step_event["reason_code"])

    def test_record_step_rejects_out_of_order_transition(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-flow-ordering",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)
        init_flow_result = _run_runtime(
            "init-flow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
        )
        self.assertEqual(0, init_flow_result.returncode)

        # Attempt to complete step 2 before step 1.
        result = _run_runtime(
            "record-step",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
            "--step-id",
            "confirm_tachometer_reference_half_flow",
            "--status",
            "passed",
            "--technician-name",
            "Alex Tech",
            "--note",
            "Tried to skip ahead",
        )

        self.assertEqual(2, result.returncode)
        self.assertIn("invalid step transition", result.stdout)
        self.assertIn("cannot be marked passed before", result.stdout)

        lines = (
            self.run_dir / "logs" / "events.jsonl"
        ).read_text(encoding="utf-8").strip().splitlines()
        parsed_events = [json.loads(line) for line in lines]
        rejection_events = [entry for entry in parsed_events if entry["event"] == "flow_step_rejected"]
        self.assertGreaterEqual(len(rejection_events), 1)
        self.assertEqual("PREREQ_ORDER", rejection_events[-1]["reason_code"])

    def test_record_step_rejects_skip_when_step_not_marked_skippable(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-flow-skip-rule",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)
        init_flow_result = _run_runtime(
            "init-flow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
        )
        self.assertEqual(0, init_flow_result.returncode)

        result = _run_runtime(
            "record-step",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
            "--step-id",
            "half_design_airflow_auto",
            "--status",
            "skipped",
            "--technician-name",
            "Alex Tech",
            "--note",
            "Skipping first step",
        )

        self.assertEqual(2, result.returncode)
        self.assertIn("invalid step transition", result.stdout)
        self.assertIn("is not skippable", result.stdout)

        lines = (
            self.run_dir / "logs" / "events.jsonl"
        ).read_text(encoding="utf-8").strip().splitlines()
        parsed_events = [json.loads(line) for line in lines]
        rejection_events = [entry for entry in parsed_events if entry["event"] == "flow_step_rejected"]
        self.assertGreaterEqual(len(rejection_events), 1)
        self.assertEqual("STEP_NOT_SKIPPABLE", rejection_events[-1]["reason_code"])

    def test_record_step_rejects_when_explicit_dependency_not_satisfied(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-flow-explicit-dependency",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)
        init_flow_result = _run_runtime(
            "init-flow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
        )
        self.assertEqual(0, init_flow_result.returncode)

        # Inject an explicit dependency on a later step to prove dependency checks
        # are enforced independently of index ordering.
        flow_state_path = self.run_dir / "state" / "flows" / "FCU-01A.json"
        flow_state = json.loads(flow_state_path.read_text(encoding="utf-8"))
        for step in flow_state["steps"]:
            if step["step_id"] == "half_design_airflow_auto":
                step["requires_step_ids"] = ["heating_test"]
                break
        flow_state_path.write_text(json.dumps(flow_state, indent=2), encoding="utf-8")

        result = _run_runtime(
            "record-step",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
            "--step-id",
            "half_design_airflow_auto",
            "--status",
            "passed",
            "--technician-name",
            "Alex Tech",
            "--note",
            "Attempting explicit dependency bypass",
        )

        self.assertEqual(2, result.returncode)
        self.assertIn("invalid step transition", result.stdout)
        self.assertIn("requires completed dependency", result.stdout)

        lines = (
            self.run_dir / "logs" / "events.jsonl"
        ).read_text(encoding="utf-8").strip().splitlines()
        parsed_events = [json.loads(line) for line in lines]
        rejection_events = [entry for entry in parsed_events if entry["event"] == "flow_step_rejected"]
        self.assertGreaterEqual(len(rejection_events), 1)
        self.assertEqual("DEPENDENCY_UNSATISFIED", rejection_events[-1]["reason_code"])

    def test_record_step_rejection_appends_history_and_preserves_status(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-flow-rejection-history",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)
        init_flow_result = _run_runtime(
            "init-flow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
        )
        self.assertEqual(0, init_flow_result.returncode)

        result = _run_runtime(
            "record-step",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
            "--step-id",
            "confirm_tachometer_reference_half_flow",
            "--status",
            "passed",
            "--technician-name",
            "Alex Tech",
            "--note",
            "Attempting out-of-order transition",
        )
        self.assertEqual(2, result.returncode)

        flow_state_path = self.run_dir / "state" / "flows" / "FCU-01A.json"
        flow_state = json.loads(flow_state_path.read_text(encoding="utf-8"))
        step = [
            s for s in flow_state["steps"] if s["step_id"] == "confirm_tachometer_reference_half_flow"
        ][0]
        self.assertEqual("pending", step["status"])
        self.assertIn("history", step)
        self.assertGreaterEqual(len(step["history"]), 1)
        rejection_entry = step["history"][-1]
        self.assertTrue(rejection_entry["rejected"])
        self.assertEqual("pending", rejection_entry["previous_status"])
        self.assertEqual("passed", rejection_entry["attempted_status"])
        self.assertEqual("pending", rejection_entry["new_status"])
        self.assertEqual("PREREQ_ORDER", rejection_entry["reason_code"])
        self.assertEqual("PREREQ_ORDER", rejection_entry["rejection_reason_code"])

    def test_record_step_rejects_when_dependency_id_missing_from_flow(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-flow-missing-dependency",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)
        init_flow_result = _run_runtime(
            "init-flow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
        )
        self.assertEqual(0, init_flow_result.returncode)

        flow_state_path = self.run_dir / "state" / "flows" / "FCU-01A.json"
        flow_state = json.loads(flow_state_path.read_text(encoding="utf-8"))
        for step in flow_state["steps"]:
            if step["step_id"] == "half_design_airflow_auto":
                step["requires_step_ids"] = ["missing-step-id"]
                break
        flow_state_path.write_text(json.dumps(flow_state, indent=2), encoding="utf-8")

        result = _run_runtime(
            "record-step",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
            "--step-id",
            "half_design_airflow_auto",
            "--status",
            "passed",
            "--technician-name",
            "Alex Tech",
            "--note",
            "Attempting transition with missing dependency id",
        )
        self.assertEqual(2, result.returncode)
        self.assertIn("not present in flow", result.stdout)

        lines = (
            self.run_dir / "logs" / "events.jsonl"
        ).read_text(encoding="utf-8").strip().splitlines()
        parsed_events = [json.loads(line) for line in lines]
        rejection_events = [entry for entry in parsed_events if entry["event"] == "flow_step_rejected"]
        self.assertGreaterEqual(len(rejection_events), 1)
        self.assertEqual("DEPENDENCY_UNSATISFIED", rejection_events[-1]["reason_code"])

    def test_record_step_appends_history_for_multiple_updates(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-flow-history-multi",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)
        init_flow_result = _run_runtime(
            "init-flow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
        )
        self.assertEqual(0, init_flow_result.returncode)

        first_result = _run_runtime(
            "record-step",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
            "--step-id",
            "half_design_airflow_auto",
            "--status",
            "passed",
            "--technician-name",
            "Alex Tech",
            "--note",
            "First transition",
        )
        self.assertEqual(0, first_result.returncode)

        second_result = _run_runtime(
            "record-step",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
            "--step-id",
            "half_design_airflow_auto",
            "--status",
            "manual_passed",
            "--technician-name",
            "Alex Tech",
            "--note",
            "Second transition",
        )
        self.assertEqual(0, second_result.returncode)

        flow_state_path = self.run_dir / "state" / "flows" / "FCU-01A.json"
        flow_state = json.loads(flow_state_path.read_text(encoding="utf-8"))
        step = [s for s in flow_state["steps"] if s["step_id"] == "half_design_airflow_auto"][
            0
        ]
        self.assertEqual("manual_passed", step["status"])
        self.assertGreaterEqual(len(step["records"]), 2)
        self.assertGreaterEqual(len(step["history"]), 2)
        first_history = step["history"][-2]
        second_history = step["history"][-1]
        self.assertEqual("pending", first_history["previous_status"])
        self.assertEqual("passed", first_history["new_status"])
        self.assertEqual("passed", second_history["previous_status"])
        self.assertEqual("manual_passed", second_history["new_status"])
        self.assertEqual("status_update", first_history["reason_code"])
        self.assertEqual("status_update", second_history["reason_code"])

    def test_record_step_rejects_pending_status_record(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-flow-pending-record",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)
        init_flow_result = _run_runtime(
            "init-flow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
        )
        self.assertEqual(0, init_flow_result.returncode)

        result = _run_runtime(
            "record-step",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
            "--step-id",
            "half_design_airflow_auto",
            "--status",
            "pending",
            "--technician-name",
            "Alex Tech",
            "--note",
            "Trying to write pending",
        )
        self.assertEqual(2, result.returncode)
        self.assertIn("cannot record step with status 'pending'", result.stdout)

    def test_record_step_failed_requires_prior_steps_completed(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-flow-fail-order",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)
        init_flow_result = _run_runtime(
            "init-flow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
        )
        self.assertEqual(0, init_flow_result.returncode)

        result = _run_runtime(
            "record-step",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
            "--step-id",
            "confirm_tachometer_reference_half_flow",
            "--status",
            "failed",
            "--technician-name",
            "Alex Tech",
            "--note",
            "Out of order fail",
        )
        self.assertEqual(2, result.returncode)
        self.assertIn("invalid step transition", result.stdout)

        lines = (
            self.run_dir / "logs" / "events.jsonl"
        ).read_text(encoding="utf-8").strip().splitlines()
        parsed_events = [json.loads(line) for line in lines]
        rejection_events = [entry for entry in parsed_events if entry["event"] == "flow_step_rejected"]
        self.assertGreaterEqual(len(rejection_events), 1)
        self.assertEqual("PREREQ_ORDER", rejection_events[-1]["reason_code"])

    def test_record_step_failed_allowed_when_prerequisites_met(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-flow-fail-ok",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)
        init_flow_result = _run_runtime(
            "init-flow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
        )
        self.assertEqual(0, init_flow_result.returncode)

        first = _run_runtime(
            "record-step",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
            "--step-id",
            "half_design_airflow_auto",
            "--status",
            "passed",
            "--technician-name",
            "Alex Tech",
            "--note",
            "Airflow ok",
        )
        self.assertEqual(0, first.returncode)

        second = _run_runtime(
            "record-step",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
            "--step-id",
            "confirm_tachometer_reference_half_flow",
            "--status",
            "failed",
            "--technician-name",
            "Alex Tech",
            "--note",
            "Tachometer mismatch",
        )
        self.assertEqual(0, second.returncode)

        flow_state = json.loads(
            (self.run_dir / "state" / "flows" / "FCU-01A.json").read_text(encoding="utf-8")
        )
        step = [
            s for s in flow_state["steps"] if s["step_id"] == "confirm_tachometer_reference_half_flow"
        ][0]
        self.assertEqual("failed", step["status"])
        self.assertEqual("failed", step["history"][-1]["new_status"])

    def test_record_step_prior_failed_blocks_later_steps(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-flow-prior-fail-block",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)
        init_flow_result = _run_runtime(
            "init-flow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
        )
        self.assertEqual(0, init_flow_result.returncode)

        fail_first = _run_runtime(
            "record-step",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
            "--step-id",
            "half_design_airflow_auto",
            "--status",
            "failed",
            "--technician-name",
            "Alex Tech",
            "--note",
            "Could not reach half design",
        )
        self.assertEqual(0, fail_first.returncode)

        second = _run_runtime(
            "record-step",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-01A",
            "--step-id",
            "confirm_tachometer_reference_half_flow",
            "--status",
            "passed",
            "--technician-name",
            "Alex Tech",
            "--note",
            "Try after prior failed",
        )
        self.assertEqual(2, second.returncode)
        self.assertIn("before 'half_design_airflow_auto' is completed", second.stdout)

    def test_record_step_bacnet_point_checkout_gate_and_commissioning_report(self) -> None:
        from test_import_compiler import _write_profile

        server = _FakeBipUdpServer(device_instance=21001, analog_input_present=21.0, msv_present=1)
        server.start()
        try:
            time.sleep(0.05)
            self.run_dir.mkdir(parents=True, exist_ok=True)
            profiles_dir = self.run_dir / "profiles-local"
            profiles_dir.mkdir(parents=True, exist_ok=True)
            _write_profile(
                profiles_dir / "unit-profile-gate.json",
                profile_id="fcu_gate_v1",
                display_name="Gate test",
                read_allowlist=["ai_sat", "msv_test_mode"],
                point_checkout=[
                    {"object_id": "ai_sat", "property": "presentValue"},
                    {"object_id": "msv_test_mode", "property": "presentValue"},
                ],
                commissioning_flow=[
                    {
                        "step_id": "gate_point_checkout",
                        "label": "Gate",
                        "step_type": "bacnet_point_checkout",
                        "report_ref": "test.gate",
                    }
                ],
            )
            csv_path = self.run_dir / "controllers-gate.csv"
            with csv_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(
                    handle,
                    fieldnames=[
                        "controller_label",
                        "profile_id",
                        "bacnet_device_instance",
                        "bacnet_ip",
                        "bacnet_port",
                        "building_floor",
                        "notes",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "controller_label": "FCU-GATE",
                        "profile_id": "fcu_gate_v1",
                        "bacnet_device_instance": "21001",
                        "bacnet_ip": "127.0.0.1",
                        "bacnet_port": str(server.port),
                        "building_floor": "L01",
                        "notes": "gate",
                    }
                )

            init_result = _run_runtime(
                "init-run",
                "--run-dir",
                str(self.run_dir),
                "--job-id",
                "job-gate",
                "--controllers-csv",
                str(csv_path),
                "--profiles-dir",
                str(profiles_dir),
                "--scenarios-dir",
                str(ROOT / "docs" / "examples" / "simulator-scenarios"),
            )
            self.assertEqual(0, init_result.returncode)
            compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
            self.assertEqual(0, compile_result.returncode)

            init_flow = _run_runtime(
                "init-flow",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-GATE",
            )
            self.assertEqual(0, init_flow.returncode)

            rec = _run_runtime(
                "record-step",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-GATE",
                "--step-id",
                "gate_point_checkout",
                "--status",
                "passed",
                "--technician-name",
                "Alex Tech",
                "--note",
                "Gate with BACnet",
                "--bacnet-timeout-seconds",
                "0.5",
                "--bacnet-retries",
                "1",
            )
            self.assertEqual(0, rec.returncode)

            report_path = self.run_dir / "artifacts" / "commissioning_report.json"
            self.assertTrue(report_path.exists())
            report_doc = json.loads(report_path.read_text(encoding="utf-8"))
            self.assertEqual(1, len(report_doc.get("entries", [])))
            ent = report_doc["entries"][0]
            self.assertEqual("point_checkout_after_step", ent["kind"])
            self.assertEqual("test.gate", ent["report_ref"])
            self.assertTrue(ent["all_read_ok"])

            export_path = self.run_dir / "artifacts" / "commissioning-report-copy.json"
            ex = _run_runtime(
                "export-commissioning-report",
                "--run-dir",
                str(self.run_dir),
                "--output-json",
                str(export_path),
            )
            self.assertEqual(0, ex.returncode)
            self.assertTrue(export_path.exists())

            mod = _run_runtime(
                "append-commissioning-modulation-sample",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-GATE",
                "--read",
                "ai_sat",
                "--technician-name",
                "Alex Tech",
                "--note",
                "unified csv test",
                "--timeout-seconds",
                "0.5",
                "--retries",
                "1",
            )
            self.assertEqual(0, mod.returncode)
            unified_csv = self.run_dir / "artifacts" / "commissioning-unified.csv"
            uni = _run_runtime(
                "export-commissioning-report",
                "--run-dir",
                str(self.run_dir),
                "--output-csv-unified",
                str(unified_csv),
            )
            self.assertEqual(0, uni.returncode)
            text = unified_csv.read_text(encoding="utf-8")
            self.assertIn("point_checkout_after_step", text)
            self.assertIn("thermal_modulation_sample", text)
            self.assertIn("step_status", text)
            self.assertIn("ai_sat", text)

            html_path = self.run_dir / "artifacts" / "commissioning-unified.html"
            h = _run_runtime(
                "export-commissioning-report",
                "--run-dir",
                str(self.run_dir),
                "--output-html",
                str(html_path),
            )
            self.assertEqual(0, h.returncode)
            html_text = html_path.read_text(encoding="utf-8")
            self.assertIn("<!DOCTYPE html>", html_text)
            self.assertIn("<table>", html_text)
            self.assertIn("job-gate", html_text)
            self.assertIn("point_checkout_after_step", html_text)
            self.assertIn("thermal_modulation_sample", html_text)
            self.assertIn("Print to PDF", html_text)

            xlsx_path = self.run_dir / "artifacts" / "commissioning-unified.xlsx"
            x = _run_runtime(
                "export-commissioning-report",
                "--run-dir",
                str(self.run_dir),
                "--output-xlsx",
                str(xlsx_path),
            )
            self.assertEqual(0, x.returncode)
            self.assertTrue(xlsx_path.is_file())
            with zipfile.ZipFile(xlsx_path, "r") as zf:
                names = zf.namelist()
            self.assertTrue(any(n.endswith("xl/worksheets/sheet1.xml") for n in names))

            pdf_path = self.run_dir / "artifacts" / "commissioning-unified.pdf"
            p = _run_runtime(
                "export-commissioning-report",
                "--run-dir",
                str(self.run_dir),
                "--output-pdf",
                str(pdf_path),
            )
            self.assertEqual(0, p.returncode)
            self.assertTrue(pdf_path.is_file())
            self.assertTrue(pdf_path.read_bytes().startswith(b"%PDF"))
        finally:
            server.stop()

    def test_export_commissioning_report_errors_when_missing(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-cr-missing",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        r = _run_runtime("export-commissioning-report", "--run-dir", str(self.run_dir))
        self.assertEqual(2, r.returncode)
        self.assertIn("commissioning report not found", r.stdout)

    def test_export_commissioning_report_allow_empty_stub(self) -> None:
        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-cr-stub",
            "--controllers-csv",
            str(ROOT / "docs" / "examples" / "site-controllers.template.csv"),
            "--profiles-dir",
            str(ROOT / "docs" / "examples"),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        out = self.run_dir / "artifacts" / "cr-stub.json"
        r = _run_runtime(
            "export-commissioning-report",
            "--run-dir",
            str(self.run_dir),
            "--output-json",
            str(out),
            "--allow-empty",
        )
        self.assertEqual(0, r.returncode)
        self.assertTrue(out.exists())
        doc = json.loads(out.read_text(encoding="utf-8"))
        self.assertEqual([], doc.get("entries", []))
        self.assertEqual("job-cr-stub", doc.get("job_id"))

        html_only = self.run_dir / "artifacts" / "empty-report.html"
        r2 = _run_runtime(
            "export-commissioning-report",
            "--run-dir",
            str(self.run_dir),
            "--allow-empty",
            "--output-html",
            str(html_only),
        )
        self.assertEqual(0, r2.returncode)
        self.assertIn("(no entries)", html_only.read_text(encoding="utf-8"))

        x_empty = self.run_dir / "artifacts" / "empty-report.xlsx"
        r3 = _run_runtime(
            "export-commissioning-report",
            "--run-dir",
            str(self.run_dir),
            "--allow-empty",
            "--output-xlsx",
            str(x_empty),
        )
        self.assertEqual(0, r3.returncode)
        from openpyxl import load_workbook

        wb = load_workbook(x_empty)
        ws = wb.active
        self.assertEqual(1, ws.max_row)
        self.assertEqual("entry_ts", ws.cell(row=1, column=1).value)

        pdf_empty = self.run_dir / "artifacts" / "empty-report.pdf"
        r4 = _run_runtime(
            "export-commissioning-report",
            "--run-dir",
            str(self.run_dir),
            "--allow-empty",
            "--output-pdf",
            str(pdf_empty),
        )
        self.assertEqual(0, r4.returncode)
        self.assertTrue(pdf_empty.read_bytes().startswith(b"%PDF"))

    def test_record_step_point_checkout_failure_leaves_step_pending(self) -> None:
        from test_import_compiler import _write_profile

        self.run_dir.mkdir(parents=True, exist_ok=True)
        profiles_dir = self.run_dir / "profiles-fail"
        profiles_dir.mkdir(parents=True, exist_ok=True)
        _write_profile(
            profiles_dir / "unit-profile-fail.json",
            profile_id="fcu_fail_v1",
            display_name="Fail test",
            read_allowlist=["ai_sat", "msv_test_mode"],
            point_checkout=[
                {"object_id": "ai_sat", "property": "presentValue"},
            ],
            commissioning_flow=[
                {
                    "step_id": "gate_fail",
                    "label": "Gate",
                    "step_type": "bacnet_point_checkout",
                }
            ],
        )
        csv_path = self.run_dir / "controllers-fail.csv"
        with csv_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=[
                    "controller_label",
                    "profile_id",
                    "bacnet_device_instance",
                    "bacnet_ip",
                    "bacnet_port",
                    "building_floor",
                    "notes",
                ],
            )
            writer.writeheader()
            writer.writerow(
                {
                    "controller_label": "FCU-FAIL",
                    "profile_id": "fcu_fail_v1",
                    "bacnet_device_instance": "21001",
                    "bacnet_ip": "127.0.0.1",
                    "bacnet_port": "1",
                    "building_floor": "L01",
                    "notes": "no listener",
                }
            )

        init_result = _run_runtime(
            "init-run",
            "--run-dir",
            str(self.run_dir),
            "--job-id",
            "job-fail",
            "--controllers-csv",
            str(csv_path),
            "--profiles-dir",
            str(profiles_dir),
            "--scenarios-dir",
            str(ROOT / "docs" / "examples" / "simulator-scenarios"),
        )
        self.assertEqual(0, init_result.returncode)
        compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
        self.assertEqual(0, compile_result.returncode)
        init_flow = _run_runtime(
            "init-flow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-FAIL",
        )
        self.assertEqual(0, init_flow.returncode)

        rec = _run_runtime(
            "record-step",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-FAIL",
            "--step-id",
            "gate_fail",
            "--status",
            "passed",
            "--technician-name",
            "Alex Tech",
            "--note",
            "should fail BACnet",
            "--bacnet-timeout-seconds",
            "0.05",
            "--bacnet-retries",
            "1",
        )
        self.assertEqual(2, rec.returncode)
        self.assertIn("BACnet point checkout failed", rec.stdout)

        flow_path = self.run_dir / "state" / "flows" / "FCU-FAIL.json"
        flow_state = json.loads(flow_path.read_text(encoding="utf-8"))
        step = [s for s in flow_state["steps"] if s["step_id"] == "gate_fail"][0]
        self.assertEqual("pending", step["status"])
        report_path = self.run_dir / "artifacts" / "commissioning_report.json"
        self.assertFalse(report_path.exists())

    def test_record_step_run_point_checkout_on_pass_after_prior_pass(self) -> None:
        from test_import_compiler import _write_profile

        server = _FakeBipUdpServer(device_instance=21001, analog_input_present=19.0, msv_present=2)
        server.start()
        try:
            time.sleep(0.05)
            self.run_dir.mkdir(parents=True, exist_ok=True)
            profiles_dir = self.run_dir / "profiles-chain"
            profiles_dir.mkdir(parents=True, exist_ok=True)
            _write_profile(
                profiles_dir / "unit-profile-chain.json",
                profile_id="fcu_chain_v1",
                display_name="Chain",
                read_allowlist=["ai_sat", "msv_test_mode"],
                point_checkout=[
                    {"object_id": "ai_sat", "property": "presentValue"},
                    {"object_id": "msv_test_mode", "property": "presentValue"},
                ],
                commissioning_flow=[
                    {"step_id": "prep", "label": "Prep"},
                    {
                        "step_id": "readout",
                        "label": "Readout",
                        "run_point_checkout_on_pass": True,
                        "report_ref": "chain.after_prep",
                    },
                ],
            )
            csv_path = self.run_dir / "controllers-chain.csv"
            with csv_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(
                    handle,
                    fieldnames=[
                        "controller_label",
                        "profile_id",
                        "bacnet_device_instance",
                        "bacnet_ip",
                        "bacnet_port",
                        "building_floor",
                        "notes",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "controller_label": "FCU-CHAIN",
                        "profile_id": "fcu_chain_v1",
                        "bacnet_device_instance": "21001",
                        "bacnet_ip": "127.0.0.1",
                        "bacnet_port": str(server.port),
                        "building_floor": "L01",
                        "notes": "",
                    }
                )

            init_result = _run_runtime(
                "init-run",
                "--run-dir",
                str(self.run_dir),
                "--job-id",
                "job-chain",
                "--controllers-csv",
                str(csv_path),
                "--profiles-dir",
                str(profiles_dir),
                "--scenarios-dir",
                str(ROOT / "docs" / "examples" / "simulator-scenarios"),
            )
            self.assertEqual(0, init_result.returncode)
            compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
            self.assertEqual(0, compile_result.returncode)
            self.assertEqual(
                0,
                _run_runtime(
                    "init-flow",
                    "--run-dir",
                    str(self.run_dir),
                    "--controller-label",
                    "FCU-CHAIN",
                ).returncode,
            )
            self.assertEqual(
                0,
                _run_runtime(
                    "record-step",
                    "--run-dir",
                    str(self.run_dir),
                    "--controller-label",
                    "FCU-CHAIN",
                    "--step-id",
                    "prep",
                    "--status",
                    "passed",
                    "--technician-name",
                    "Alex Tech",
                    "--note",
                    "prep ok",
                ).returncode,
            )
            self.assertEqual(
                0,
                _run_runtime(
                    "record-step",
                    "--run-dir",
                    str(self.run_dir),
                    "--controller-label",
                    "FCU-CHAIN",
                    "--step-id",
                    "readout",
                    "--status",
                    "passed",
                    "--technician-name",
                    "Alex Tech",
                    "--note",
                    "with checkout",
                    "--bacnet-timeout-seconds",
                    "0.5",
                    "--bacnet-retries",
                    "1",
                ).returncode,
            )
            report_path = self.run_dir / "artifacts" / "commissioning_report.json"
            doc = json.loads(report_path.read_text(encoding="utf-8"))
            self.assertEqual(1, len(doc["entries"]))
            self.assertEqual("chain.after_prep", doc["entries"][0]["report_ref"])
        finally:
            server.stop()

    def test_append_modulation_sample_and_export_modulation_csv(self) -> None:
        server = _FakeBipUdpServer(device_instance=21001, analog_input_present=20.0, msv_present=3)
        server.start()
        try:
            time.sleep(0.05)
            self.run_dir.mkdir(parents=True, exist_ok=True)
            csv_path = self.run_dir / "controllers-local.csv"
            with csv_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(
                    handle,
                    fieldnames=[
                        "controller_label",
                        "profile_id",
                        "bacnet_device_instance",
                        "bacnet_ip",
                        "bacnet_port",
                        "building_floor",
                        "notes",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "controller_label": "FCU-MOD",
                        "profile_id": "fcu_2pipe_chw_electric_heat_v1",
                        "bacnet_device_instance": "21001",
                        "bacnet_ip": "127.0.0.1",
                        "bacnet_port": str(server.port),
                        "building_floor": "L01",
                        "notes": "mod",
                    }
                )
            init_result = _run_runtime(
                "init-run",
                "--run-dir",
                str(self.run_dir),
                "--job-id",
                "job-mod",
                "--controllers-csv",
                str(csv_path),
                "--profiles-dir",
                str(ROOT / "docs" / "examples"),
                "--scenarios-dir",
                str(ROOT / "docs" / "examples" / "simulator-scenarios"),
            )
            self.assertEqual(0, init_result.returncode)
            compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
            self.assertEqual(0, compile_result.returncode)

            app = _run_runtime(
                "append-commissioning-modulation-sample",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-MOD",
                "--read",
                "ai_sat",
                "--read",
                "msv_test_mode",
                "--technician-name",
                "Alex Tech",
                "--note",
                "sweep t0",
                "--step-id",
                "heating_test",
                "--report-ref",
                "thermal_tests_for_report.heating",
                "--timeout-seconds",
                "0.5",
                "--retries",
                "1",
            )
            self.assertEqual(0, app.returncode)
            report_path = self.run_dir / "artifacts" / "commissioning_report.json"
            doc = json.loads(report_path.read_text(encoding="utf-8"))
            self.assertEqual("0.2-commissioning-report", doc.get("schema_version"))
            kinds = [e.get("kind") for e in doc.get("entries", [])]
            self.assertIn("thermal_modulation_sample", kinds)

            csv_out = self.run_dir / "artifacts" / "modulation.csv"
            ex = _run_runtime(
                "export-commissioning-report",
                "--run-dir",
                str(self.run_dir),
                "--output-csv",
                str(csv_out),
            )
            self.assertEqual(0, ex.returncode)
            text = csv_out.read_text(encoding="utf-8")
            self.assertIn("ai_sat", text)
            self.assertIn("msv_test_mode", text)
            self.assertIn("thermal_modulation_sample", text)
        finally:
            server.stop()

    def test_append_modulation_batch_from_json(self) -> None:
        server = _FakeBipUdpServer(device_instance=21001, analog_input_present=11.0, msv_present=4)
        server.start()
        try:
            time.sleep(0.05)
            self.run_dir.mkdir(parents=True, exist_ok=True)
            profiles_dir = self.run_dir / "profiles-batch"
            profiles_dir.mkdir(parents=True, exist_ok=True)
            (profiles_dir / "unit-batch.json").write_text(
                json.dumps(
                    {
                        "schema_version": "0.1-example",
                        "profile_id": "fcu_batch_v1",
                        "display_name": "Batch",
                        "commissioning_read_allowlist": ["ai_sat"],
                        "objects": [
                            {
                                "id": "ai_sat",
                                "bacnet": {"object_type": "analogInput", "instance": 2},
                                "writable": False,
                            }
                        ],
                        "commissioning_flow": [{"step_id": "s1", "label": "S"}],
                    },
                    indent=2,
                ),
                encoding="utf-8",
            )
            csv_path = self.run_dir / "controllers-batch.csv"
            with csv_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(
                    handle,
                    fieldnames=[
                        "controller_label",
                        "profile_id",
                        "bacnet_device_instance",
                        "bacnet_ip",
                        "bacnet_port",
                        "building_floor",
                        "notes",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "controller_label": "FCU-BATCH",
                        "profile_id": "fcu_batch_v1",
                        "bacnet_device_instance": "21001",
                        "bacnet_ip": "127.0.0.1",
                        "bacnet_port": str(server.port),
                        "building_floor": "L01",
                        "notes": "",
                    }
                )
            init_result = _run_runtime(
                "init-run",
                "--run-dir",
                str(self.run_dir),
                "--job-id",
                "job-batch-mod",
                "--controllers-csv",
                str(csv_path),
                "--profiles-dir",
                str(profiles_dir),
                "--scenarios-dir",
                str(ROOT / "docs" / "examples" / "simulator-scenarios"),
            )
            self.assertEqual(0, init_result.returncode)
            compile_result = _run_runtime("compile-import", "--run-dir", str(self.run_dir))
            self.assertEqual(0, compile_result.returncode)

            batch_file = self.run_dir / "batch-samples.json"
            batch_file.write_text(
                json.dumps(
                    [
                        {
                            "controller_label": "FCU-BATCH",
                            "reads": ["ai_sat"],
                            "technician_name": "Batch Tech",
                            "report_ref": "batch.t1",
                        }
                    ],
                    indent=2,
                ),
                encoding="utf-8",
            )
            r = _run_runtime(
                "append-commissioning-modulation-batch",
                "--run-dir",
                str(self.run_dir),
                "--input-json",
                str(batch_file),
                "--timeout-seconds",
                "0.5",
                "--retries",
                "1",
            )
            self.assertEqual(0, r.returncode)
            doc = json.loads(
                (self.run_dir / "artifacts" / "commissioning_report.json").read_text(encoding="utf-8")
            )
            self.assertTrue(any(e.get("kind") == "thermal_modulation_batch" for e in doc["entries"]))
        finally:
            server.stop()

    def test_bacnet_modulation_sweep_heating_step(self) -> None:
        server = _FakeBipUdpServer(
            device_instance=21001,
            analog_input_present=23.5,
            msv_present=1,
            av_heat_present=10.0,
        )
        server.start()
        try:
            time.sleep(0.05)
            self.run_dir.mkdir(parents=True, exist_ok=True)
            profiles_dir = self.run_dir / "profiles-sweep"
            profiles_dir.mkdir(parents=True, exist_ok=True)
            (profiles_dir / "unit-sweep.json").write_text(
                json.dumps(
                    {
                        "schema_version": "0.1-example",
                        "profile_id": "fcu_sweep_v1",
                        "display_name": "Sweep",
                        "commissioning_write_allowlist": ["av_electric_heat_command"],
                        "commissioning_read_allowlist": ["ai_sat", "av_electric_heat_command"],
                        "objects": [
                            {
                                "id": "ai_sat",
                                "bacnet": {"object_type": "analogInput", "instance": 2},
                                "writable": False,
                            },
                            {
                                "id": "av_electric_heat_command",
                                "bacnet": {"object_type": "analogValue", "instance": 4},
                                "writable": True,
                            },
                        ],
                        "commissioning_flow": [
                            {
                                "step_id": "heating_test",
                                "label": "Heating",
                                "report_ref": "thermal_tests_for_report.heating",
                                "actions": [
                                    {
                                        "type": "modulate_actuator_log_sat_for_report",
                                        "command_object_id": "av_electric_heat_command",
                                        "result_supply_temperature_object_id": "ai_sat",
                                        "optional_context_object_ids": ["av_electric_heat_command"],
                                    }
                                ],
                            }
                        ],
                    },
                    indent=2,
                ),
                encoding="utf-8",
            )
            csv_path = self.run_dir / "controllers-sweep.csv"
            with csv_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(
                    handle,
                    fieldnames=[
                        "controller_label",
                        "profile_id",
                        "bacnet_device_instance",
                        "bacnet_ip",
                        "bacnet_port",
                        "building_floor",
                        "notes",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "controller_label": "FCU-SWEEP",
                        "profile_id": "fcu_sweep_v1",
                        "bacnet_device_instance": "21001",
                        "bacnet_ip": "127.0.0.1",
                        "bacnet_port": str(server.port),
                        "building_floor": "L01",
                        "notes": "",
                    }
                )
            self.assertEqual(
                0,
                _run_runtime(
                    "init-run",
                    "--run-dir",
                    str(self.run_dir),
                    "--job-id",
                    "job-sweep",
                    "--controllers-csv",
                    str(csv_path),
                    "--profiles-dir",
                    str(profiles_dir),
                    "--scenarios-dir",
                    str(ROOT / "docs" / "examples" / "simulator-scenarios"),
                ).returncode,
            )
            self.assertEqual(0, _run_runtime("compile-import", "--run-dir", str(self.run_dir)).returncode)
            self.assertEqual(
                0,
                _run_runtime(
                    "init-flow",
                    "--run-dir",
                    str(self.run_dir),
                    "--controller-label",
                    "FCU-SWEEP",
                ).returncode,
            )
            sw = _run_runtime(
                "bacnet-modulation-sweep",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-SWEEP",
                "--step-id",
                "heating_test",
                "--command-percent",
                "55",
                "--dwell-seconds",
                "0.05",
                "--technician-name",
                "Alex Tech",
                "--note",
                "sweep test",
                "--timeout-seconds",
                "0.5",
                "--retries",
                "1",
            )
            self.assertEqual(0, sw.returncode)
            doc = json.loads(
                (self.run_dir / "artifacts" / "commissioning_report.json").read_text(encoding="utf-8")
            )
            kinds = [e.get("kind") for e in doc.get("entries", [])]
            self.assertIn("thermal_modulation_sweep", kinds)
            sweep = [e for e in doc["entries"] if e.get("kind") == "thermal_modulation_sweep"][0]
            self.assertEqual(55.0, sweep["command_percent"])
            ids = {r["logical_object_id"] for r in sweep["readings"]}
            self.assertEqual({"ai_sat", "av_electric_heat_command"}, ids)

            csv_out = self.run_dir / "artifacts" / "sweep.csv"
            self.assertEqual(
                0,
                _run_runtime(
                    "export-commissioning-report",
                    "--run-dir",
                    str(self.run_dir),
                    "--output-csv",
                    str(csv_out),
                ).returncode,
            )
            csv_text = csv_out.read_text(encoding="utf-8")
            self.assertIn("thermal_modulation_sweep", csv_text)
            self.assertIn("av_electric_heat_command", csv_text)
        finally:
            server.stop()

    def test_bacnet_modulation_sweep_multi_percent_session_rat(self) -> None:
        server = _FakeBipUdpServer(
            device_instance=21001,
            analog_input_present=23.5,
            msv_present=1,
            av_heat_present=10.0,
        )
        server.start()
        try:
            time.sleep(0.05)
            self.run_dir.mkdir(parents=True, exist_ok=True)
            profiles_dir = self.run_dir / "profiles-sweep-rat"
            profiles_dir.mkdir(parents=True, exist_ok=True)
            (profiles_dir / "unit-sweep-rat.json").write_text(
                json.dumps(
                    {
                        "schema_version": "0.1-example",
                        "profile_id": "fcu_sweep_rat_v1",
                        "display_name": "Sweep RAT session",
                        "commissioning_write_allowlist": ["av_electric_heat_command"],
                        "commissioning_read_allowlist": [
                            "ai_sat",
                            "av_electric_heat_command",
                        ],
                        "objects": [
                            {
                                "id": "ai_sat",
                                "bacnet": {"object_type": "analogInput", "instance": 2},
                                "writable": False,
                            },
                            {
                                "id": "av_electric_heat_command",
                                "bacnet": {"object_type": "analogValue", "instance": 4},
                                "writable": True,
                            },
                        ],
                        "commissioning_flow": [
                            {
                                "step_id": "heating_test",
                                "label": "Heating",
                                "report_ref": "thermal_tests_for_report.heating",
                                "actions": [
                                    {
                                        "type": "modulate_actuator_log_sat_for_report",
                                        "command_object_id": "av_electric_heat_command",
                                        "result_supply_temperature_object_id": "ai_sat",
                                        "result_return_temperature_object_id": "ai_rat_missing",
                                        "session_return_air_temperature_key": "rat_degC",
                                    }
                                ],
                            }
                        ],
                    },
                    indent=2,
                ),
                encoding="utf-8",
            )
            csv_path = self.run_dir / "controllers-sweep-rat.csv"
            with csv_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(
                    handle,
                    fieldnames=[
                        "controller_label",
                        "profile_id",
                        "bacnet_device_instance",
                        "bacnet_ip",
                        "bacnet_port",
                        "building_floor",
                        "notes",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "controller_label": "FCU-RAT",
                        "profile_id": "fcu_sweep_rat_v1",
                        "bacnet_device_instance": "21001",
                        "bacnet_ip": "127.0.0.1",
                        "bacnet_port": str(server.port),
                        "building_floor": "L01",
                        "notes": "",
                    }
                )
            self.assertEqual(
                0,
                _run_runtime(
                    "init-run",
                    "--run-dir",
                    str(self.run_dir),
                    "--job-id",
                    "job-sweep-rat",
                    "--controllers-csv",
                    str(csv_path),
                    "--profiles-dir",
                    str(profiles_dir),
                    "--scenarios-dir",
                    str(ROOT / "docs" / "examples" / "simulator-scenarios"),
                ).returncode,
            )
            self.assertEqual(0, _run_runtime("compile-import", "--run-dir", str(self.run_dir)).returncode)
            self.assertEqual(
                0,
                _run_runtime(
                    "init-flow",
                    "--run-dir",
                    str(self.run_dir),
                    "--controller-label",
                    "FCU-RAT",
                ).returncode,
            )
            self.assertEqual(
                0,
                _run_runtime(
                    "set-session-value",
                    "--run-dir",
                    str(self.run_dir),
                    "--controller-label",
                    "FCU-RAT",
                    "--key",
                    "rat_degC",
                    "--value",
                    "21.25",
                    "--technician-name",
                    "Alex Tech",
                    "--note",
                    "manual rat",
                ).returncode,
            )
            sw = _run_runtime(
                "bacnet-modulation-sweep",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-RAT",
                "--step-id",
                "heating_test",
                "--command-percents",
                "12,24",
                "--dwell-seconds",
                "0.05",
                "--technician-name",
                "Alex Tech",
                "--note",
                "multi",
                "--timeout-seconds",
                "0.5",
                "--retries",
                "1",
            )
            self.assertEqual(0, sw.returncode)
            doc = json.loads(
                (self.run_dir / "artifacts" / "commissioning_report.json").read_text(encoding="utf-8")
            )
            sweeps = [e for e in doc["entries"] if e.get("kind") == "thermal_modulation_sweep"]
            self.assertEqual(2, len(sweeps))
            self.assertEqual(12.0, sweeps[0]["command_percent"])
            self.assertEqual(24.0, sweeps[1]["command_percent"])
            rat_rows = [r for r in sweeps[0]["readings"] if r.get("logical_object_id") == "rat_degC"]
            self.assertEqual(1, len(rat_rows))
            self.assertEqual("read_ok", rat_rows[0]["status"])
            self.assertEqual("21.25", rat_rows[0]["value_str"])
            self.assertEqual("session", rat_rows[0].get("source"))
        finally:
            server.stop()

    def test_record_step_modulation_on_pass_requires_percents(self) -> None:
        from test_import_compiler import _write_profile

        server = _FakeBipUdpServer(
            device_instance=21001,
            analog_input_present=20.0,
            msv_present=1,
            av_heat_present=5.0,
        )
        server.start()
        try:
            time.sleep(0.05)
            self.run_dir.mkdir(parents=True, exist_ok=True)
            profiles_dir = self.run_dir / "profiles-rec-mod"
            profiles_dir.mkdir(parents=True, exist_ok=True)
            _write_profile(
                profiles_dir / "unit-rec-mod.json",
                profile_id="fcu_rec_mod_v1",
                display_name="Rec mod",
                write_allowlist=["av_electric_heat_command"],
                read_allowlist=["ai_sat", "av_electric_heat_command"],
                objects=[
                    {
                        "id": "msv_test_mode",
                        "writable": True,
                        "bacnet": {"object_type": "multiStateValue", "instance": 50},
                    },
                    {
                        "id": "ai_sat",
                        "writable": False,
                        "bacnet": {"object_type": "analogInput", "instance": 2},
                    },
                    {
                        "id": "av_electric_heat_command",
                        "writable": True,
                        "bacnet": {"object_type": "analogValue", "instance": 4},
                    },
                ],
                commissioning_flow=[
                    {"step_id": "prep", "label": "Prep"},
                    {
                        "step_id": "heat",
                        "label": "Heat",
                        "actions": [
                            {
                                "type": "modulate_actuator_log_sat_for_report",
                                "command_object_id": "av_electric_heat_command",
                                "result_supply_temperature_object_id": "ai_sat",
                            }
                        ],
                    },
                ],
            )
            csv_path = self.run_dir / "controllers-rec-mod.csv"
            with csv_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(
                    handle,
                    fieldnames=[
                        "controller_label",
                        "profile_id",
                        "bacnet_device_instance",
                        "bacnet_ip",
                        "bacnet_port",
                        "building_floor",
                        "notes",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "controller_label": "FCU-REC",
                        "profile_id": "fcu_rec_mod_v1",
                        "bacnet_device_instance": "21001",
                        "bacnet_ip": "127.0.0.1",
                        "bacnet_port": str(server.port),
                        "building_floor": "L01",
                        "notes": "",
                    }
                )
            self.assertEqual(
                0,
                _run_runtime(
                    "init-run",
                    "--run-dir",
                    str(self.run_dir),
                    "--job-id",
                    "job-rec-mod",
                    "--controllers-csv",
                    str(csv_path),
                    "--profiles-dir",
                    str(profiles_dir),
                    "--scenarios-dir",
                    str(ROOT / "docs" / "examples" / "simulator-scenarios"),
                ).returncode,
            )
            self.assertEqual(0, _run_runtime("compile-import", "--run-dir", str(self.run_dir)).returncode)
            self.assertEqual(
                0,
                _run_runtime(
                    "init-flow",
                    "--run-dir",
                    str(self.run_dir),
                    "--controller-label",
                    "FCU-REC",
                ).returncode,
            )
            self.assertEqual(
                0,
                _run_runtime(
                    "record-step",
                    "--run-dir",
                    str(self.run_dir),
                    "--controller-label",
                    "FCU-REC",
                    "--step-id",
                    "prep",
                    "--status",
                    "passed",
                    "--technician-name",
                    "Alex Tech",
                    "--note",
                    "prep",
                ).returncode,
            )
            bad = _run_runtime(
                "record-step",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-REC",
                "--step-id",
                "heat",
                "--status",
                "passed",
                "--technician-name",
                "Alex Tech",
                "--note",
                "no percents",
            )
            self.assertEqual(2, bad.returncode)
            self.assertIn("--modulation-command-percents", bad.stdout)
            ok = _run_runtime(
                "record-step",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-REC",
                "--step-id",
                "heat",
                "--status",
                "passed",
                "--technician-name",
                "Alex Tech",
                "--note",
                "with percents",
                "--modulation-command-percents",
                "15,30",
                "--modulation-dwell-seconds",
                "0.05",
                "--bacnet-timeout-seconds",
                "0.5",
                "--bacnet-retries",
                "1",
            )
            self.assertEqual(0, ok.returncode)
            doc = json.loads(
                (self.run_dir / "artifacts" / "commissioning_report.json").read_text(encoding="utf-8")
            )
            sweeps = [e for e in doc["entries"] if e.get("kind") == "thermal_modulation_sweep"]
            self.assertEqual(2, len(sweeps))
            self.assertEqual("record_step", sweeps[0].get("trigger"))
            flow_path = self.run_dir / "state" / "flows" / "FCU-REC.json"
            heat = [s for s in json.loads(flow_path.read_text(encoding="utf-8"))["steps"] if s["step_id"] == "heat"][0]
            self.assertIn("last_modulation_sweep", heat)
            self.assertEqual(2, heat["last_modulation_sweep"]["sweep_steps"])
        finally:
            server.stop()

    def test_record_step_skip_when_requires_session_flag(self) -> None:
        from test_import_compiler import _write_profile

        self.run_dir.mkdir(parents=True, exist_ok=True)
        profiles_dir = self.run_dir / "profiles-skip-gate"
        profiles_dir.mkdir(parents=True, exist_ok=True)
        _write_profile(
            profiles_dir / "unit-skip-gate.json",
            profile_id="fcu_skip_gate_v1",
            display_name="Skip gate",
            read_allowlist=["ai_sat"],
            commissioning_flow=[
                {"step_id": "prep", "label": "Prep"},
                {
                    "step_id": "cool_plant",
                    "label": "Cooling plant test",
                    "skippable": True,
                    "skip_when": ["chilled_water_not_ready", "plant_not_commissioned"],
                },
            ],
        )
        csv_path = self.run_dir / "controllers-skip-gate.csv"
        with csv_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=[
                    "controller_label",
                    "profile_id",
                    "bacnet_device_instance",
                    "bacnet_ip",
                    "bacnet_port",
                    "building_floor",
                    "notes",
                ],
            )
            writer.writeheader()
            writer.writerow(
                {
                    "controller_label": "FCU-SKIP",
                    "profile_id": "fcu_skip_gate_v1",
                    "bacnet_device_instance": "21001",
                    "bacnet_ip": "192.168.1.50",
                    "bacnet_port": "47808",
                    "building_floor": "L01",
                    "notes": "",
                }
            )
        self.assertEqual(
            0,
            _run_runtime(
                "init-run",
                "--run-dir",
                str(self.run_dir),
                "--job-id",
                "job-skip-gate",
                "--controllers-csv",
                str(csv_path),
                "--profiles-dir",
                str(profiles_dir),
                "--scenarios-dir",
                str(ROOT / "docs" / "examples" / "simulator-scenarios"),
            ).returncode,
        )
        self.assertEqual(0, _run_runtime("compile-import", "--run-dir", str(self.run_dir)).returncode)
        self.assertEqual(
            0,
            _run_runtime(
                "init-flow",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-SKIP",
            ).returncode,
        )
        self.assertEqual(
            0,
            _run_runtime(
                "record-step",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-SKIP",
                "--step-id",
                "prep",
                "--status",
                "passed",
                "--technician-name",
                "Alex Tech",
                "--note",
                "prep",
            ).returncode,
        )
        blocked = _run_runtime(
            "record-step",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-SKIP",
            "--step-id",
            "cool_plant",
            "--status",
            "skipped",
            "--technician-name",
            "Alex Tech",
            "--note",
            "no chw",
        )
        self.assertEqual(2, blocked.returncode)
        self.assertIn("chilled_water_not_ready", blocked.stdout)
        events = (self.run_dir / "logs" / "events.jsonl").read_text(encoding="utf-8").strip().splitlines()
        rej = [json.loads(line) for line in events if json.loads(line).get("event") == "flow_step_rejected"]
        self.assertEqual("SKIP_GATE", rej[-1]["reason_code"])
        self.assertEqual(
            0,
            _run_runtime(
                "set-session-value",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-SKIP",
                "--key",
                "chilled_water_not_ready",
                "--value",
                "true",
                "--technician-name",
                "Alex Tech",
                "--note",
                "plant down",
            ).returncode,
        )
        allowed = _run_runtime(
            "record-step",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-SKIP",
            "--step-id",
            "cool_plant",
            "--status",
            "skipped",
            "--technician-name",
            "Alex Tech",
            "--note",
            "skip with reason",
        )
        self.assertEqual(0, allowed.returncode)

    def test_chw_valve_stroke_confirm_prompt_then_record_step_pass(self) -> None:
        from test_import_compiler import _write_profile

        server = _FakeBipUdpServer(
            device_instance=21001,
            analog_input_present=20.0,
            msv_present=6,
            ao_valve_present=0.0,
        )
        server.start()
        try:
            time.sleep(0.05)
            self.run_dir.mkdir(parents=True, exist_ok=True)
            profiles_dir = self.run_dir / "profiles-chw-stroke"
            profiles_dir.mkdir(parents=True, exist_ok=True)
            _write_profile(
                profiles_dir / "unit-chw-stroke.json",
                profile_id="fcu_chw_stroke_v1",
                display_name="CHW stroke",
                write_allowlist=["msv_test_mode", "ao_chw_valve"],
                read_allowlist=["ai_sat", "msv_test_mode", "ao_chw_valve"],
                objects=[
                    {
                        "id": "msv_test_mode",
                        "writable": True,
                        "bacnet": {"object_type": "multiStateValue", "instance": 50},
                    },
                    {
                        "id": "ai_sat",
                        "writable": False,
                        "bacnet": {"object_type": "analogInput", "instance": 2},
                    },
                    {
                        "id": "ao_chw_valve",
                        "writable": True,
                        "bacnet": {"object_type": "analogOutput", "instance": 5},
                    },
                ],
                commissioning_flow=[
                    {"step_id": "prep", "label": "Prep"},
                    {
                        "step_id": "cooling_valve_stroke_no_chw",
                        "label": "Stroke no CHW",
                        "arms_test_mode_state_key": "chw_valve_stroke_no_plant",
                        "actions": [
                            {
                                "type": "write_analog_percent",
                                "object_id": "ao_chw_valve",
                                "value": 100,
                            },
                            {
                                "type": "operator_prompt_confirm",
                                "prompt_id": "chw_valve_at_100",
                                "prompt_text": "Confirm 100%",
                            },
                            {
                                "type": "write_analog_percent",
                                "object_id": "ao_chw_valve",
                                "value": 0,
                            },
                            {
                                "type": "operator_prompt_confirm",
                                "prompt_id": "chw_valve_at_0",
                                "prompt_text": "Confirm 0%",
                            },
                        ],
                    },
                ],
            )
            csv_path = self.run_dir / "controllers-chw-stroke.csv"
            with csv_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(
                    handle,
                    fieldnames=[
                        "controller_label",
                        "profile_id",
                        "bacnet_device_instance",
                        "bacnet_ip",
                        "bacnet_port",
                        "building_floor",
                        "notes",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "controller_label": "FCU-STROKE",
                        "profile_id": "fcu_chw_stroke_v1",
                        "bacnet_device_instance": "21001",
                        "bacnet_ip": "127.0.0.1",
                        "bacnet_port": str(server.port),
                        "building_floor": "L01",
                        "notes": "",
                    }
                )
            self.assertEqual(
                0,
                _run_runtime(
                    "init-run",
                    "--run-dir",
                    str(self.run_dir),
                    "--job-id",
                    "job-chw-stroke",
                    "--controllers-csv",
                    str(csv_path),
                    "--profiles-dir",
                    str(profiles_dir),
                    "--scenarios-dir",
                    str(ROOT / "docs" / "examples" / "simulator-scenarios"),
                ).returncode,
            )
            self.assertEqual(0, _run_runtime("compile-import", "--run-dir", str(self.run_dir)).returncode)
            job = json.loads(
                (self.run_dir / "state" / "runtime-job.json").read_text(encoding="utf-8")
            )
            stroke_def = [
                s
                for s in job["controllers"][0]["commissioning_flow"]
                if s["step_id"] == "cooling_valve_stroke_no_chw"
            ][0]
            self.assertEqual("chw_valve_stroke_no_plant", stroke_def["arms_test_mode_state_key"])

            self.assertEqual(
                0,
                _run_runtime(
                    "init-flow",
                    "--run-dir",
                    str(self.run_dir),
                    "--controller-label",
                    "FCU-STROKE",
                ).returncode,
            )
            self.assertEqual(
                0,
                _run_runtime(
                    "record-step",
                    "--run-dir",
                    str(self.run_dir),
                    "--controller-label",
                    "FCU-STROKE",
                    "--step-id",
                    "prep",
                    "--status",
                    "passed",
                    "--technician-name",
                    "Alex Tech",
                    "--note",
                    "prep",
                ).returncode,
            )

            blocked = _run_runtime(
                "record-step",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-STROKE",
                "--step-id",
                "cooling_valve_stroke_no_chw",
                "--status",
                "passed",
                "--technician-name",
                "Alex Tech",
                "--note",
                "no confirms",
            )
            self.assertEqual(2, blocked.returncode)
            self.assertIn("commissioning-confirm-prompt", blocked.stdout)
            rej = [
                json.loads(line)
                for line in (self.run_dir / "logs" / "events.jsonl")
                .read_text(encoding="utf-8")
                .strip()
                .splitlines()
                if json.loads(line).get("event") == "flow_step_rejected"
            ]
            self.assertEqual("PROMPTS_NOT_CONFIRMED", rej[-1]["reason_code"])

            for pid in ("chw_valve_at_100", "chw_valve_at_0"):
                c = _run_runtime(
                    "commissioning-confirm-prompt",
                    "--run-dir",
                    str(self.run_dir),
                    "--controller-label",
                    "FCU-STROKE",
                    "--step-id",
                    "cooling_valve_stroke_no_chw",
                    "--prompt-id",
                    pid,
                    "--technician-name",
                    "Alex Tech",
                    "--note",
                    f"confirm {pid}",
                    "--bacnet-timeout-seconds",
                    "0.5",
                    "--bacnet-retries",
                    "1",
                )
                self.assertEqual(0, c.returncode, c.stdout + c.stderr)

            sess = json.loads(
                (
                    self.run_dir / "state" / "sessions" / "FCU-STROKE.json"
                ).read_text(encoding="utf-8")
            )
            self.assertEqual("true", sess["values"]["prompt_confirm.chw_valve_at_100"]["value"])

            ok = _run_runtime(
                "record-step",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-STROKE",
                "--step-id",
                "cooling_valve_stroke_no_chw",
                "--status",
                "passed",
                "--technician-name",
                "Alex Tech",
                "--note",
                "after confirms",
                "--no-run-modulation-on-pass",
            )
            self.assertEqual(0, ok.returncode)
        finally:
            server.stop()

    def test_airflow_half_flow_adjust_then_tach_confirm_gate(self) -> None:
        from test_import_compiler import _write_profile

        server = _FakeBipUdpServer(
            device_instance=21001,
            analog_input_present=21.0,
            msv_present=3,
            av_tacho_present=42.5,
            av_supply_fan_present=0.0,
        )
        server.start()
        try:
            time.sleep(0.05)
            self.run_dir.mkdir(parents=True, exist_ok=True)
            profiles_dir = self.run_dir / "profiles-airflow"
            profiles_dir.mkdir(parents=True, exist_ok=True)
            _write_profile(
                profiles_dir / "unit-airflow.json",
                profile_id="fcu_airflow_slice_v1",
                display_name="Airflow slice",
                write_allowlist=["msv_test_mode", "av_supply_fan_command"],
                read_allowlist=[
                    "ai_sat",
                    "msv_test_mode",
                    "av_supply_fan_command",
                    "av_supply_fan_tacho_value",
                ],
                objects=[
                    {
                        "id": "msv_test_mode",
                        "writable": True,
                        "bacnet": {"object_type": "multiStateValue", "instance": 50},
                    },
                    {
                        "id": "ai_sat",
                        "writable": False,
                        "bacnet": {"object_type": "analogInput", "instance": 2},
                    },
                    {
                        "id": "av_supply_fan_command",
                        "writable": True,
                        "bacnet": {"object_type": "analogValue", "instance": 3},
                    },
                    {
                        "id": "av_supply_fan_tacho_value",
                        "writable": False,
                        "bacnet": {"object_type": "analogValue", "instance": 1},
                    },
                ],
                commissioning_flow=[
                    {"step_id": "prep", "label": "Prep"},
                    {
                        "step_id": "half_design_airflow_auto",
                        "label": "Half design auto",
                        "arms_test_mode_state_key": "airflow_verify",
                        "actions": [
                            {
                                "type": "automatic_airflow_adjustment",
                                "actuator_object_id": "av_supply_fan_command",
                                "target_flow_ratio_of_design": 0.5,
                                "measurement_branch_id": "supply_terminal_main",
                                "tolerance_ratio": 0.05,
                                "tachometer_reference_session_key": (
                                    "fan_tachometer_reference_at_half_design_flow"
                                ),
                            }
                        ],
                    },
                    {
                        "step_id": "confirm_tachometer_reference_half_flow",
                        "label": "Confirm tacho",
                        "actions": [
                            {
                                "type": "operator_confirm_tachometer_reference",
                                "read_object_id": "av_supply_fan_tacho_value",
                                "session_key": "fan_tachometer_reference_at_half_design_flow",
                            }
                        ],
                    },
                ],
                unit_specs={"design_supply_airflow_L_s": 0.85},
            )
            csv_path = self.run_dir / "controllers-airflow.csv"
            with csv_path.open("w", newline="", encoding="utf-8") as handle:
                writer = csv.DictWriter(
                    handle,
                    fieldnames=[
                        "controller_label",
                        "profile_id",
                        "bacnet_device_instance",
                        "bacnet_ip",
                        "bacnet_port",
                        "building_floor",
                        "notes",
                    ],
                )
                writer.writeheader()
                writer.writerow(
                    {
                        "controller_label": "FCU-AIR",
                        "profile_id": "fcu_airflow_slice_v1",
                        "bacnet_device_instance": "21001",
                        "bacnet_ip": "127.0.0.1",
                        "bacnet_port": str(server.port),
                        "building_floor": "L01",
                        "notes": "",
                    }
                )
            self.assertEqual(
                0,
                _run_runtime(
                    "init-run",
                    "--run-dir",
                    str(self.run_dir),
                    "--job-id",
                    "job-airflow",
                    "--controllers-csv",
                    str(csv_path),
                    "--profiles-dir",
                    str(profiles_dir),
                    "--scenarios-dir",
                    str(ROOT / "docs" / "examples" / "simulator-scenarios"),
                ).returncode,
            )
            self.assertEqual(0, _run_runtime("compile-import", "--run-dir", str(self.run_dir)).returncode)
            rj = json.loads(
                (self.run_dir / "state" / "runtime-job.json").read_text(encoding="utf-8")
            )
            ctrl0 = rj["controllers"][0]
            self.assertEqual(
                0.85,
                ctrl0["commissioning_meta"]["unit_specs"]["design_supply_airflow_L_s"],
            )

            self.assertEqual(
                0,
                _run_runtime(
                    "init-flow",
                    "--run-dir",
                    str(self.run_dir),
                    "--controller-label",
                    "FCU-AIR",
                ).returncode,
            )
            self.assertEqual(
                0,
                _run_runtime(
                    "record-step",
                    "--run-dir",
                    str(self.run_dir),
                    "--controller-label",
                    "FCU-AIR",
                    "--step-id",
                    "prep",
                    "--status",
                    "passed",
                    "--technician-name",
                    "Alex Tech",
                    "--note",
                    "prep ok",
                    "--no-run-modulation-on-pass",
                ).returncode,
            )

            adj = _run_runtime(
                "commissioning-airflow-adjust-write",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-AIR",
                "--step-id",
                "half_design_airflow_auto",
                "--fan-command-percent",
                "50",
                "--technician-name",
                "Alex Tech",
                "--note",
                "half target",
                "--bacnet-timeout-seconds",
                "0.5",
                "--bacnet-retries",
                "1",
            )
            self.assertEqual(0, adj.returncode, adj.stdout + adj.stderr)
            self.assertIn("design_supply_airflow_L_s", adj.stdout)

            rd = _run_runtime(
                "bacnet-read",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-AIR",
                "--object-id",
                "av_supply_fan_command",
                "--timeout-seconds",
                "0.5",
                "--retries",
                "1",
            )
            self.assertEqual(0, rd.returncode)
            self.assertIn('"value_str": "50.0"', rd.stdout)

            blocked_half = _run_runtime(
                "record-step",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-AIR",
                "--step-id",
                "half_design_airflow_auto",
                "--status",
                "passed",
                "--technician-name",
                "Alex Tech",
                "--note",
                "skip tacho",
                "--no-run-modulation-on-pass",
            )
            self.assertEqual(2, blocked_half.returncode)
            self.assertIn(
                "commissioning-confirm-tachometer-reference", blocked_half.stdout
            )

            blocked_confirm = _run_runtime(
                "record-step",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-AIR",
                "--step-id",
                "confirm_tachometer_reference_half_flow",
                "--status",
                "passed",
                "--technician-name",
                "Alex Tech",
                "--note",
                "no read",
                "--no-run-modulation-on-pass",
            )
            self.assertEqual(2, blocked_confirm.returncode)

            ct = _run_runtime(
                "commissioning-confirm-tachometer-reference",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-AIR",
                "--step-id",
                "confirm_tachometer_reference_half_flow",
                "--technician-name",
                "Alex Tech",
                "--note",
                "saw tacho",
                "--bacnet-timeout-seconds",
                "0.5",
                "--bacnet-retries",
                "1",
            )
            self.assertEqual(0, ct.returncode, ct.stdout + ct.stderr)
            self.assertIn("42.5", ct.stdout)

            ok_half = _run_runtime(
                "record-step",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-AIR",
                "--step-id",
                "half_design_airflow_auto",
                "--status",
                "passed",
                "--technician-name",
                "Alex Tech",
                "--note",
                "after tacho on next step",
                "--no-run-modulation-on-pass",
            )
            self.assertEqual(0, ok_half.returncode)

            ok_confirm = _run_runtime(
                "record-step",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-AIR",
                "--step-id",
                "confirm_tachometer_reference_half_flow",
                "--status",
                "passed",
                "--technician-name",
                "Alex Tech",
                "--note",
                "confirmed",
                "--no-run-modulation-on-pass",
            )
            self.assertEqual(0, ok_confirm.returncode)
        finally:
            server.stop()

    def test_manual_airflow_record_then_record_step_pass(self) -> None:
        from test_import_compiler import _write_profile

        self.run_dir.mkdir(parents=True, exist_ok=True)
        profiles_dir = self.run_dir / "profiles-manual-air"
        profiles_dir.mkdir(parents=True, exist_ok=True)
        _write_profile(
            profiles_dir / "unit-manual-air.json",
            profile_id="fcu_manual_air_v1",
            display_name="Manual air",
            write_allowlist=["msv_test_mode"],
            read_allowlist=["msv_test_mode"],
            objects=[
                {
                    "id": "msv_test_mode",
                    "writable": True,
                    "bacnet": {"object_type": "multiStateValue", "instance": 50},
                },
            ],
            commissioning_flow=[
                {"step_id": "prep", "label": "Prep"},
                {
                    "step_id": "manual_airflow_only",
                    "label": "Manual airflow",
                    "actions": [
                        {
                            "type": "manual_airflow_verification_assisted",
                            "branch_ids": ["supply_main"],
                        }
                    ],
                },
            ],
            airflow_verification={
                "branches": [
                    {
                        "id": "supply_main",
                        "design_flow_L_s": 0.5,
                        "measurement": {
                            "allowed_tools": ["balometer"],
                        },
                    }
                ]
            },
        )
        csv_path = self.run_dir / "controllers-manual-air.csv"
        with csv_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=[
                    "controller_label",
                    "profile_id",
                    "bacnet_device_instance",
                    "bacnet_ip",
                    "bacnet_port",
                    "building_floor",
                    "notes",
                ],
            )
            writer.writeheader()
            writer.writerow(
                {
                    "controller_label": "FCU-MAN",
                    "profile_id": "fcu_manual_air_v1",
                    "bacnet_device_instance": "21001",
                    "bacnet_ip": "127.0.0.1",
                    "bacnet_port": "47808",
                    "building_floor": "L01",
                    "notes": "",
                }
            )
        self.assertEqual(
            0,
            _run_runtime(
                "init-run",
                "--run-dir",
                str(self.run_dir),
                "--job-id",
                "job-man-air",
                "--controllers-csv",
                str(csv_path),
                "--profiles-dir",
                str(profiles_dir),
                "--scenarios-dir",
                str(ROOT / "docs" / "examples" / "simulator-scenarios"),
            ).returncode,
        )
        self.assertEqual(0, _run_runtime("compile-import", "--run-dir", str(self.run_dir)).returncode)
        rj = json.loads((self.run_dir / "state" / "runtime-job.json").read_text(encoding="utf-8"))
        self.assertEqual(
            "supply_main",
            rj["controllers"][0]["commissioning_meta"]["airflow_verification"]["branches"][0]["id"],
        )

        self.assertEqual(
            0,
            _run_runtime(
                "init-flow",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-MAN",
            ).returncode,
        )
        self.assertEqual(
            0,
            _run_runtime(
                "record-step",
                "--run-dir",
                str(self.run_dir),
                "--controller-label",
                "FCU-MAN",
                "--step-id",
                "prep",
                "--status",
                "passed",
                "--technician-name",
                "Alex Tech",
                "--note",
                "prep",
                "--no-run-modulation-on-pass",
            ).returncode,
        )

        blocked = _run_runtime(
            "record-step",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-MAN",
            "--step-id",
            "manual_airflow_only",
            "--status",
            "passed",
            "--technician-name",
            "Alex Tech",
            "--note",
            "no measurement",
            "--no-run-modulation-on-pass",
        )
        self.assertEqual(2, blocked.returncode)
        self.assertIn("commissioning-record-manual-airflow", blocked.stdout)

        rec = _run_runtime(
            "commissioning-record-manual-airflow",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-MAN",
            "--step-id",
            "manual_airflow_only",
            "--branch-id",
            "supply_main",
            "--measured-flow-L-s",
            "0.44",
            "--measurement-tool",
            "balometer",
            "--technician-name",
            "Alex Tech",
            "--note",
            "traverse",
        )
        self.assertEqual(0, rec.returncode, rec.stdout + rec.stderr)
        self.assertIn("manual_airflow_measured_supply_main_L_s", rec.stdout)

        ok = _run_runtime(
            "record-step",
            "--run-dir",
            str(self.run_dir),
            "--controller-label",
            "FCU-MAN",
            "--step-id",
            "manual_airflow_only",
            "--status",
            "passed",
            "--technician-name",
            "Alex Tech",
            "--note",
            "after measurement",
            "--no-run-modulation-on-pass",
        )
        self.assertEqual(0, ok.returncode)

        report_path = self.run_dir / "artifacts" / "commissioning_report.json"
        self.assertTrue(report_path.is_file())
        rdoc = json.loads(report_path.read_text(encoding="utf-8"))
        kinds = [e.get("kind") for e in rdoc.get("entries", []) if isinstance(e, dict)]
        self.assertIn("manual_airflow_measurement", kinds)
        man_ent = next(
            e
            for e in rdoc["entries"]
            if isinstance(e, dict) and e.get("kind") == "manual_airflow_measurement"
        )
        self.assertEqual("supply_main", man_ent.get("branch_id"))
        self.assertAlmostEqual(0.44, float(man_ent.get("measured_flow_L_s", 0)))

        unified_csv = self.run_dir / "artifacts" / "man-air-unified.csv"
        exu = _run_runtime(
            "export-commissioning-report",
            "--run-dir",
            str(self.run_dir),
            "--output-csv-unified",
            str(unified_csv),
        )
        self.assertEqual(0, exu.returncode)
        csv_text = unified_csv.read_text(encoding="utf-8")
        self.assertIn("manual_airflow_measurement", csv_text)
        self.assertIn("measurement_branch_id", csv_text)
        self.assertIn("supply_main", csv_text)
        self.assertIn("0.44", csv_text)
