#!/usr/bin/env python3
"""Runtime CLI skeleton for commissioning workflows."""

from __future__ import annotations

import argparse
from typing import Any
import csv
import html
import json
import importlib.util
import shutil
import sys
import time
from datetime import UTC, datetime
from pathlib import Path

from repo_root import repo_root


ROOT = repo_root()
IMPORT_COMPILER = ROOT / "tools" / "import" / "compile_job.py"
SIMULATOR_ORCH = ROOT / "tools" / "simulator" / "orchestrator.py"
BACNET_ADAPTER = ROOT / "tools" / "bacnet" / "adapter.py"
OPERATOR_GUI_SERVER = ROOT / "tools" / "operator_gui_server.py"

_bacnet_adapter_singleton = None
_compile_job_module = None
_orchestrator_module = None


def _compile_job_module_loaded():
    global _compile_job_module
    if _compile_job_module is not None:
        return _compile_job_module
    spec = importlib.util.spec_from_file_location(
        "runtime_import_compile_job", IMPORT_COMPILER
    )
    if spec is None or spec.loader is None:
        raise RuntimeError(f"unable to load import compiler: {IMPORT_COMPILER}")
    mod = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = mod
    spec.loader.exec_module(mod)
    _compile_job_module = mod
    return mod


def _orchestrator_module_loaded():
    global _orchestrator_module
    if _orchestrator_module is not None:
        return _orchestrator_module
    spec = importlib.util.spec_from_file_location(
        "runtime_simulator_orchestrator", SIMULATOR_ORCH
    )
    if spec is None or spec.loader is None:
        raise RuntimeError(f"unable to load simulator orchestrator: {SIMULATOR_ORCH}")
    mod = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = mod
    spec.loader.exec_module(mod)
    _orchestrator_module = mod
    return mod


def _bacnet_adapter():
    """Lazy singleton :class:`CommissioningBACnetAdapter` (see ``tools/bacnet/adapter.py``)."""
    global _bacnet_adapter_singleton
    if _bacnet_adapter_singleton is not None:
        return _bacnet_adapter_singleton
    spec = importlib.util.spec_from_file_location(
        "runtime_commissioning_bacnet", BACNET_ADAPTER
    )
    if spec is None or spec.loader is None:
        raise RuntimeError(f"unable to load BACnet adapter module: {BACNET_ADAPTER}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    _bacnet_adapter_singleton = module.CommissioningBACnetAdapter(ROOT)
    return _bacnet_adapter_singleton


def _utc_timestamp() -> str:
    return datetime.now(UTC).isoformat().replace("+00:00", "Z")


def _append_event(log_path: Path, event: str, payload: dict | None = None) -> None:
    payload = payload or {}
    entry = {"ts": _utc_timestamp(), "event": event, **payload}
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with log_path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(entry, sort_keys=True) + "\n")


def _parse_run_config(run_dir: Path) -> dict:
    config_path = run_dir / "config" / "runtime-config.json"
    return json.loads(config_path.read_text(encoding="utf-8"))


def _probe_bip(
    host: str,
    port: int,
    expected_device_instance: int,
    timeout_seconds: float,
    retries: int,
) -> dict:
    return _bacnet_adapter().probe_device(
        host=host,
        port=port,
        expected_device_instance=expected_device_instance,
        timeout_seconds=timeout_seconds,
        retries=retries,
    )


def _flows_dir(run_dir: Path) -> Path:
    return run_dir / "state" / "flows"


def _flow_state_path(run_dir: Path, controller_label: str) -> Path:
    return _flows_dir(run_dir) / f"{controller_label}.json"


def _sessions_dir(run_dir: Path) -> Path:
    return run_dir / "state" / "sessions"


def _session_state_path(run_dir: Path, controller_label: str) -> Path:
    return _sessions_dir(run_dir) / f"{controller_label}.json"


def _flow_backups_dir(run_dir: Path) -> Path:
    return run_dir / "state" / "flow_backups"


def _step_status_counts(steps: list[dict]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for item in steps:
        status = str(item.get("status", "pending"))
        counts[status] = counts.get(status, 0) + 1
    return counts


def _sequencing_complete_statuses() -> frozenset[str]:
    return frozenset({"passed", "manual_passed", "skipped"})


def _is_sequencing_complete_status(status: str) -> bool:
    """Prior steps must reach this before later steps can record pass/fail/skip outcomes."""
    return status in _sequencing_complete_statuses()


def _next_open_step(steps: list[dict]) -> dict | None:
    """First step not in a sequencing-complete terminal state (for operator 'next' hint)."""
    complete = _sequencing_complete_statuses()
    for item in steps:
        status = str(item.get("status", "pending"))
        if status not in complete:
            return {
                "step_id": item.get("step_id"),
                "label": item.get("label"),
                "status": status,
            }
    return None


def _find_modulate_actuator_action(step: dict) -> dict | None:
    actions = step.get("actions")
    if not isinstance(actions, list):
        return None
    for act in actions:
        if isinstance(act, dict) and str(act.get("type", "")).strip() == "modulate_actuator_log_sat_for_report":
            return act
    return None


def _find_automatic_airflow_adjustment(step: dict) -> dict | None:
    actions = step.get("actions")
    if not isinstance(actions, list):
        return None
    for act in actions:
        if isinstance(act, dict) and str(act.get("type", "")).strip() == "automatic_airflow_adjustment":
            return act
    return None


def _find_operator_confirm_tachometer(step: dict) -> dict | None:
    actions = step.get("actions")
    if not isinstance(actions, list):
        return None
    for act in actions:
        if isinstance(act, dict) and str(act.get("type", "")).strip() == "operator_confirm_tachometer_reference":
            return act
    return None


def _find_manual_airflow_verification(step: dict) -> dict | None:
    actions = step.get("actions")
    if not isinstance(actions, list):
        return None
    for act in actions:
        if isinstance(act, dict) and str(act.get("type", "")).strip() == "manual_airflow_verification_assisted":
            return act
    return None


def _suggested_cli_commands_for_step(step: dict) -> list[str]:
    """Human-readable CLI hints for scripting operators (no subprocess)."""
    out: list[str] = []
    stype = str(step.get("step_type", "")).strip()
    sid = str(step.get("step_id", "")).strip()
    if stype == "bacnet_point_checkout":
        out.append("bacnet-point-checkout --run-dir <run-dir> --controller-label <label>")
    if _find_modulate_actuator_action(step) is not None:
        out.append(
            "bacnet-modulation-sweep --run-dir <run-dir> --controller-label <label> "
            f"--step-id {sid or '<step_id>'} --modulation-command-percents ... "
            "[--technician-name ...]"
        )
    if _find_automatic_airflow_adjustment(step) is not None:
        out.append(
            "commissioning-airflow-adjust-write --run-dir <run-dir> --controller-label <label> "
            f"--step-id {sid or '<step_id>'} --fan-command-percent <0-100> --technician-name ..."
        )
    if _find_operator_confirm_tachometer(step) is not None:
        out.append(
            "commissioning-confirm-tachometer-reference --run-dir <run-dir> "
            f"--controller-label <label> --step-id {sid or '<step_id>'} --technician-name ..."
        )
    if _find_manual_airflow_verification(step) is not None:
        out.append(
            "commissioning-record-manual-airflow --run-dir <run-dir> --controller-label <label> "
            f"--step-id {sid or '<step_id>'} --branch-id <id> --measured-flow-L-s <n> --measurement-tool ..."
        )
    arms = str(step.get("arms_test_mode_state_key", "")).strip()
    if arms == "chw_valve_stroke_no_plant":
        out.append(
            "commissioning-confirm-prompt --run-dir <run-dir> --controller-label <label> "
            f"--step-id {sid or '<step_id>'} --prompt-id <id> --technician-name ..."
        )
    if not out:
        out.append(
            "record-step --run-dir <run-dir> --controller-label <label> "
            f"--step-id {sid or '<step_id>'} --status passed|failed|skipped|manual_passed ..."
        )
    return out


def _step_guidance_blocked_reasons(
    step: dict, steps_by_id: dict[str, dict], session_vals: dict[str, str]
) -> list[str]:
    """Why a pending step may be blocked (prereqs / skip gates); empty if not blocked or not pending."""
    complete = _sequencing_complete_statuses()
    status = str(step.get("status", "pending"))
    if status in complete:
        return []
    reasons: list[str] = []
    raw_req = step.get("requires_step_ids")
    if isinstance(raw_req, list):
        for rid in raw_req:
            rid_s = str(rid).strip()
            if not rid_s:
                continue
            dep = steps_by_id.get(rid_s)
            if dep is None:
                reasons.append(f"missing_prerequisite_step:{rid_s}")
                continue
            dst = str(dep.get("status", "pending"))
            if not _is_terminal_prereq_status(dst):
                reasons.append(f"prerequisite_not_complete:{rid_s}({dst})")
    if step.get("skippable") is True:
        raw_sw = step.get("skip_when")
        if isinstance(raw_sw, list) and raw_sw:
            missing = []
            for code in raw_sw:
                c = str(code).strip()
                if not c:
                    continue
                if not _session_flag_truthy(session_vals.get(c, "")):
                    missing.append(c)
            if missing:
                reasons.append(f"skip_when_session_not_set:{','.join(missing)}")
    return reasons


def _flow_steps_by_id(steps: list[dict]) -> dict[str, dict]:
    out: dict[str, dict] = {}
    for item in steps:
        if not isinstance(item, dict):
            continue
        sid = str(item.get("step_id", "")).strip()
        if sid:
            out[sid] = item
    return out


def _compact_flow_guidance(steps: list[dict]) -> dict[str, Any]:
    """Operator-oriented commissioning flow summary (CLI guided slice)."""
    complete = _sequencing_complete_statuses()
    items: list[dict[str, Any]] = []
    for item in steps:
        if not isinstance(item, dict):
            continue
        sid = str(item.get("step_id", "")).strip()
        if not sid:
            continue
        st = str(item.get("status", "pending"))
        row: dict[str, Any] = {
            "step_id": sid,
            "label": str(item.get("label", "")).strip(),
            "status": st,
        }
        arms = str(item.get("arms_test_mode_state_key", "")).strip()
        if arms:
            row["arms_test_mode_state_key"] = arms
        stype = str(item.get("step_type", "")).strip()
        if stype and stype != "standard":
            row["step_type"] = stype
        raw_req = item.get("requires_step_ids")
        if isinstance(raw_req, list):
            req_ids = [str(x).strip() for x in raw_req if str(x).strip()]
            if req_ids:
                row["requires_step_ids"] = req_ids
        if item.get("skippable") is True:
            raw_sw = item.get("skip_when")
            if isinstance(raw_sw, list) and any(str(c).strip() for c in raw_sw):
                row["skip_when"] = [str(c).strip() for c in raw_sw if str(c).strip()]
        items.append(row)
    pending = sum(1 for i in items if str(i.get("status", "")) not in complete)
    return {
        "step_count": len(items),
        "pending_step_count": pending,
        "all_sequencing_complete": pending == 0,
        "next_open_step": _next_open_step(steps),
        "steps": items,
    }


def _enrich_guidance_for_operator_json(
    base: dict[str, Any], steps: list[dict], session_vals: dict[str, str]
) -> dict[str, Any]:
    """Attach suggested_cli_commands and blocked_reasons per step."""
    steps_by_id = _flow_steps_by_id([s for s in steps if isinstance(s, dict)])
    out_steps: list[dict[str, Any]] = []
    for row in base.get("steps", []):
        if not isinstance(row, dict):
            continue
        sid = str(row.get("step_id", "")).strip()
        full = steps_by_id.get(sid, {})
        enriched = dict(row)
        enriched["suggested_cli_commands"] = _suggested_cli_commands_for_step(full)
        enriched["blocked_reasons"] = _step_guidance_blocked_reasons(
            full, steps_by_id, session_vals
        )
        out_steps.append(enriched)
    base = dict(base)
    base["steps"] = out_steps
    return base


def _is_terminal_prereq_status(status: str) -> bool:
    return status in {"passed", "manual_passed", "skipped"}


def _lookup_step_by_id(steps: list[dict], step_id: str) -> dict | None:
    for item in steps:
        if item.get("step_id") == step_id:
            return item
    return None


def _session_values_map(session_state: dict) -> dict[str, str]:
    """Return raw string values for session keys (empty if missing)."""
    out: dict[str, str] = {}
    values = session_state.get("values")
    if not isinstance(values, dict):
        return out
    for key, meta in values.items():
        k = str(key).strip()
        if not k:
            continue
        if isinstance(meta, dict):
            out[k] = str(meta.get("value", ""))
        else:
            out[k] = str(meta)
    return out


def _session_flag_truthy(raw: str) -> bool:
    return str(raw).strip().lower() in {"1", "true", "yes", "y", "on"}


def _load_session_state(run_dir: Path, controller_label: str) -> dict | None:
    path = _session_state_path(run_dir, controller_label)
    if not path.is_file():
        return None
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None
    return data if isinstance(data, dict) else None


def _chw_valve_stroke_prompt_confirm_session_keys(step: dict) -> list[str]:
    """Session keys ``prompt_confirm.<prompt_id>`` for CHW valve stroke + operator prompts."""
    actions = step.get("actions")
    if not isinstance(actions, list):
        return []
    keys: list[str] = []
    prev_write_oid = ""
    for act in actions:
        if not isinstance(act, dict):
            continue
        t = str(act.get("type", "")).strip()
        if t == "write_analog_percent":
            prev_write_oid = str(act.get("object_id", "")).strip()
        elif t == "operator_prompt_confirm":
            pid = str(act.get("prompt_id", "")).strip()
            if pid and prev_write_oid == "ao_chw_valve":
                keys.append(f"prompt_confirm.{pid}")
    return keys


def _cooling_valve_stroke_no_chw_gating(step: dict) -> bool:
    """Steps that require BACnet writes + recorded confirmations before pass."""
    if str(step.get("step_id", "")).strip() == "cooling_valve_stroke_no_chw":
        return bool(_chw_valve_stroke_prompt_confirm_session_keys(step))
    arms = str(step.get("arms_test_mode_state_key", "")).strip()
    if arms == "chw_valve_stroke_no_plant":
        return bool(_chw_valve_stroke_prompt_confirm_session_keys(step))
    return False


def _operator_confirm_tachometer_session_key(step: dict) -> str | None:
    """Session key from ``operator_confirm_tachometer_reference`` action, if any."""
    actions = step.get("actions")
    if not isinstance(actions, list):
        return None
    for act in actions:
        if not isinstance(act, dict):
            continue
        if str(act.get("type", "")).strip() != "operator_confirm_tachometer_reference":
            continue
        session_key = str(act.get("session_key", "")).strip()
        read_oid = str(act.get("read_object_id", "")).strip()
        if session_key and read_oid:
            return session_key
    return None


def _tachometer_confirmation_gating(step: dict) -> bool:
    return _operator_confirm_tachometer_session_key(step) is not None


def _default_manual_airflow_session_key(branch_id: str) -> str:
    bid = str(branch_id).strip()
    return f"manual_airflow_measured_{bid}_L_s"


def _manual_airflow_verification_session_keys(step: dict) -> list[str]:
    """Session keys required before pass for ``manual_airflow_verification_assisted`` steps."""
    actions = step.get("actions")
    if not isinstance(actions, list):
        return []
    keys: list[str] = []
    for act in actions:
        if not isinstance(act, dict):
            continue
        if str(act.get("type", "")).strip() != "manual_airflow_verification_assisted":
            continue
        raw_map = act.get("session_keys")
        overrides: dict[str, str] = {}
        if isinstance(raw_map, dict):
            for k, v in raw_map.items():
                kk = str(k).strip()
                vv = str(v).strip()
                if kk and vv:
                    overrides[kk] = vv
        raw_branches = act.get("branch_ids")
        if not isinstance(raw_branches, list):
            continue
        for bid in raw_branches:
            b = str(bid).strip()
            if not b:
                continue
            keys.append(overrides.get(b, _default_manual_airflow_session_key(b)))
    return keys


def _manual_airflow_verification_gating(step: dict) -> bool:
    return bool(_manual_airflow_verification_session_keys(step))


def _session_has_recorded_measurement(vals: dict[str, str], key: str) -> bool:
    raw = str(vals.get(key, "")).strip()
    return bool(raw)


def _airflow_adjust_tachometer_reference_session_key(step: dict) -> str | None:
    """Optional ``tachometer_reference_session_key`` on ``automatic_airflow_adjustment`` action."""
    actions = step.get("actions")
    if not isinstance(actions, list):
        return None
    for act in actions:
        if not isinstance(act, dict):
            continue
        if str(act.get("type", "")).strip() != "automatic_airflow_adjustment":
            continue
        sk = str(act.get("tachometer_reference_session_key", "")).strip()
        return sk or None
    return None


def _validate_step_transition(
    steps: list[dict],
    step: dict,
    requested_status: str,
    *,
    session_values: dict[str, str] | None = None,
) -> dict[str, str] | None:
    """Return reason dict when transition is invalid, otherwise None."""
    if requested_status == "pending":
        return {
            "reason_code": "pending_not_recordable",
            "message": (
                "cannot record step with status 'pending'; "
                "use passed, failed, skipped, or manual_passed to record outcomes"
            ),
        }

    if requested_status == "skipped" and step.get("skippable") is not True:
        return {
            "reason_code": "step_not_skippable",
            "message": f"step '{step.get('step_id')}' is not skippable",
        }

    if requested_status == "skipped":
        raw_sw = step.get("skip_when")
        if isinstance(raw_sw, list) and raw_sw:
            codes = [str(c).strip() for c in raw_sw if str(c).strip()]
            if codes:
                vals = session_values if session_values is not None else {}
                if not any(_session_flag_truthy(vals.get(code, "")) for code in codes):
                    joined = ", ".join(codes)
                    return {
                        "reason_code": "skip_reason_not_recorded",
                        "message": (
                            f"step '{step.get('step_id')}' requires a matching session flag "
                            f"before skip (set-session-value): one of [{joined}] must be truthy "
                            f"(e.g. true/1/yes)"
                        ),
                    }

    if requested_status in {"passed", "manual_passed"}:
        if _cooling_valve_stroke_no_chw_gating(step):
            vals = session_values if session_values is not None else {}
            pend = _chw_valve_stroke_prompt_confirm_session_keys(step)
            missing = [k for k in pend if not _session_flag_truthy(vals.get(k, ""))]
            if missing:
                return {
                    "reason_code": "operator_prompts_not_confirmed",
                    "message": (
                        f"step '{step.get('step_id')}' requires operator prompt confirmations "
                        f"before pass; use commissioning-confirm-prompt for each prompt_id "
                        f"(missing session keys: {missing})"
                    ),
                }

    if requested_status in {"passed", "manual_passed", "failed"}:
        step_id = step.get("step_id")
        requires_step_ids = step.get("requires_step_ids", [])
        if isinstance(requires_step_ids, list):
            for required_id in requires_step_ids:
                required = _lookup_step_by_id(steps, str(required_id))
                if required is None:
                    return {
                        "reason_code": "dependency_missing_from_flow",
                        "message": (
                            f"step '{step_id}' requires completed dependency "
                            f"'{required_id}' which is not present in flow"
                        ),
                    }
                required_status = str(required.get("status", "pending"))
                if not _is_terminal_prereq_status(required_status):
                    return {
                        "reason_code": "dependency_not_completed",
                        "message": (
                            f"step '{step_id}' requires completed dependency "
                            f"'{required_id}'"
                        ),
                    }
        current_index = next(
            (idx for idx, item in enumerate(steps) if item.get("step_id") == step_id),
            None,
        )
        if current_index is None:
            return {
                "reason_code": "step_not_found_in_sequence",
                "message": f"step '{step_id}' not found in flow sequence",
            }
        for prev in steps[:current_index]:
            prev_status = str(prev.get("status", "pending"))
            if not _is_sequencing_complete_status(prev_status):
                return {
                    "reason_code": "prior_step_incomplete",
                    "message": (
                        f"step '{step_id}' cannot be marked {requested_status} "
                        f"before '{prev.get('step_id')}' is completed"
                    ),
                }

    if requested_status in {"passed", "manual_passed"}:
        vals = session_values if session_values is not None else {}
        sk_air = _airflow_adjust_tachometer_reference_session_key(step)
        if sk_air and not _session_flag_truthy(vals.get(sk_air, "")):
            return {
                "reason_code": "tachometer_reference_not_confirmed",
                "message": (
                    f"step '{step.get('step_id')}' requires technician confirmation of the "
                    f"BACnet tachometer before pass (profile automatic_airflow_adjustment "
                    f"tachometer_reference_session_key); run commissioning-confirm-tachometer-reference "
                    f"(missing session key: {sk_air!r})"
                ),
            }
        if _tachometer_confirmation_gating(step):
            sk = _operator_confirm_tachometer_session_key(step)
            if sk and not _session_flag_truthy(vals.get(sk, "")):
                return {
                    "reason_code": "tachometer_reference_not_confirmed",
                    "message": (
                        f"step '{step.get('step_id')}' requires technician confirmation "
                        f"of the BACnet tachometer reading before pass; run "
                        f"commissioning-confirm-tachometer-reference (missing session key: {sk!r})"
                    ),
                }
        if _manual_airflow_verification_gating(step):
            pend = _manual_airflow_verification_session_keys(step)
            missing = [k for k in pend if not _session_has_recorded_measurement(vals, k)]
            if missing:
                return {
                    "reason_code": "manual_airflow_measurement_missing",
                    "message": (
                        f"step '{step.get('step_id')}' requires measured airflow (L/s) in session "
                        f"before pass; run commissioning-record-manual-airflow for each branch "
                        f"(missing keys: {missing})"
                    ),
                }
    return None


def _normalize_rejection_reason(reason_code: str) -> str:
    mapping = {
        "step_not_skippable": "STEP_NOT_SKIPPABLE",
        "skip_reason_not_recorded": "SKIP_GATE",
        "operator_prompts_not_confirmed": "PROMPTS_NOT_CONFIRMED",
        "tachometer_reference_not_confirmed": "TACHOMETER_REFERENCE_NOT_CONFIRMED",
        "manual_airflow_measurement_missing": "MANUAL_AIRFLOW_MEASUREMENT_MISSING",
        "dependency_not_completed": "DEPENDENCY_UNSATISFIED",
        "dependency_missing_from_flow": "DEPENDENCY_UNSATISFIED",
        "prior_step_incomplete": "PREREQ_ORDER",
        "step_not_found_in_sequence": "PREREQ_ORDER",
        "pending_not_recordable": "INVALID_TRANSITION",
    }
    return mapping.get(reason_code, "INVALID_TRANSITION")


def cmd_init_run(args: argparse.Namespace) -> int:
    run_dir = args.run_dir
    artifacts_dir = run_dir / "artifacts"
    state_dir = run_dir / "state"
    logs_dir = run_dir / "logs"
    config_dir = run_dir / "config"
    log_path = logs_dir / "events.jsonl"
    config_path = config_dir / "runtime-config.json"

    run_dir.mkdir(parents=True, exist_ok=True)
    artifacts_dir.mkdir(parents=True, exist_ok=True)
    state_dir.mkdir(parents=True, exist_ok=True)
    logs_dir.mkdir(parents=True, exist_ok=True)
    config_dir.mkdir(parents=True, exist_ok=True)

    config = {
        "job_id": args.job_id,
        "controllers_csv": str(args.controllers_csv.resolve()),
        "profiles_dir": str(args.profiles_dir.resolve()),
        "scenarios_dir": str(args.scenarios_dir.resolve()),
    }
    config_path.write_text(json.dumps(config, indent=2), encoding="utf-8")
    _append_event(log_path, "run_initialized", {"run_dir": str(run_dir.resolve())})
    branding = _ensure_default_branding_placeholder(artifacts_dir)
    if branding is not None and branding.is_file():
        _append_event(
            log_path,
            "branding_placeholder_installed",
            {"logo_png": str(branding.resolve())},
        )
    print(f"run_initialized=true run_dir={run_dir}")
    return 0


def cmd_compile_import(args: argparse.Namespace) -> int:
    run_dir = args.run_dir
    config = _parse_run_config(run_dir)
    state_dir = run_dir / "state"
    logs_path = run_dir / "logs" / "events.jsonl"
    output_json = state_dir / "runtime-job.json"
    report_json = state_dir / "import-report.json"

    cj = _compile_job_module_loaded()
    code = cj.run_compile(
        Path(config["controllers_csv"]),
        Path(config["profiles_dir"]),
        output_json,
        report_json,
    )

    _append_event(
        logs_path,
        "import_compiled",
        {"exit_code": code, "report_json": str(report_json.resolve())},
    )
    return code


def cmd_validate_import(args: argparse.Namespace) -> int:
    """Compile import into an isolated directory without overwriting runtime-job.json."""
    run_dir = args.run_dir
    logs_path = run_dir / "logs" / "events.jsonl"
    config = _parse_run_config(run_dir)
    out_dir = args.output_dir or (run_dir / "artifacts" / "import-validation")
    out_dir.mkdir(parents=True, exist_ok=True)
    output_json = out_dir / "runtime-job.json"
    report_json = out_dir / "import-report.json"

    cj = _compile_job_module_loaded()
    code = cj.run_compile(
        Path(config["controllers_csv"]),
        Path(config["profiles_dir"]),
        output_json,
        report_json,
    )

    _append_event(
        logs_path,
        "import_validated",
        {
            "exit_code": code,
            "validation_dir": str(out_dir.resolve()),
            "report_json": str(report_json.resolve()),
        },
    )
    print(f"import_validated=true validation_dir={out_dir.resolve()}")
    return code


def cmd_print_job_graph(args: argparse.Namespace) -> int:
    run_dir = args.run_dir
    logs_path = run_dir / "logs" / "events.jsonl"
    runtime_job_path = run_dir / "state" / "runtime-job.json"
    if not runtime_job_path.is_file():
        print(
            f"error: runtime job missing at {runtime_job_path}; run compile-import first"
        )
        return 2

    runtime_job = json.loads(runtime_job_path.read_text(encoding="utf-8"))
    config = _parse_run_config(run_dir)
    lines: list[str] = []
    lines.append(f"job_id={config.get('job_id')}")
    lines.append(
        f"controller_count={runtime_job.get('summary', {}).get('controller_count', len(runtime_job.get('controllers', [])))}"
    )
    for row in runtime_job.get("controllers", []):
        label = str(row.get("controller_label", "")).strip()
        flow = row.get("commissioning_flow", [])
        step_count = len(flow) if isinstance(flow, list) else 0
        skip_gated_steps = 0
        modulation_action_steps = 0
        if isinstance(flow, list):
            for st in flow:
                if not isinstance(st, dict):
                    continue
                raw_sw = st.get("skip_when")
                if (
                    st.get("skippable") is True
                    and isinstance(raw_sw, list)
                    and any(str(c).strip() for c in raw_sw)
                ):
                    skip_gated_steps += 1
                if _find_modulate_actuator_action(st) is not None:
                    modulation_action_steps += 1
        objs = row.get("objects_by_id", {})
        obj_count = len(objs) if isinstance(objs, dict) else 0
        w_allow = row.get("commissioning_write_allowlist", [])
        r_allow = row.get("commissioning_read_allowlist", [])
        checkout = row.get("point_checkout", [])
        pc_count = len(checkout) if isinstance(checkout, list) else 0
        wn = len(w_allow) if isinstance(w_allow, list) else 0
        rn = len(r_allow) if isinstance(r_allow, list) else 0
        lines.append(
            f"  {label} profile_id={row.get('profile_id')} "
            f"steps={step_count} objects_by_id={obj_count} "
            f"write_allowlist={wn} read_allowlist={rn} point_checkout={pc_count} "
            f"skip_gated_steps={skip_gated_steps} modulation_action_steps={modulation_action_steps}"
        )

    text = "\n".join(lines) + "\n"
    print(text, end="")
    _append_event(
        logs_path,
        "job_graph_printed",
        {"line_count": len(lines)},
    )
    return 0


def cmd_verify_simulator(args: argparse.Namespace) -> int:
    run_dir = args.run_dir
    config = _parse_run_config(run_dir)
    artifacts_dir = run_dir / "artifacts"
    logs_path = run_dir / "logs" / "events.jsonl"
    output_file = artifacts_dir / "simulator" / f"{args.profile}-{args.scenario}.json"

    orch = _orchestrator_module_loaded()
    code = orch.run_orchestrator(
        args.profile,
        args.scenario,
        strict=True,
        output="json",
        output_file=output_file,
        controllers_csv=Path(config["controllers_csv"]),
        scenarios_dir=Path(config["scenarios_dir"]),
    )

    _append_event(
        logs_path,
        "simulator_verified",
        {
            "exit_code": code,
            "profile": args.profile,
            "scenario": args.scenario,
            "artifact_json": str(output_file.resolve()),
        },
    )
    return code


def cmd_init_flow(args: argparse.Namespace) -> int:
    run_dir = args.run_dir
    logs_path = run_dir / "logs" / "events.jsonl"
    runtime_job_path = run_dir / "state" / "runtime-job.json"
    flow_state_path = _flow_state_path(run_dir, args.controller_label)

    runtime_job = json.loads(runtime_job_path.read_text(encoding="utf-8"))
    controllers = runtime_job.get("controllers", [])
    target = None
    for controller in controllers:
        if controller.get("controller_label") == args.controller_label:
            target = controller
            break
    if target is None:
        print(f"error: controller not found in runtime job: {args.controller_label}")
        return 2

    if flow_state_path.is_file() and not bool(getattr(args, "force", False)):
        print(
            "error: commissioning flow state already exists for this controller; "
            "use init-flow --force with --reset-technician-name and --reset-reason to replace"
        )
        return 2

    if flow_state_path.is_file() and bool(getattr(args, "force", False)):
        tech = str(getattr(args, "reset_technician_name", "") or "").strip()
        reason = str(getattr(args, "reset_reason", "") or "").strip()
        if not tech or not reason:
            print(
                "error: --force requires non-empty --reset-technician-name and --reset-reason"
            )
            return 2
        backup_dir = _flow_backups_dir(run_dir)
        backup_dir.mkdir(parents=True, exist_ok=True)
        stamp = _utc_timestamp().replace(":", "-")
        backup_path = backup_dir / f"{args.controller_label}-{stamp}.json"
        backup_path.write_bytes(flow_state_path.read_bytes())
        _append_event(
            logs_path,
            "flow_reinitialized",
            {
                "controller_label": args.controller_label,
                "previous_flow_backup_json": str(backup_path.resolve()),
                "reset_technician_name": tech,
                "reset_reason": reason,
            },
        )

    step_defs = target.get("commissioning_flow", [])
    steps = []
    for step in step_defs:
        step_id = str(step.get("step_id", "")).strip()
        if not step_id:
            continue
        requires_step_ids = []
        if isinstance(step.get("requires_step_ids"), list):
            for required_id in step["requires_step_ids"]:
                text = str(required_id).strip()
                if text:
                    requires_step_ids.append(text)
        step_row = {
            "step_id": step_id,
            "label": str(step.get("label", "")).strip(),
            "status": "pending",
            "step_type": str(step.get("step_type", "standard")).strip() or "standard",
            "run_point_checkout_on_pass": bool(step.get("run_point_checkout_on_pass")),
            "skippable": step.get("skippable") is True,
            "requires_step_ids": requires_step_ids,
            "records": [],
            "history": [],
        }
        raw_skip_when = step.get("skip_when")
        if isinstance(raw_skip_when, list) and raw_skip_when:
            skip_codes = [str(c).strip() for c in raw_skip_when if str(c).strip()]
            if skip_codes:
                step_row["skip_when"] = skip_codes
        report_ref = str(step.get("report_ref", "")).strip()
        if report_ref:
            step_row["report_ref"] = report_ref
        arms_key = str(step.get("arms_test_mode_state_key", "")).strip()
        if arms_key:
            step_row["arms_test_mode_state_key"] = arms_key
        raw_actions = step.get("actions")
        if isinstance(raw_actions, list) and raw_actions:
            step_row["actions"] = raw_actions
        steps.append(step_row)

    flow_state = {
        "controller_label": args.controller_label,
        "profile_id": target.get("profile_id"),
        "initialized_at": _utc_timestamp(),
        "steps": steps,
    }
    flow_state_path.parent.mkdir(parents=True, exist_ok=True)
    flow_state_path.write_text(json.dumps(flow_state, indent=2), encoding="utf-8")
    _append_event(
        logs_path,
        "flow_initialized",
        {
            "controller_label": args.controller_label,
            "step_count": len(steps),
            "flow_state_json": str(flow_state_path.resolve()),
        },
    )
    print(
        f"flow_initialized=true controller_label={args.controller_label} "
        f"steps={len(steps)}"
    )
    return 0


def cmd_list_flows(args: argparse.Namespace) -> int:
    run_dir = args.run_dir
    logs_path = run_dir / "logs" / "events.jsonl"
    flows_root = _flows_dir(run_dir)
    flows: list[dict] = []

    if flows_root.is_dir():
        for path in sorted(flows_root.glob("*.json")):
            try:
                data = json.loads(path.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError):
                continue
            label = str(data.get("controller_label", "")).strip()
            if not label:
                label = path.stem
            steps = data.get("steps", [])
            if not isinstance(steps, list):
                steps = []
            flows.append(
                {
                    "controller_label": label,
                    "profile_id": data.get("profile_id"),
                    "flow_state_json": str(path.resolve()),
                    "step_count": len(steps),
                    "status_counts": _step_status_counts(steps),
                }
            )

    payload = {"flow_count": len(flows), "flows": flows}
    _append_event(
        logs_path,
        "flows_listed",
        {
            "flow_count": len(flows),
            "controller_labels": [row["controller_label"] for row in flows],
        },
    )
    print(json.dumps(payload, sort_keys=True))
    return 0


def cmd_show_flow(args: argparse.Namespace) -> int:
    run_dir = args.run_dir
    logs_path = run_dir / "logs" / "events.jsonl"
    flow_state_path = _flow_state_path(run_dir, args.controller_label)
    if not flow_state_path.is_file():
        print(
            f"error: flow state not found for controller_label={args.controller_label}"
        )
        return 2

    flow_state = json.loads(flow_state_path.read_text(encoding="utf-8"))
    _append_event(
        logs_path,
        "flow_viewed",
        {
            "controller_label": args.controller_label,
            "flow_state_json": str(flow_state_path.resolve()),
        },
    )
    print(json.dumps(flow_state, sort_keys=True))
    return 0


def cmd_commissioning_guided_next(args: argparse.Namespace) -> int:
    """Print compact commissioning guidance: next step + step list (CLI guided slice)."""
    run_dir = args.run_dir
    logs_path = run_dir / "logs" / "events.jsonl"
    flow_state_path = _flow_state_path(run_dir, args.controller_label)
    if not flow_state_path.is_file():
        print(
            "error: commissioning flow not initialized; run init-flow first "
            f"(missing {flow_state_path})"
        )
        return 2

    flow_state = json.loads(flow_state_path.read_text(encoding="utf-8"))
    steps = flow_state.get("steps", [])
    if not isinstance(steps, list):
        steps = []
    session = _load_session_state(run_dir, args.controller_label)
    session_vals = _session_values_map(session) if session else {}

    guidance_base = _compact_flow_guidance(steps)
    guidance = _enrich_guidance_for_operator_json(guidance_base, steps, session_vals)

    payload: dict[str, Any] = {
        "controller_label": str(flow_state.get("controller_label", "")).strip()
        or args.controller_label,
        "profile_id": flow_state.get("profile_id"),
        "flow_state_json": str(flow_state_path.resolve()),
        "guidance": guidance,
        "session_keys": sorted(session_vals.keys()),
    }
    _append_event(
        logs_path,
        "commissioning_guided_next_viewed",
        {"controller_label": args.controller_label},
    )
    print(json.dumps(payload, indent=2, sort_keys=True))
    return 0


def cmd_set_session_value(args: argparse.Namespace) -> int:
    run_dir = args.run_dir
    logs_path = run_dir / "logs" / "events.jsonl"
    flow_state_path = _flow_state_path(run_dir, args.controller_label)
    if not flow_state_path.is_file():
        print(
            "error: commissioning flow not initialized; run init-flow first "
            f"(missing {flow_state_path})"
        )
        return 2

    key = str(args.key).strip()
    if not key:
        print("error: session key must be non-empty")
        return 2
    if len(key) > 128:
        print("error: session key exceeds maximum length (128)")
        return 2

    session_path = _session_state_path(run_dir, args.controller_label)
    session_path.parent.mkdir(parents=True, exist_ok=True)
    if session_path.is_file():
        session_state = json.loads(session_path.read_text(encoding="utf-8"))
    else:
        session_state = {
            "controller_label": args.controller_label,
            "updated_at": _utc_timestamp(),
            "values": {},
        }

    if not isinstance(session_state.get("values"), dict):
        session_state["values"] = {}

    session_state["values"][key] = {
        "value": str(args.value),
        "technician_name": args.technician_name,
        "note": args.note,
        "ts": _utc_timestamp(),
    }
    session_state["updated_at"] = _utc_timestamp()
    session_path.write_text(json.dumps(session_state, indent=2), encoding="utf-8")

    _append_event(
        logs_path,
        "session_value_set",
        {
            "controller_label": args.controller_label,
            "session_key": key,
            "session_state_json": str(session_path.resolve()),
        },
    )
    print(
        f"session_value_set=true controller_label={args.controller_label} key={key}"
    )
    return 0


def cmd_commissioning_confirm_prompt(args: argparse.Namespace) -> int:
    """Re-issue CHW valve write for a profile prompt, then record ``prompt_confirm.<id>`` in session."""
    run_dir = args.run_dir
    logs_path = run_dir / "logs" / "events.jsonl"
    runtime_job_path = run_dir / "state" / "runtime-job.json"
    flow_path = _flow_state_path(run_dir, args.controller_label)
    if not flow_path.is_file():
        print(f"error: flow state missing; run init-flow first ({flow_path})")
        return 2
    runtime_job = json.loads(runtime_job_path.read_text(encoding="utf-8"))
    target = None
    for controller in runtime_job.get("controllers", []):
        if controller.get("controller_label") == args.controller_label:
            target = controller
            break
    if target is None:
        print(f"error: controller not found in runtime job: {args.controller_label}")
        return 2

    flow_state = json.loads(flow_path.read_text(encoding="utf-8"))
    step = _lookup_step_by_id(flow_state.get("steps", []), args.step_id)
    if step is None:
        print(f"error: step_id not found in flow state: {args.step_id}")
        return 2

    prompt_id = str(args.prompt_id).strip()
    if not prompt_id:
        print("error: prompt_id must be non-empty")
        return 2

    actions = step.get("actions")
    if not isinstance(actions, list):
        print("error: step has no actions in flow state")
        return 2

    write_pct: float | None = None
    found_prompt = False
    for idx, act in enumerate(actions):
        if not isinstance(act, dict):
            continue
        if str(act.get("type", "")).strip() != "operator_prompt_confirm":
            continue
        if str(act.get("prompt_id", "")).strip() != prompt_id:
            continue
        found_prompt = True
        for j in range(idx - 1, -1, -1):
            prev = actions[j]
            if not isinstance(prev, dict):
                continue
            if str(prev.get("type", "")).strip() != "write_analog_percent":
                continue
            if str(prev.get("object_id", "")).strip() != "ao_chw_valve":
                continue
            try:
                write_pct = float(prev.get("value"))
            except (TypeError, ValueError):
                print("error: preceding write_analog_percent has invalid value")
                return 2
            break
        break

    if not found_prompt or write_pct is None:
        print(
            f"error: no operator_prompt_confirm with prompt_id={prompt_id!r} "
            "after write_analog_percent on ao_chw_valve in this step"
        )
        return 2

    arms = str(step.get("arms_test_mode_state_key", "")).strip()
    step_id = str(step.get("step_id", "")).strip()
    if arms == "chw_valve_stroke_no_plant" or step_id == "cooling_valve_stroke_no_chw":
        one = _bacnet_read_one(
            controller_label=args.controller_label,
            target=target,
            object_id="msv_test_mode",
            property_name="presentValue",
            timeout_seconds=float(args.bacnet_timeout_seconds),
            retries=int(args.bacnet_retries),
            bacnet_bind_port=int(getattr(args, "bacnet_bind_port", 0) or 0),
            apdu_timeout_override=args.apdu_timeout,
        )
        if one.get("status") != "read_ok":
            print(
                "error: BACnet read of msv_test_mode failed (required to arm "
                "chw_valve_stroke_no_plant before valve writes)"
            )
            print(json.dumps(one, indent=2, sort_keys=True))
            return 2
        try:
            msv_state = int(float(str(one.get("read", {}).get("value_str", ""))))
        except (TypeError, ValueError):
            print("error: could not parse msv_test_mode presentValue as integer state")
            return 2
        if msv_state != 6:
            print(
                f"error: msv_test_mode must be state 6 (chw_valve_stroke_no_plant); "
                f"got {msv_state} — write MSV first, then retry"
            )
            return 2

    cmd_res = _resolve_profile_object_bacnet(target, "ao_chw_valve")
    if cmd_res is None:
        print("error: ao_chw_valve not in profile objects_by_id")
        return 2
    cmd_ot, cmd_oi = cmd_res
    objects_by_id = target.get("objects_by_id", {})
    if not isinstance(objects_by_id, dict) or "ao_chw_valve" not in objects_by_id:
        print("error: ao_chw_valve missing from objects_by_id")
        return 2
    if not bool(objects_by_id["ao_chw_valve"].get("writable")):
        print("error: ao_chw_valve is not writable in profile")
        return 2

    addr = target.get("bacnet", {})
    host = str(addr.get("host", "")).strip()
    port = int(addr.get("port", 0))
    expected_instance = int(addr.get("device_instance", 0))
    bacnet_ad = _bacnet_adapter()
    target_addr = bacnet_ad.format_ipv4_target(host, port)
    try:
        bind_port = int(getattr(args, "bacnet_bind_port", 0) or 0)
    except ValueError:
        bind_port = 0
    try:
        apdu_t = bacnet_ad.commissioning_apdu_timeout_seconds(args.apdu_timeout)
    except (TypeError, ValueError) as err:
        print(f"error: invalid --apdu-timeout: {err}")
        return 2
    who_t = bacnet_ad.effective_who_is_timeout(
        float(args.bacnet_timeout_seconds), int(args.bacnet_retries)
    )

    write_res = bacnet_ad.write_present_value(
        bind_port=bind_port,
        target_address=target_addr,
        expected_device_instance=expected_instance,
        object_type=cmd_ot,
        object_instance=cmd_oi,
        value=float(write_pct),
        who_is_timeout=who_t,
        apdu_timeout=apdu_t,
    )
    if write_res.get("status") != "write_ok":
        print(json.dumps({"status": "write_failed", "write": write_res}, indent=2, sort_keys=True))
        return 2

    session_key = f"prompt_confirm.{prompt_id}"
    session_path = _session_state_path(run_dir, args.controller_label)
    session_path.parent.mkdir(parents=True, exist_ok=True)
    if session_path.is_file():
        session_state = json.loads(session_path.read_text(encoding="utf-8"))
    else:
        session_state = {
            "controller_label": args.controller_label,
            "updated_at": _utc_timestamp(),
            "values": {},
        }
    if not isinstance(session_state.get("values"), dict):
        session_state["values"] = {}
    session_state["values"][session_key] = {
        "value": "true",
        "technician_name": str(args.technician_name).strip(),
        "note": str(getattr(args, "note", "") or ""),
        "ts": _utc_timestamp(),
        "prompt_id": prompt_id,
        "step_id": args.step_id,
        "command_percent": float(write_pct),
    }
    session_state["updated_at"] = _utc_timestamp()
    session_path.write_text(json.dumps(session_state, indent=2), encoding="utf-8")

    _append_event(
        logs_path,
        "commissioning_prompt_confirmed",
        {
            "controller_label": args.controller_label,
            "step_id": args.step_id,
            "prompt_id": prompt_id,
            "session_key": session_key,
            "command_percent": float(write_pct),
            "session_state_json": str(session_path.resolve()),
        },
    )
    rep_path = _append_commissioning_report_entry(
        run_dir,
        {
            "ts": _utc_timestamp(),
            "kind": "valve_prompt_confirmation",
            "controller_label": args.controller_label,
            "step_id": args.step_id,
            "report_ref": "",
            "technician_name": str(args.technician_name).strip(),
            "note": str(getattr(args, "note", "") or ""),
            "prompt_id": prompt_id,
            "session_key": session_key,
            "command_object_id": "ao_chw_valve",
            "command_percent": float(write_pct),
        },
    )
    _append_event(
        logs_path,
        "commissioning_valve_prompt_report_appended",
        {"commissioning_report_json": str(rep_path.resolve())},
    )
    print(
        json.dumps(
            {
                "prompt_confirmed": True,
                "controller_label": args.controller_label,
                "step_id": args.step_id,
                "prompt_id": prompt_id,
                "session_key": session_key,
                "write_percent": float(write_pct),
                "commissioning_report_json": str(rep_path.resolve()),
            },
            indent=2,
            sort_keys=True,
        )
    )
    return 0


def _find_flow_step_actions(flow_state: dict, step_id: str) -> list[dict] | None:
    for item in flow_state.get("steps", []):
        if not isinstance(item, dict):
            continue
        if str(item.get("step_id", "")).strip() == str(step_id).strip():
            acts = item.get("actions")
            return acts if isinstance(acts, list) else None
    return None


def _find_action_by_type(actions: list[dict], action_type: str) -> dict | None:
    want = str(action_type).strip()
    for act in actions:
        if not isinstance(act, dict):
            continue
        if str(act.get("type", "")).strip() == want:
            return act
    return None


def cmd_commissioning_confirm_tachometer_reference(args: argparse.Namespace) -> int:
    """Read tachometer BACnet point and record ``session_key`` as operator-confirmed reference."""
    run_dir = args.run_dir
    logs_path = run_dir / "logs" / "events.jsonl"
    runtime_job_path = run_dir / "state" / "runtime-job.json"
    flow_path = _flow_state_path(run_dir, args.controller_label)
    if not flow_path.is_file():
        print(f"error: flow state missing; run init-flow first ({flow_path})")
        return 2
    runtime_job = json.loads(runtime_job_path.read_text(encoding="utf-8"))
    target = None
    for controller in runtime_job.get("controllers", []):
        if controller.get("controller_label") == args.controller_label:
            target = controller
            break
    if target is None:
        print(f"error: controller not found in runtime job: {args.controller_label}")
        return 2

    flow_state = json.loads(flow_path.read_text(encoding="utf-8"))
    actions = _find_flow_step_actions(flow_state, args.step_id)
    if actions is None:
        print(f"error: step_id not found in flow state: {args.step_id}")
        return 2
    act = _find_action_by_type(actions, "operator_confirm_tachometer_reference")
    if act is None:
        print(
            "error: step has no operator_confirm_tachometer_reference action "
            "(wrong step_id or profile)"
        )
        return 2
    read_oid = str(act.get("read_object_id", "")).strip()
    session_key = str(act.get("session_key", "")).strip()
    if not read_oid or not session_key:
        print("error: operator_confirm_tachometer_reference missing read_object_id or session_key")
        return 2

    one = _bacnet_read_one(
        controller_label=args.controller_label,
        target=target,
        object_id=read_oid,
        property_name="presentValue",
        timeout_seconds=float(args.bacnet_timeout_seconds),
        retries=int(args.bacnet_retries),
        bacnet_bind_port=int(getattr(args, "bacnet_bind_port", 0) or 0),
        apdu_timeout_override=args.apdu_timeout,
    )
    if one.get("status") != "read_ok":
        print("error: BACnet read of tachometer object failed")
        print(json.dumps(one, indent=2, sort_keys=True))
        return 2
    read_block = one.get("read") if isinstance(one.get("read"), dict) else {}
    value_str = str(read_block.get("value_str", "")).strip()

    session_path = _session_state_path(run_dir, args.controller_label)
    session_path.parent.mkdir(parents=True, exist_ok=True)
    if session_path.is_file():
        session_state = json.loads(session_path.read_text(encoding="utf-8"))
    else:
        session_state = {
            "controller_label": args.controller_label,
            "updated_at": _utc_timestamp(),
            "values": {},
        }
    if not isinstance(session_state.get("values"), dict):
        session_state["values"] = {}
    # ``value`` must be truthy per _session_flag_truthy (numeric BACnet text alone is not).
    session_state["values"][session_key] = {
        "value": "true",
        "reading_value_str": value_str,
        "technician_name": str(args.technician_name).strip(),
        "note": str(getattr(args, "note", "") or ""),
        "ts": _utc_timestamp(),
        "step_id": args.step_id,
        "read_object_id": read_oid,
        "confirmed": "true",
    }
    session_state["updated_at"] = _utc_timestamp()
    session_path.write_text(json.dumps(session_state, indent=2), encoding="utf-8")

    _append_event(
        logs_path,
        "tachometer_reference_confirmed",
        {
            "controller_label": args.controller_label,
            "step_id": args.step_id,
            "session_key": session_key,
            "read_object_id": read_oid,
            "reading_value_str": value_str,
            "session_state_json": str(session_path.resolve()),
        },
    )
    rep_path = _append_commissioning_report_entry(
        run_dir,
        {
            "ts": _utc_timestamp(),
            "kind": "tachometer_reference_confirmation",
            "controller_label": args.controller_label,
            "step_id": args.step_id,
            "report_ref": "",
            "technician_name": str(args.technician_name).strip(),
            "note": str(getattr(args, "note", "") or ""),
            "session_key": session_key,
            "read_object_id": read_oid,
            "reading_value_str": value_str,
        },
    )
    _append_event(
        logs_path,
        "commissioning_tachometer_report_appended",
        {"commissioning_report_json": str(rep_path.resolve())},
    )
    print(
        json.dumps(
            {
                "tachometer_reference_confirmed": True,
                "controller_label": args.controller_label,
                "step_id": args.step_id,
                "session_key": session_key,
                "read_object_id": read_oid,
                "reading_value_str": value_str,
                "commissioning_report_json": str(rep_path.resolve()),
            },
            indent=2,
            sort_keys=True,
        )
    )
    return 0


def cmd_commissioning_airflow_adjust_write(args: argparse.Namespace) -> int:
    """Write fan / actuator command percent for an ``automatic_airflow_adjustment`` profile step."""
    run_dir = args.run_dir
    logs_path = run_dir / "logs" / "events.jsonl"
    runtime_job_path = run_dir / "state" / "runtime-job.json"
    flow_path = _flow_state_path(run_dir, args.controller_label)
    if not flow_path.is_file():
        print(f"error: flow state missing; run init-flow first ({flow_path})")
        return 2
    runtime_job = json.loads(runtime_job_path.read_text(encoding="utf-8"))
    target = None
    for controller in runtime_job.get("controllers", []):
        if controller.get("controller_label") == args.controller_label:
            target = controller
            break
    if target is None:
        print(f"error: controller not found in runtime job: {args.controller_label}")
        return 2

    flow_state = json.loads(flow_path.read_text(encoding="utf-8"))
    actions = _find_flow_step_actions(flow_state, args.step_id)
    if actions is None:
        print(f"error: step_id not found in flow state: {args.step_id}")
        return 2
    act = _find_action_by_type(actions, "automatic_airflow_adjustment")
    if act is None:
        print(
            "error: step has no automatic_airflow_adjustment action "
            "(wrong step_id or profile)"
        )
        return 2
    actuator_oid = str(act.get("actuator_object_id", "")).strip()
    if not actuator_oid:
        print("error: automatic_airflow_adjustment missing actuator_object_id")
        return 2

    try:
        pct = float(args.fan_command_percent)
    except (TypeError, ValueError):
        print("error: --fan-command-percent must be a number")
        return 2
    if pct < 0.0 or pct > 100.0:
        print("error: --fan-command-percent must be between 0 and 100")
        return 2

    step_row = _lookup_step_by_id(flow_state.get("steps", []), args.step_id)
    arms = str(step_row.get("arms_test_mode_state_key", "")).strip() if step_row else ""
    if arms == "airflow_verify":
        one = _bacnet_read_one(
            controller_label=args.controller_label,
            target=target,
            object_id="msv_test_mode",
            property_name="presentValue",
            timeout_seconds=float(args.bacnet_timeout_seconds),
            retries=int(args.bacnet_retries),
            bacnet_bind_port=int(getattr(args, "bacnet_bind_port", 0) or 0),
            apdu_timeout_override=args.apdu_timeout,
        )
        if one.get("status") != "read_ok":
            print(
                "error: BACnet read of msv_test_mode failed (required when step arms airflow_verify)"
            )
            print(json.dumps(one, indent=2, sort_keys=True))
            return 2
        try:
            msv_state = int(float(str(one.get("read", {}).get("value_str", ""))))
        except (TypeError, ValueError):
            print("error: could not parse msv_test_mode presentValue as integer state")
            return 2
        if msv_state != 3:
            print(
                f"error: msv_test_mode must be state 3 (airflow_verify) for this step; got {msv_state}"
            )
            return 2

    cmd_res = _resolve_profile_object_bacnet(target, actuator_oid)
    if cmd_res is None:
        print(f"error: {actuator_oid!r} not in profile objects_by_id")
        return 2
    cmd_ot, cmd_oi = cmd_res
    objects_by_id = target.get("objects_by_id", {})
    if not isinstance(objects_by_id, dict) or actuator_oid not in objects_by_id:
        print(f"error: {actuator_oid!r} missing from objects_by_id")
        return 2
    if not bool(objects_by_id[actuator_oid].get("writable")):
        print(f"error: {actuator_oid!r} is not writable in profile")
        return 2
    w_allow = target.get("commissioning_write_allowlist", [])
    if not isinstance(w_allow, list) or actuator_oid not in {
        str(x).strip() for x in w_allow if str(x).strip()
    }:
        print(
            f"error: {actuator_oid!r} not in commissioning_write_allowlist "
            f"(add it to the unit profile for this site)"
        )
        return 2

    addr = target.get("bacnet", {})
    host = str(addr.get("host", "")).strip()
    port = int(addr.get("port", 0))
    expected_instance = int(addr.get("device_instance", 0))
    bacnet_ad = _bacnet_adapter()
    target_addr = bacnet_ad.format_ipv4_target(host, port)
    try:
        bind_port = int(getattr(args, "bacnet_bind_port", 0) or 0)
    except ValueError:
        bind_port = 0
    try:
        apdu_t = bacnet_ad.commissioning_apdu_timeout_seconds(args.apdu_timeout)
    except (TypeError, ValueError) as err:
        print(f"error: invalid --apdu-timeout: {err}")
        return 2
    who_t = bacnet_ad.effective_who_is_timeout(
        float(args.bacnet_timeout_seconds), int(args.bacnet_retries)
    )

    write_res = bacnet_ad.write_present_value(
        bind_port=bind_port,
        target_address=target_addr,
        expected_device_instance=expected_instance,
        object_type=cmd_ot,
        object_instance=cmd_oi,
        value=float(pct),
        who_is_timeout=who_t,
        apdu_timeout=apdu_t,
    )
    if write_res.get("status") != "write_ok":
        print(json.dumps({"status": "write_failed", "write": write_res}, indent=2, sort_keys=True))
        return 2

    meta = target.get("commissioning_meta") if isinstance(target.get("commissioning_meta"), dict) else {}
    unit_specs = meta.get("unit_specs") if isinstance(meta.get("unit_specs"), dict) else {}
    design_flow = unit_specs.get("design_supply_airflow_L_s")
    try:
        ratio = float(act.get("target_flow_ratio_of_design", 0.5))
    except (TypeError, ValueError):
        ratio = 0.5

    _append_event(
        logs_path,
        "airflow_adjust_command_written",
        {
            "controller_label": args.controller_label,
            "step_id": args.step_id,
            "actuator_object_id": actuator_oid,
            "fan_command_percent": float(pct),
            "target_flow_ratio_of_design": ratio,
            "design_supply_airflow_L_s": design_flow,
        },
    )
    dflow_out: float | str = ""
    if design_flow is not None:
        try:
            dflow_out = float(design_flow)
        except (TypeError, ValueError):
            dflow_out = str(design_flow)
    rep_path = _append_commissioning_report_entry(
        run_dir,
        {
            "ts": _utc_timestamp(),
            "kind": "airflow_adjust_command",
            "controller_label": args.controller_label,
            "step_id": args.step_id,
            "report_ref": "",
            "technician_name": str(args.technician_name).strip(),
            "note": str(getattr(args, "note", "") or ""),
            "actuator_object_id": actuator_oid,
            "command_percent": float(pct),
            "target_flow_ratio_of_design": ratio,
            "design_supply_airflow_L_s": dflow_out,
        },
    )
    _append_event(
        logs_path,
        "commissioning_airflow_adjust_report_appended",
        {"commissioning_report_json": str(rep_path.resolve())},
    )
    print(
        json.dumps(
            {
                "airflow_adjust_written": True,
                "controller_label": args.controller_label,
                "step_id": args.step_id,
                "actuator_object_id": actuator_oid,
                "fan_command_percent": float(pct),
                "target_flow_ratio_of_design": ratio,
                "design_supply_airflow_L_s": design_flow,
                "commissioning_report_json": str(rep_path.resolve()),
            },
            indent=2,
            sort_keys=True,
        )
    )
    return 0


def _runtime_controller_row(runtime_job: dict, label: str) -> dict | None:
    for c in runtime_job.get("controllers", []) or []:
        if isinstance(c, dict) and str(c.get("controller_label", "")).strip() == label:
            return c
    return None


def cmd_commissioning_airflow_closed_loop_iterate(args: argparse.Namespace) -> int:
    """Iterative fan % adjustments toward target L/s using BACnet flow feedback (Tier B1).

    Profile ``automatic_airflow_adjustment`` may include optional ``closed_loop``::

        {
          "enabled": true,
          "flow_read_object_id": "av_supply_airflow_present",
          "flow_read_controller_label": "HRV-01",
          "target_flow_ratio_of_design": 0.5,
          "tolerance_L_s": 0.05,
          "max_iterations": 6,
          "gain": 0.15,
          "min_command_percent": 15.0,
          "max_command_percent": 95.0,
          "initial_command_percent": 50.0
        }

    ``flow_read_controller_label`` defaults to ``--controller-label`` when omitted.
    """
    run_dir = args.run_dir
    logs_path = run_dir / "logs" / "events.jsonl"
    runtime_job_path = run_dir / "state" / "runtime-job.json"
    flow_path = _flow_state_path(run_dir, args.controller_label)
    if not flow_path.is_file():
        print(f"error: flow state missing; run init-flow first ({flow_path})")
        return 2
    runtime_job = json.loads(runtime_job_path.read_text(encoding="utf-8"))
    target = _runtime_controller_row(runtime_job, args.controller_label)
    if target is None:
        print(f"error: controller not found in runtime job: {args.controller_label}")
        return 2

    flow_state = json.loads(flow_path.read_text(encoding="utf-8"))
    actions = _find_flow_step_actions(flow_state, args.step_id)
    if actions is None:
        print(f"error: step_id not found in flow state: {args.step_id}")
        return 2
    act = _find_action_by_type(actions, "automatic_airflow_adjustment")
    if act is None:
        print("error: step has no automatic_airflow_adjustment action")
        return 2
    cl = act.get("closed_loop")
    if not isinstance(cl, dict) or not bool(cl.get("enabled")):
        print("error: profile automatic_airflow_adjustment.closed_loop.enabled is not true")
        return 2

    flow_oid = str(cl.get("flow_read_object_id", "")).strip()
    if not flow_oid:
        print("error: closed_loop.flow_read_object_id is required")
        return 2
    read_label = str(cl.get("flow_read_controller_label", "") or "").strip() or args.controller_label
    read_target = _runtime_controller_row(runtime_job, read_label)
    if read_target is None:
        print(f"error: flow_read_controller_label not in job: {read_label!r}")
        return 2

    try:
        ratio = float(act.get("target_flow_ratio_of_design", cl.get("target_flow_ratio_of_design", 0.5)))
    except (TypeError, ValueError):
        ratio = 0.5
    meta = target.get("commissioning_meta") if isinstance(target.get("commissioning_meta"), dict) else {}
    unit_specs = meta.get("unit_specs") if isinstance(meta.get("unit_specs"), dict) else {}
    design_flow = unit_specs.get("design_supply_airflow_L_s")
    try:
        design_f = float(design_flow)
    except (TypeError, ValueError):
        print("error: commissioning_meta.unit_specs.design_supply_airflow_L_s required as number for closed loop")
        return 2
    target_L = design_f * ratio

    try:
        tol = float(cl.get("tolerance_L_s", 0.08))
    except (TypeError, ValueError):
        tol = 0.08
    try:
        max_iter = int(cl.get("max_iterations", 8))
    except (TypeError, ValueError):
        max_iter = 8
    try:
        gain = float(cl.get("gain", 0.12))
    except (TypeError, ValueError):
        gain = 0.12
    try:
        min_pct = float(cl.get("min_command_percent", 10.0))
    except (TypeError, ValueError):
        min_pct = 10.0
    try:
        max_pct = float(cl.get("max_command_percent", 100.0))
    except (TypeError, ValueError):
        max_pct = 100.0
    if getattr(args, "initial_fan_command_percent", None) is not None:
        pct = float(args.initial_fan_command_percent)
    else:
        try:
            pct = float(cl.get("initial_command_percent", 50.0))
        except (TypeError, ValueError):
            pct = 50.0
    pct = max(min_pct, min(max_pct, pct))

    actuator_oid = str(act.get("actuator_object_id", "")).strip()
    if not actuator_oid:
        print("error: automatic_airflow_adjustment missing actuator_object_id")
        return 2

    step_row = _lookup_step_by_id(flow_state.get("steps", []), args.step_id)
    arms = str(step_row.get("arms_test_mode_state_key", "")).strip() if step_row else ""
    if arms == "airflow_verify":
        one = _bacnet_read_one(
            controller_label=args.controller_label,
            target=target,
            object_id="msv_test_mode",
            property_name="presentValue",
            timeout_seconds=float(args.bacnet_timeout_seconds),
            retries=int(args.bacnet_retries),
            bacnet_bind_port=int(getattr(args, "bacnet_bind_port", 0) or 0),
            apdu_timeout_override=args.apdu_timeout,
        )
        if one.get("status") != "read_ok":
            print("error: BACnet read of msv_test_mode failed (airflow_verify)")
            print(json.dumps(one, indent=2, sort_keys=True))
            return 2
        try:
            msv_state = int(float(str(one.get("read", {}).get("value_str", ""))))
        except (TypeError, ValueError):
            print("error: could not parse msv_test_mode presentValue")
            return 2
        if msv_state != 3:
            print(f"error: msv_test_mode must be state 3; got {msv_state}")
            return 2

    cmd_res = _resolve_profile_object_bacnet(target, actuator_oid)
    if cmd_res is None:
        print(f"error: {actuator_oid!r} not in profile objects_by_id")
        return 2
    cmd_ot, cmd_oi = cmd_res
    objects_by_id = target.get("objects_by_id", {})
    if not isinstance(objects_by_id, dict) or actuator_oid not in objects_by_id:
        print(f"error: {actuator_oid!r} missing from objects_by_id")
        return 2
    if not bool(objects_by_id[actuator_oid].get("writable")):
        print(f"error: {actuator_oid!r} is not writable")
        return 2
    w_allow = target.get("commissioning_write_allowlist", [])
    if not isinstance(w_allow, list) or actuator_oid not in {
        str(x).strip() for x in w_allow if str(x).strip()
    }:
        print(f"error: {actuator_oid!r} not in commissioning_write_allowlist")
        return 2

    r_allow = read_target.get("commissioning_read_allowlist", [])
    if not isinstance(r_allow, list) or flow_oid not in {str(x).strip() for x in r_allow if str(x).strip()}:
        print(
            f"error: flow_read_object_id {flow_oid!r} not on read allowlist of controller {read_label!r}"
        )
        return 2
    if flow_oid not in (read_target.get("objects_by_id") or {}):
        print(f"error: {flow_oid!r} not in objects_by_id for {read_label!r}")
        return 2

    addr = target.get("bacnet", {})
    host = str(addr.get("host", "")).strip()
    port = int(addr.get("port", 0))
    expected_instance = int(addr.get("device_instance", 0))
    bacnet_ad = _bacnet_adapter()
    target_addr = bacnet_ad.format_ipv4_target(host, port)
    try:
        bind_port = int(getattr(args, "bacnet_bind_port", 0) or 0)
    except ValueError:
        bind_port = 0
    try:
        apdu_t = bacnet_ad.commissioning_apdu_timeout_seconds(args.apdu_timeout)
    except (TypeError, ValueError) as err:
        print(f"error: invalid --apdu-timeout: {err}")
        return 2
    who_t = bacnet_ad.effective_who_is_timeout(
        float(args.bacnet_timeout_seconds), int(args.bacnet_retries)
    )

    iterations: list[dict[str, Any]] = []
    ok = False
    last_flow: float | None = None
    for i in range(max(1, max_iter)):
        write_res = bacnet_ad.write_present_value(
            bind_port=bind_port,
            target_address=target_addr,
            expected_device_instance=expected_instance,
            object_type=cmd_ot,
            object_instance=cmd_oi,
            value=float(pct),
            who_is_timeout=who_t,
            apdu_timeout=apdu_t,
        )
        if write_res.get("status") != "write_ok":
            iterations.append(
                {"iteration": i + 1, "command_percent": pct, "write": write_res, "flow_read": None}
            )
            print(json.dumps({"ok": False, "reason": "write_failed", "iterations": iterations}, indent=2))
            return 2

        fr = _bacnet_read_one(
            controller_label=read_label,
            target=read_target,
            object_id=flow_oid,
            property_name="presentValue",
            timeout_seconds=float(args.bacnet_timeout_seconds),
            retries=int(args.bacnet_retries),
            bacnet_bind_port=bind_port,
            apdu_timeout_override=args.apdu_timeout,
        )
        flow_val: float | None = None
        if fr.get("status") == "read_ok":
            try:
                flow_val = float(str(fr.get("read", {}).get("value_str", "")))
            except (TypeError, ValueError):
                flow_val = None
        last_flow = flow_val
        iterations.append(
            {
                "iteration": i + 1,
                "command_percent": pct,
                "measured_flow_L_s": flow_val,
                "target_flow_L_s": target_L,
                "read_controller": read_label,
                "read_object_id": flow_oid,
            }
        )
        if flow_val is None:
            print(json.dumps({"ok": False, "reason": "flow_read_failed", "iterations": iterations}, indent=2))
            return 2
        if abs(flow_val - target_L) <= tol:
            ok = True
            break
        adj = gain * (target_L - flow_val)
        pct = max(min_pct, min(max_pct, pct + adj))

    _append_event(
        logs_path,
        "airflow_closed_loop_iterate",
        {
            "controller_label": args.controller_label,
            "step_id": args.step_id,
            "iterations": len(iterations),
            "converged": ok,
        },
    )
    out = {
        "ok": ok,
        "controller_label": args.controller_label,
        "step_id": args.step_id,
        "target_flow_L_s": target_L,
        "tolerance_L_s": tol,
        "last_measured_flow_L_s": last_flow,
        "final_command_percent": pct,
        "iterations": iterations,
    }
    print(json.dumps(out, indent=2, sort_keys=True))
    return 0 if ok else 2


def cmd_operator_gui(args: argparse.Namespace) -> int:
    """Local browser UI for common commissioning CLI actions (Tier B2)."""
    spec = importlib.util.spec_from_file_location(
        "operator_gui_server_mod", OPERATOR_GUI_SERVER
    )
    if spec is None or spec.loader is None:
        print(f"error: cannot load {OPERATOR_GUI_SERVER}")
        return 2
    mod = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = mod
    spec.loader.exec_module(mod)
    run_dir: Path = args.run_dir
    host = str(getattr(args, "gui_host", "127.0.0.1") or "127.0.0.1")
    port = int(getattr(args, "gui_port", 8765) or 8765)
    mod.run_operator_gui_server(run_dir=run_dir, host=host, port=port)
    return 0


def cmd_commissioning_record_manual_airflow(args: argparse.Namespace) -> int:
    """Record measured airflow (L/s) for ``manual_airflow_verification_assisted`` profile step."""
    run_dir = args.run_dir
    logs_path = run_dir / "logs" / "events.jsonl"
    runtime_job_path = run_dir / "state" / "runtime-job.json"
    flow_path = _flow_state_path(run_dir, args.controller_label)
    if not flow_path.is_file():
        print(f"error: flow state missing; run init-flow first ({flow_path})")
        return 2
    runtime_job = json.loads(runtime_job_path.read_text(encoding="utf-8"))
    target = None
    for controller in runtime_job.get("controllers", []):
        if controller.get("controller_label") == args.controller_label:
            target = controller
            break
    if target is None:
        print(f"error: controller not found in runtime job: {args.controller_label}")
        return 2

    flow_state = json.loads(flow_path.read_text(encoding="utf-8"))
    actions = _find_flow_step_actions(flow_state, args.step_id)
    if actions is None:
        print(f"error: step_id not found in flow state: {args.step_id}")
        return 2

    branch_id = str(args.branch_id).strip()
    if not branch_id:
        print("error: --branch-id must be non-empty")
        return 2

    act_found: dict | None = None
    for act in actions:
        if not isinstance(act, dict):
            continue
        if str(act.get("type", "")).strip() != "manual_airflow_verification_assisted":
            continue
        raw_branches = act.get("branch_ids")
        if not isinstance(raw_branches, list):
            continue
        branch_set = {str(b).strip() for b in raw_branches if str(b).strip()}
        if branch_id in branch_set:
            act_found = act
            break

    if act_found is None:
        print(
            "error: step has no manual_airflow_verification_assisted action "
            f"listing branch_id={branch_id!r}"
        )
        return 2

    raw_map = act_found.get("session_keys")
    overrides: dict[str, str] = {}
    if isinstance(raw_map, dict):
        for k, v in raw_map.items():
            kk = str(k).strip()
            vv = str(v).strip()
            if kk and vv:
                overrides[kk] = vv
    session_key = overrides.get(branch_id, _default_manual_airflow_session_key(branch_id))

    try:
        flow_ls = float(str(args.measured_flow_L_s).strip())
    except (TypeError, ValueError):
        print("error: --measured-flow-L-s must be a number")
        return 2
    if flow_ls <= 0.0:
        print("error: --measured-flow-L-s must be > 0")
        return 2

    tool = str(getattr(args, "measurement_tool", "") or "").strip()
    if not tool:
        print("error: --measurement-tool is required (e.g. balometer)")
        return 2

    meta = target.get("commissioning_meta") if isinstance(target.get("commissioning_meta"), dict) else {}
    av = meta.get("airflow_verification") if isinstance(meta.get("airflow_verification"), dict) else {}
    branches = av.get("branches") if isinstance(av.get("branches"), list) else []
    allowed_tools: list[str] | None = None
    design_flow_branch: float | None = None
    for br in branches:
        if not isinstance(br, dict):
            continue
        if str(br.get("id", "")).strip() != branch_id:
            continue
        try:
            d = br.get("design_flow_L_s")
            if d is not None:
                design_flow_branch = float(d)
        except (TypeError, ValueError):
            design_flow_branch = None
        meas = br.get("measurement") if isinstance(br.get("measurement"), dict) else {}
        raw_allowed = meas.get("allowed_tools")
        if isinstance(raw_allowed, list) and raw_allowed:
            allowed_tools = [str(x).strip() for x in raw_allowed if str(x).strip()]
        break

    if allowed_tools is not None and tool not in allowed_tools:
        print(
            f"error: measurement_tool {tool!r} not in profile allowed_tools "
            f"for branch {branch_id!r}: {allowed_tools}"
        )
        return 2

    step_row = _lookup_step_by_id(flow_state.get("steps", []), args.step_id)
    arms = str(step_row.get("arms_test_mode_state_key", "")).strip() if step_row else ""
    if arms == "airflow_verify":
        one = _bacnet_read_one(
            controller_label=args.controller_label,
            target=target,
            object_id="msv_test_mode",
            property_name="presentValue",
            timeout_seconds=float(args.bacnet_timeout_seconds),
            retries=int(args.bacnet_retries),
            bacnet_bind_port=int(getattr(args, "bacnet_bind_port", 0) or 0),
            apdu_timeout_override=args.apdu_timeout,
        )
        if one.get("status") != "read_ok":
            print(
                "error: BACnet read of msv_test_mode failed (required when step arms airflow_verify)"
            )
            print(json.dumps(one, indent=2, sort_keys=True))
            return 2
        try:
            msv_state = int(float(str(one.get("read", {}).get("value_str", ""))))
        except (TypeError, ValueError):
            print("error: could not parse msv_test_mode presentValue as integer state")
            return 2
        if msv_state != 3:
            print(
                f"error: msv_test_mode must be state 3 (airflow_verify); got {msv_state}"
            )
            return 2

    session_path = _session_state_path(run_dir, args.controller_label)
    session_path.parent.mkdir(parents=True, exist_ok=True)
    if session_path.is_file():
        session_state = json.loads(session_path.read_text(encoding="utf-8"))
    else:
        session_state = {
            "controller_label": args.controller_label,
            "updated_at": _utc_timestamp(),
            "values": {},
        }
    if not isinstance(session_state.get("values"), dict):
        session_state["values"] = {}

    value_str = str(flow_ls)
    session_state["values"][session_key] = {
        "value": value_str,
        "technician_name": str(args.technician_name).strip(),
        "note": str(getattr(args, "note", "") or ""),
        "ts": _utc_timestamp(),
        "step_id": args.step_id,
        "branch_id": branch_id,
        "measurement_tool": tool,
        "design_flow_L_s": design_flow_branch,
    }
    session_state["updated_at"] = _utc_timestamp()
    session_path.write_text(json.dumps(session_state, indent=2), encoding="utf-8")

    _append_event(
        logs_path,
        "manual_airflow_recorded",
        {
            "controller_label": args.controller_label,
            "step_id": args.step_id,
            "branch_id": branch_id,
            "session_key": session_key,
            "measured_flow_L_s": flow_ls,
            "measurement_tool": tool,
            "session_state_json": str(session_path.resolve()),
        },
    )
    report_entry = {
        "ts": _utc_timestamp(),
        "kind": "manual_airflow_measurement",
        "controller_label": args.controller_label,
        "step_id": args.step_id,
        "report_ref": "",
        "technician_name": str(args.technician_name).strip(),
        "note": str(getattr(args, "note", "") or ""),
        "branch_id": branch_id,
        "session_key": session_key,
        "measured_flow_L_s": float(flow_ls),
        "measurement_tool": tool,
        "design_flow_L_s": design_flow_branch,
    }
    report_path = _append_commissioning_report_entry(run_dir, report_entry)
    _append_event(
        logs_path,
        "commissioning_manual_airflow_report_appended",
        {"commissioning_report_json": str(report_path.resolve())},
    )
    print(
        json.dumps(
            {
                "manual_airflow_recorded": True,
                "controller_label": args.controller_label,
                "step_id": args.step_id,
                "branch_id": branch_id,
                "session_key": session_key,
                "measured_flow_L_s": flow_ls,
                "measurement_tool": tool,
                "commissioning_report_json": str(report_path.resolve()),
            },
            indent=2,
            sort_keys=True,
        )
    )
    return 0


def cmd_show_session(args: argparse.Namespace) -> int:
    run_dir = args.run_dir
    logs_path = run_dir / "logs" / "events.jsonl"
    session_path = _session_state_path(run_dir, args.controller_label)
    if not session_path.is_file():
        print(
            f"error: session state not found for controller_label={args.controller_label}"
        )
        return 2

    session_state = json.loads(session_path.read_text(encoding="utf-8"))
    _append_event(
        logs_path,
        "session_viewed",
        {
            "controller_label": args.controller_label,
            "session_state_json": str(session_path.resolve()),
        },
    )
    print(json.dumps(session_state, sort_keys=True))
    return 0


def cmd_export_run_summary(args: argparse.Namespace) -> int:
    run_dir = args.run_dir
    logs_path = run_dir / "logs" / "events.jsonl"
    output_json = args.output_json or (run_dir / "artifacts" / "run-summary.json")
    config = _parse_run_config(run_dir)
    runtime_job_path = run_dir / "state" / "runtime-job.json"
    if not runtime_job_path.is_file():
        print(
            f"error: runtime job missing at {runtime_job_path}; run compile-import first"
        )
        return 2

    runtime_job = json.loads(runtime_job_path.read_text(encoding="utf-8"))
    import_report_path = run_dir / "state" / "import-report.json"
    import_report = None
    if import_report_path.is_file():
        import_report = json.loads(import_report_path.read_text(encoding="utf-8"))

    bip_summary_path = run_dir / "artifacts" / "bip" / "list-summary.json"
    bip_list_summary = None
    if bip_summary_path.is_file():
        bip_list_summary = json.loads(bip_summary_path.read_text(encoding="utf-8"))

    controllers_out: list[dict] = []
    for row in runtime_job.get("controllers", []):
        label = str(row.get("controller_label", "")).strip()
        flow_path = _flow_state_path(run_dir, label)
        entry: dict = {
            "controller_label": label,
            "profile_id": row.get("profile_id"),
            "flow_initialized": flow_path.is_file(),
            "flow_state_json": str(flow_path.resolve()) if flow_path.is_file() else None,
            "step_count": None,
            "status_counts": None,
            "next_open_step": None,
        }
        if flow_path.is_file():
            try:
                flow_state = json.loads(flow_path.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError):
                flow_state = {}
            steps = flow_state.get("steps", [])
            if not isinstance(steps, list):
                steps = []
            entry["step_count"] = len(steps)
            entry["status_counts"] = _step_status_counts(steps)
            entry["next_open_step"] = _next_open_step(steps)
        controllers_out.append(entry)

    summary = {
        "schema_version": "0.1-run-summary",
        "generated_at": _utc_timestamp(),
        "job_id": config.get("job_id"),
        "run_dir": str(run_dir.resolve()),
        "runtime_job_json": str(runtime_job_path.resolve()),
        "import_report_present": import_report is not None,
        "import_compile_ok": import_report.get("compile_ok") if import_report else None,
        "bip_list_summary_present": bip_list_summary is not None,
        "controllers": controllers_out,
    }
    if bool(getattr(args, "embed_import_report", False)) and import_report is not None:
        summary["import_report"] = import_report
    if bool(getattr(args, "embed_bip_list_summary", False)) and bip_list_summary is not None:
        summary["bip_list_summary"] = bip_list_summary

    output_json.parent.mkdir(parents=True, exist_ok=True)
    output_json.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    csv_path = getattr(args, "output_csv", None)
    if csv_path:
        csv_path.parent.mkdir(parents=True, exist_ok=True)
        fieldnames = [
            "controller_label",
            "profile_id",
            "flow_initialized",
            "step_count",
            "next_step_id",
            "next_step_status",
            "pending_count",
            "passed_count",
            "failed_count",
            "skipped_count",
            "manual_passed_count",
        ]
        with csv_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            for row in controllers_out:
                counts = row.get("status_counts") or {}
                nxt = row.get("next_open_step") or {}
                writer.writerow(
                    {
                        "controller_label": row.get("controller_label", ""),
                        "profile_id": row.get("profile_id", ""),
                        "flow_initialized": str(bool(row.get("flow_initialized"))).lower(),
                        "step_count": row.get("step_count") if row.get("step_count") is not None else "",
                        "next_step_id": nxt.get("step_id", "") if isinstance(nxt, dict) else "",
                        "next_step_status": nxt.get("status", "") if isinstance(nxt, dict) else "",
                        "pending_count": counts.get("pending", 0),
                        "passed_count": counts.get("passed", 0),
                        "failed_count": counts.get("failed", 0),
                        "skipped_count": counts.get("skipped", 0),
                        "manual_passed_count": counts.get("manual_passed", 0),
                    }
                )

    _append_event(
        logs_path,
        "run_summary_exported",
        {
            "summary_json": str(output_json.resolve()),
            "csv_path": str(csv_path.resolve()) if csv_path else None,
        },
    )
    print(f"run_summary_exported=true summary_json={output_json.resolve()}")
    if csv_path:
        print(f"run_summary_csv=true csv_path={csv_path.resolve()}")
    return 0


def cmd_export_commissioning_report(args: argparse.Namespace) -> int:
    """Print or copy the append-only commissioning report JSON for this run."""
    run_dir = args.run_dir
    logs_path = run_dir / "logs" / "events.jsonl"
    src = _commissioning_report_path(run_dir)
    out_path = getattr(args, "output_json", None)
    want_csv = bool(getattr(args, "output_csv", None))
    want_csv_unified = bool(getattr(args, "output_csv_unified", None))
    want_html = bool(getattr(args, "output_html", None))
    want_xlsx = bool(getattr(args, "output_xlsx", None))
    want_pdf = bool(getattr(args, "output_pdf", None))
    want_customer_html = bool(getattr(args, "output_customer_html", None))
    want_customer_pdf = bool(getattr(args, "output_customer_pdf", None))
    xlsx_extra_mod = bool(getattr(args, "xlsx_include_modulation", False))
    allow_empty = bool(getattr(args, "allow_empty", False))

    if not src.is_file():
        if allow_empty and (
            out_path
            or want_csv
            or want_csv_unified
            or want_html
            or want_xlsx
            or want_pdf
            or want_customer_html
            or want_customer_pdf
        ):
            config = _parse_run_config(run_dir)
            job_id = str(config.get("job_id", "")).strip() or "unknown-job"
            doc: dict = {
                "schema_version": "0.2-commissioning-report",
                "job_id": job_id,
                "entries": [],
            }
            text = json.dumps(doc, indent=2, sort_keys=True) + "\n"
            if out_path:
                out_path = Path(out_path)
                out_path.parent.mkdir(parents=True, exist_ok=True)
                out_path.write_text(text, encoding="utf-8")
                _append_event(
                    logs_path,
                    "commissioning_report_exported_empty_stub",
                    {"output_json": str(out_path.resolve())},
                )
                print(
                    f"commissioning_report_exported=true output_json={out_path.resolve()} "
                    "stub=true entries=0"
                )
            if want_csv:
                csv_path = Path(args.output_csv)
                csv_path.parent.mkdir(parents=True, exist_ok=True)
                fieldnames = [
                    "entry_ts",
                    "kind",
                    "controller_label",
                    "step_id",
                    "report_ref",
                    "technician_name",
                    "command_object_id",
                    "command_percent",
                    "dwell_seconds",
                    "object_id",
                    "property",
                    "status",
                    "value_str",
                    "read_source",
                ]
                with csv_path.open("w", newline="", encoding="utf-8") as handle:
                    writer = csv.DictWriter(handle, fieldnames=fieldnames)
                    writer.writeheader()
                _append_event(
                    logs_path,
                    "commissioning_report_modulation_csv_exported",
                    {"csv_path": str(csv_path.resolve())},
                )
                print(
                    f"commissioning_report_modulation_csv=true csv_path={csv_path.resolve()}"
                )
            if want_csv_unified:
                csv_unified = Path(args.output_csv_unified)
                csv_unified.parent.mkdir(parents=True, exist_ok=True)
                with csv_unified.open("w", newline="", encoding="utf-8") as handle:
                    writer = csv.DictWriter(
                        handle, fieldnames=list(COMMISSIONING_REPORT_UNIFIED_FIELDNAMES)
                    )
                    writer.writeheader()
                _append_event(
                    logs_path,
                    "commissioning_report_unified_csv_exported",
                    {"csv_path": str(csv_unified.resolve())},
                )
                print(
                    f"commissioning_report_unified_csv=true csv_path={csv_unified.resolve()}"
                )
            if want_xlsx:
                try:
                    _write_commissioning_report_unified_xlsx(
                        Path(args.output_xlsx),
                        doc,
                        include_modulation_sheet=xlsx_extra_mod,
                    )
                except RuntimeError as err:
                    print(f"error: {err}")
                    return 2
                xlsx_path = Path(args.output_xlsx)
                _append_event(
                    logs_path,
                    "commissioning_report_xlsx_exported",
                    {"xlsx_path": str(xlsx_path.resolve())},
                )
                print(
                    f"commissioning_report_xlsx=true xlsx_path={xlsx_path.resolve()}"
                )
            if want_html:
                html_path = Path(args.output_html)
                html_path.parent.mkdir(parents=True, exist_ok=True)
                unified_rows = _commissioning_report_unified_csv_rows(doc)
                html_body = _commissioning_report_unified_rows_to_html(
                    job_id,
                    str(doc.get("schema_version", "")),
                    unified_rows,
                    doc=doc,
                )
                html_path.write_text(html_body, encoding="utf-8")
                _append_event(
                    logs_path,
                    "commissioning_report_html_exported",
                    {"html_path": str(html_path.resolve())},
                )
                print(
                    f"commissioning_report_html=true html_path={html_path.resolve()}"
                )
            if want_pdf:
                try:
                    logo_p = _resolve_commissioning_pdf_logo_path(
                        run_dir, getattr(args, "pdf_logo_image", None)
                    )
                    _write_commissioning_report_unified_pdf(
                        Path(args.output_pdf),
                        doc,
                        logo_image_path=logo_p,
                    )
                except RuntimeError as err:
                    print(f"error: {err}")
                    return 2
                pdf_path = Path(args.output_pdf)
                _append_event(
                    logs_path,
                    "commissioning_report_pdf_exported",
                    {"pdf_path": str(pdf_path.resolve())},
                )
                print(
                    f"commissioning_report_pdf=true pdf_path={pdf_path.resolve()}"
                )
            if want_customer_html:
                ch_path = Path(args.output_customer_html)
                ch_path.parent.mkdir(parents=True, exist_ok=True)
                mod_rows = _commissioning_report_modulation_rows(doc)
                norm: list[dict[str, str]] = []
                for r in mod_rows:
                    out_row = {k: str(r.get(k, "") or "") for k in COMMISSIONING_REPORT_CUSTOMER_MODULATION_FIELDNAMES}
                    norm.append(out_row)
                ch_body = _customer_modulation_rows_to_html(
                    job_id, str(doc.get("schema_version", "")), norm
                )
                ch_path.write_text(ch_body, encoding="utf-8")
                _append_event(
                    logs_path,
                    "commissioning_report_customer_html_exported",
                    {"html_path": str(ch_path.resolve())},
                )
                print(
                    f"commissioning_report_customer_html=true html_path={ch_path.resolve()}"
                )
            if want_customer_pdf:
                try:
                    logo_p = _resolve_commissioning_pdf_logo_path(
                        run_dir, getattr(args, "pdf_logo_image", None)
                    )
                    _write_customer_modulation_pdf(
                        Path(args.output_customer_pdf),
                        doc,
                        logo_image_path=logo_p,
                    )
                except RuntimeError as err:
                    print(f"error: {err}")
                    return 2
                cp_path = Path(args.output_customer_pdf)
                _append_event(
                    logs_path,
                    "commissioning_report_customer_pdf_exported",
                    {"pdf_path": str(cp_path.resolve())},
                )
                print(
                    f"commissioning_report_customer_pdf=true pdf_path={cp_path.resolve()}"
                )
            if (
                not out_path
                and not want_csv
                and not want_csv_unified
                and not want_html
                and not want_xlsx
                and not want_pdf
                and not want_customer_html
                and not want_customer_pdf
            ):
                print(
                    "error: --allow-empty requires --output-json and/or "
                    "--output-csv / --output-csv-unified / --output-html / "
                    "--output-xlsx / --output-pdf / --output-customer-html / "
                    "--output-customer-pdf"
                )
                return 2
            return 0
        print(
            f"error: commissioning report not found at {src}; "
            "nothing recorded yet (e.g. record-step with BACnet point checkout gate). "
            "Use --allow-empty with --output-json and/or --output-csv / "
            "--output-csv-unified / --output-html / --output-xlsx / --output-pdf / "
            "--output-customer-html / --output-customer-pdf "
            "for empty outputs."
        )
        return 2

    text = src.read_text(encoding="utf-8")
    doc = json.loads(text)
    csv_path = getattr(args, "output_csv", None)
    if csv_path:
        csv_path = Path(csv_path)
        csv_path.parent.mkdir(parents=True, exist_ok=True)
        fieldnames = [
            "entry_ts",
            "kind",
            "controller_label",
            "step_id",
            "report_ref",
            "technician_name",
            "command_object_id",
            "command_percent",
            "dwell_seconds",
            "object_id",
            "property",
            "status",
            "value_str",
            "read_source",
        ]
        with csv_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=fieldnames)
            writer.writeheader()
            for row in _commissioning_report_modulation_rows(doc):
                writer.writerow(row)
        _append_event(
            logs_path,
            "commissioning_report_modulation_csv_exported",
            {"csv_path": str(csv_path.resolve())},
        )
        print(f"commissioning_report_modulation_csv=true csv_path={csv_path.resolve()}")

    csv_unified = getattr(args, "output_csv_unified", None)
    if csv_unified:
        csv_unified = Path(csv_unified)
        csv_unified.parent.mkdir(parents=True, exist_ok=True)
        with csv_unified.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(
                handle, fieldnames=list(COMMISSIONING_REPORT_UNIFIED_FIELDNAMES)
            )
            writer.writeheader()
            for row in _commissioning_report_unified_csv_rows(doc):
                writer.writerow(row)
        _append_event(
            logs_path,
            "commissioning_report_unified_csv_exported",
            {"csv_path": str(csv_unified.resolve())},
        )
        print(
            f"commissioning_report_unified_csv=true csv_path={csv_unified.resolve()}"
        )

    xlsx_out = getattr(args, "output_xlsx", None)
    if xlsx_out:
        try:
            _write_commissioning_report_unified_xlsx(
                Path(xlsx_out),
                doc,
                include_modulation_sheet=xlsx_extra_mod,
            )
        except RuntimeError as err:
            print(f"error: {err}")
            return 2
        xlsx_path = Path(xlsx_out)
        _append_event(
            logs_path,
            "commissioning_report_xlsx_exported",
            {"xlsx_path": str(xlsx_path.resolve())},
        )
        print(f"commissioning_report_xlsx=true xlsx_path={xlsx_path.resolve()}")

    html_out = getattr(args, "output_html", None)
    if html_out:
        html_path = Path(html_out)
        html_path.parent.mkdir(parents=True, exist_ok=True)
        job_id = str(doc.get("job_id", "")).strip() or "unknown-job"
        schema_v = str(doc.get("schema_version", "")).strip()
        unified_rows = _commissioning_report_unified_csv_rows(doc)
        html_body = _commissioning_report_unified_rows_to_html(
            job_id, schema_v, unified_rows, doc=doc
        )
        html_path.write_text(html_body, encoding="utf-8")
        _append_event(
            logs_path,
            "commissioning_report_html_exported",
            {"html_path": str(html_path.resolve())},
        )
        print(f"commissioning_report_html=true html_path={html_path.resolve()}")

    pdf_out = getattr(args, "output_pdf", None)
    if pdf_out:
        try:
            logo_p = _resolve_commissioning_pdf_logo_path(
                run_dir, getattr(args, "pdf_logo_image", None)
            )
            _write_commissioning_report_unified_pdf(
                Path(pdf_out),
                doc,
                logo_image_path=logo_p,
            )
        except RuntimeError as err:
            print(f"error: {err}")
            return 2
        pdf_path = Path(pdf_out)
        _append_event(
            logs_path,
            "commissioning_report_pdf_exported",
            {"pdf_path": str(pdf_path.resolve())},
        )
        print(f"commissioning_report_pdf=true pdf_path={pdf_path.resolve()}")

    customer_html_out = getattr(args, "output_customer_html", None)
    if customer_html_out:
        html_path = Path(customer_html_out)
        html_path.parent.mkdir(parents=True, exist_ok=True)
        job_id = str(doc.get("job_id", "")).strip() or "unknown-job"
        schema_v = str(doc.get("schema_version", "")).strip()
        mod_rows = _commissioning_report_modulation_rows(doc)
        norm: list[dict[str, str]] = []
        for r in mod_rows:
            norm.append(
                {k: str(r.get(k, "") or "") for k in COMMISSIONING_REPORT_CUSTOMER_MODULATION_FIELDNAMES}
            )
        html_body = _customer_modulation_rows_to_html(job_id, schema_v, norm)
        html_path.write_text(html_body, encoding="utf-8")
        _append_event(
            logs_path,
            "commissioning_report_customer_html_exported",
            {"html_path": str(html_path.resolve())},
        )
        print(
            f"commissioning_report_customer_html=true html_path={html_path.resolve()}"
        )

    customer_pdf_out = getattr(args, "output_customer_pdf", None)
    if customer_pdf_out:
        try:
            logo_p = _resolve_commissioning_pdf_logo_path(
                run_dir, getattr(args, "pdf_logo_image", None)
            )
            _write_customer_modulation_pdf(
                Path(customer_pdf_out),
                doc,
                logo_image_path=logo_p,
            )
        except RuntimeError as err:
            print(f"error: {err}")
            return 2
        cp_path = Path(customer_pdf_out)
        _append_event(
            logs_path,
            "commissioning_report_customer_pdf_exported",
            {"pdf_path": str(cp_path.resolve())},
        )
        print(
            f"commissioning_report_customer_pdf=true pdf_path={cp_path.resolve()}"
        )

    if out_path:
        out_path = Path(out_path)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(text, encoding="utf-8")
        _append_event(
            logs_path,
            "commissioning_report_exported",
            {"output_json": str(out_path.resolve())},
        )
        print(f"commissioning_report_exported=true output_json={out_path.resolve()}")
        return 0

    print(text, end="")
    _append_event(logs_path, "commissioning_report_printed", {})
    return 0


def cmd_dry_run_bacnet_write(args: argparse.Namespace) -> int:
    run_dir = args.run_dir
    logs_path = run_dir / "logs" / "events.jsonl"
    runtime_job_path = run_dir / "state" / "runtime-job.json"
    runtime_job = json.loads(runtime_job_path.read_text(encoding="utf-8"))

    target = None
    for controller in runtime_job.get("controllers", []):
        if controller.get("controller_label") == args.controller_label:
            target = controller
            break
    if target is None:
        print(f"error: controller not found in runtime job: {args.controller_label}")
        return 2

    object_id = str(args.object_id).strip()
    profile_allow = target.get("commissioning_write_allowlist", [])
    if not isinstance(profile_allow, list) or not profile_allow:
        print(
            "error: profile has no commissioning_write_allowlist; "
            "add a non-empty array of logical object ids to the unit profile JSON"
        )
        return 2
    allowed = {str(x).strip() for x in profile_allow if str(x).strip()}
    if object_id not in allowed:
        print(
            f"error: object_id not in profile commissioning_write_allowlist: {object_id!r} "
            f"(allowed: {sorted(allowed)})"
        )
        return 2

    objects_by_id = target.get("objects_by_id", {})
    if not isinstance(objects_by_id, dict) or object_id not in objects_by_id:
        print(
            f"error: object_id not found in compiled runtime job for controller: {object_id}"
        )
        return 2
    meta = objects_by_id[object_id]
    if not isinstance(meta, dict) or not bool(meta.get("writable")):
        print(f"error: object {object_id!r} is not writable in profile")
        return 2
    bacnet = meta.get("bacnet", {})
    if not isinstance(bacnet, dict):
        print(f"error: invalid objects_by_id entry for {object_id!r}")
        return 2
    type_name = str(bacnet.get("object_type", "")).strip()
    try:
        object_instance = int(bacnet.get("instance"))
    except (TypeError, ValueError):
        print(f"error: invalid BACnet instance for {object_id!r}")
        return 2

    bacnet_ad = _bacnet_adapter()
    object_type_int = bacnet_ad.object_type_name_to_int(type_name)
    if object_type_int is None:
        print(f"error: unsupported BACnet object_type for writes: {type_name!r}")
        return 2

    addr = target.get("bacnet", {})
    host = str(addr.get("host", "")).strip()
    port = int(addr.get("port", 0))
    expected_instance = int(addr.get("device_instance", 0))

    dry_run = not bool(getattr(args, "execute", False))
    result = bacnet_ad.plan_write_property(
        host=host,
        port=port,
        expected_device_instance=expected_instance,
        object_type=object_type_int,
        object_instance=object_instance,
        property_id=bacnet_ad.present_value_property_id,
        value=int(args.value),
        timeout_seconds=args.timeout_seconds,
        retries=args.retries,
        dry_run=True,
    )
    result["controller_label"] = args.controller_label
    result["profile_object_id"] = object_id
    result["technician_name"] = args.technician_name
    result["note"] = args.note

    if not dry_run and result.get("status") == "dry_run_allowed":
        try:
            bind_port = int(getattr(args, "bacnet_bind_port", 0) or 0)
        except ValueError:
            bind_port = 0
        who_is_timeout = bacnet_ad.effective_who_is_timeout(
            args.timeout_seconds, args.retries
        )
        try:
            apdu_timeout = bacnet_ad.commissioning_apdu_timeout_seconds(args.apdu_timeout)
        except (TypeError, ValueError) as err:
            print(f"error: invalid --apdu-timeout: {err}")
            return 2
        result["bacnet_timeouts"] = {
            "who_is_timeout_seconds": who_is_timeout,
            "apdu_timeout_seconds": apdu_timeout,
        }
        try:
            exec_result = bacnet_ad.write_present_value(
                bind_port=bind_port,
                target_address=bacnet_ad.format_ipv4_target(host, port),
                expected_device_instance=expected_instance,
                object_type=object_type_int,
                object_instance=object_instance,
                value=int(args.value),
                who_is_timeout=who_is_timeout,
                apdu_timeout=apdu_timeout,
            )
        except (OSError, RuntimeError) as err:
            print(f"error: failed to load BACnet stack: {err}")
            return 2
        except ModuleNotFoundError as err:
            print(
                "error: bacpypes3 is required for --execute "
                f"(pip install -r requirements.txt): {err}"
            )
            return 2
        except Exception as err:  # noqa: BLE001 — surface client failures to operator
            result["execute_error"] = str(err)
            result["status"] = "execute_failed"
        else:
            result["execute"] = exec_result
            result["status"] = str(exec_result.get("status", "execute_failed"))
            result["dry_run"] = False

    plans_dir = run_dir / "artifacts" / "bacnet_write_plans"
    plans_dir.mkdir(parents=True, exist_ok=True)
    artifact = plans_dir / f"{args.controller_label}-{object_id}.json"
    artifact.write_text(json.dumps(result, indent=2, sort_keys=True), encoding="utf-8")

    if result.get("status") == "write_ok":
        event = "bacnet_write_executed"
    elif result.get("status") in {"dry_run_allowed", "use_bacpypes_client"}:
        event = "bacnet_write_planned"
    else:
        event = "bacnet_write_blocked"
    _append_event(
        logs_path,
        event,
        {
            "controller_label": args.controller_label,
            "object_id": object_id,
            "status": result.get("status"),
            "artifact_json": str(artifact.resolve()),
        },
    )
    print(json.dumps(result, sort_keys=True))
    if result.get("status") in {"dry_run_allowed", "write_ok"}:
        return 0
    return 2


def _bacnet_read_one(
    *,
    controller_label: str,
    target: dict,
    object_id: str,
    property_name: str,
    timeout_seconds: float,
    retries: int,
    bacnet_bind_port: int,
    apdu_timeout_override: float | None = None,
) -> dict:
    """Perform one BACnet read; returns result dict (may set status read_ok / blocked / errors)."""
    object_id = str(object_id).strip()
    profile_allow = target.get("commissioning_read_allowlist", [])
    if not isinstance(profile_allow, list) or not profile_allow:
        return {
            "controller_label": controller_label,
            "profile_object_id": object_id,
            "status": "config_error",
            "message": "profile has no commissioning_read_allowlist",
        }
    allowed = {str(x).strip() for x in profile_allow if str(x).strip()}
    if object_id not in allowed:
        return {
            "controller_label": controller_label,
            "profile_object_id": object_id,
            "status": "config_error",
            "message": f"object_id not in commissioning_read_allowlist (allowed: {sorted(allowed)})",
        }

    objects_by_id = target.get("objects_by_id", {})
    if not isinstance(objects_by_id, dict) or object_id not in objects_by_id:
        return {
            "controller_label": controller_label,
            "profile_object_id": object_id,
            "status": "config_error",
            "message": "object_id not found in objects_by_id",
        }
    meta = objects_by_id[object_id]
    if not isinstance(meta, dict):
        return {
            "controller_label": controller_label,
            "profile_object_id": object_id,
            "status": "config_error",
            "message": "invalid objects_by_id entry",
        }
    bacnet = meta.get("bacnet", {})
    if not isinstance(bacnet, dict):
        return {
            "controller_label": controller_label,
            "profile_object_id": object_id,
            "status": "config_error",
            "message": "invalid objects_by_id entry",
        }
    type_name = str(bacnet.get("object_type", "")).strip()
    try:
        object_instance = int(bacnet.get("instance"))
    except (TypeError, ValueError):
        return {
            "controller_label": controller_label,
            "profile_object_id": object_id,
            "status": "config_error",
            "message": "invalid BACnet instance",
        }

    bacnet_ad = _bacnet_adapter()
    object_type_int = bacnet_ad.object_type_name_to_int(type_name)
    if object_type_int is None:
        return {
            "controller_label": controller_label,
            "profile_object_id": object_id,
            "status": "config_error",
            "message": f"unsupported BACnet object_type: {type_name!r}",
        }

    addr = target.get("bacnet", {})
    host = str(addr.get("host", "")).strip()
    port = int(addr.get("port", 0))
    expected_instance = int(addr.get("device_instance", 0))

    probe = bacnet_ad.probe_device(
        host=host,
        port=port,
        expected_device_instance=expected_instance,
        timeout_seconds=timeout_seconds,
        retries=retries,
    )
    result: dict = {
        "controller_label": controller_label,
        "profile_object_id": object_id,
        "property": property_name,
        "probe": probe,
    }
    if probe.get("status") != "reachable_verified":
        result["status"] = "blocked_probe_failed"
        return result

    try:
        bind_port = int(bacnet_bind_port or 0)
    except ValueError:
        bind_port = 0
    who_is_timeout = bacnet_ad.effective_who_is_timeout(timeout_seconds, retries)
    try:
        apdu_timeout = bacnet_ad.commissioning_apdu_timeout_seconds(apdu_timeout_override)
    except ValueError as err:
        return {
            "controller_label": controller_label,
            "profile_object_id": object_id,
            "status": "config_error",
            "message": str(err),
        }
    result["bacnet_timeouts"] = {
        "who_is_timeout_seconds": who_is_timeout,
        "apdu_timeout_seconds": apdu_timeout,
    }
    prop = str(property_name or "presentValue").strip() or "presentValue"
    try:
        read_result = bacnet_ad.read_present_value(
            bind_port=bind_port,
            target_address=bacnet_ad.format_ipv4_target(host, port),
            expected_device_instance=expected_instance,
            object_type=object_type_int,
            object_instance=object_instance,
            property_name=prop,
            who_is_timeout=who_is_timeout,
            apdu_timeout=apdu_timeout,
        )
    except (OSError, RuntimeError) as err:
        result["status"] = "client_load_failed"
        result["message"] = str(err)
        return result
    except ModuleNotFoundError as err:
        result["status"] = "bacpypes_missing"
        result["message"] = str(err)
        return result
    except Exception as err:  # noqa: BLE001
        result["status"] = "read_failed"
        result["read_error"] = str(err)
        return result

    result["read"] = read_result
    result["status"] = str(read_result.get("status", "read_failed"))
    return result


def cmd_bacnet_read(args: argparse.Namespace) -> int:
    run_dir = args.run_dir
    logs_path = run_dir / "logs" / "events.jsonl"
    runtime_job_path = run_dir / "state" / "runtime-job.json"
    runtime_job = json.loads(runtime_job_path.read_text(encoding="utf-8"))

    target = None
    for controller in runtime_job.get("controllers", []):
        if controller.get("controller_label") == args.controller_label:
            target = controller
            break
    if target is None:
        print(f"error: controller not found in runtime job: {args.controller_label}")
        return 2

    object_id = str(args.object_id).strip()
    prop = str(args.property or "presentValue").strip() or "presentValue"
    try:
        bind_port = int(getattr(args, "bacnet_bind_port", 0) or 0)
    except ValueError:
        bind_port = 0

    try:
        _bacnet_adapter().commissioning_apdu_timeout_seconds(args.apdu_timeout)
    except (TypeError, ValueError) as err:
        print(f"error: invalid --apdu-timeout: {err}")
        return 2

    result = _bacnet_read_one(
        controller_label=args.controller_label,
        target=target,
        object_id=object_id,
        property_name=prop,
        timeout_seconds=args.timeout_seconds,
        retries=args.retries,
        bacnet_bind_port=bind_port,
        apdu_timeout_override=args.apdu_timeout,
    )
    if result.get("status") == "config_error":
        print(f"error: {result.get('message', 'invalid configuration')}")
        return 2
    if result.get("status") == "client_load_failed":
        print(f"error: failed to load BACnet client: {result.get('message')}")
        return 2
    if result.get("status") == "bacpypes_missing":
        print(
            "error: bacpypes3 is required for bacnet-read "
            f"(pip install -r requirements.txt): {result.get('message')}"
        )
        return 2

    reads_dir = run_dir / "artifacts" / "bacnet_reads"
    reads_dir.mkdir(parents=True, exist_ok=True)
    artifact = reads_dir / f"{args.controller_label}-{object_id}.json"
    artifact.write_text(json.dumps(result, indent=2, sort_keys=True), encoding="utf-8")

    event = (
        "bacnet_read_ok"
        if result.get("status") == "read_ok"
        else "bacnet_read_blocked"
    )
    _append_event(
        logs_path,
        event,
        {
            "controller_label": args.controller_label,
            "object_id": object_id,
            "status": result.get("status"),
            "artifact_json": str(artifact.resolve()),
        },
    )
    print(json.dumps(result, sort_keys=True))
    return 0 if result.get("status") == "read_ok" else 2


def _run_point_checkout_reads(
    *,
    controller_label: str,
    target: dict,
    timeout_seconds: float,
    retries: int,
    bacnet_bind_port: int,
    apdu_timeout_override: float | None,
    strict: bool,
) -> tuple[list[dict], bool]:
    """Execute profile ``point_checkout`` reads; returns (rows, all_read_ok)."""
    checkout = target.get("point_checkout", [])
    if not isinstance(checkout, list) or not checkout:
        return [], False
    rows: list[dict] = []
    all_ok = True
    for entry in checkout:
        if not isinstance(entry, dict):
            continue
        oid = str(entry.get("object_id", "")).strip()
        if not oid:
            continue
        prop = str(entry.get("property", "presentValue")).strip() or "presentValue"
        one = _bacnet_read_one(
            controller_label=controller_label,
            target=target,
            object_id=oid,
            property_name=prop,
            timeout_seconds=timeout_seconds,
            retries=retries,
            bacnet_bind_port=bacnet_bind_port,
            apdu_timeout_override=apdu_timeout_override,
        )
        rows.append(one)
        if one.get("status") != "read_ok":
            all_ok = False
            if strict:
                break
    return rows, all_ok


def _commissioning_report_path(run_dir: Path) -> Path:
    return run_dir / "artifacts" / "commissioning_report.json"


def _append_commissioning_report_entry(run_dir: Path, entry: dict) -> Path:
    """Append one entry to ``artifacts/commissioning_report.json`` (create if missing)."""
    path = _commissioning_report_path(run_dir)
    path.parent.mkdir(parents=True, exist_ok=True)
    config = _parse_run_config(run_dir)
    job_id = str(config.get("job_id", "")).strip() or "unknown-job"
    if path.is_file():
        doc = json.loads(path.read_text(encoding="utf-8"))
    else:
        doc = {
            "schema_version": "0.2-commissioning-report",
            "job_id": job_id,
            "entries": [],
        }
    if not isinstance(doc.get("entries"), list):
        doc["entries"] = []
    doc["job_id"] = job_id
    # Bump schema when older runs pick up new entry kinds (backward compatible readers).
    if str(doc.get("schema_version", "")).strip() == "0.1-commissioning-report":
        doc["schema_version"] = "0.2-commissioning-report"
    doc["entries"].append(entry)
    path.write_text(json.dumps(doc, indent=2, sort_keys=True), encoding="utf-8")
    return path


def _parse_read_spec(text: str) -> tuple[str, str]:
    """Parse ``object_id`` or ``object_id:property`` for modulation reads."""
    raw = str(text).strip()
    if not raw:
        raise ValueError("empty read spec")
    if ":" in raw:
        oid, prop = raw.split(":", 1)
        oid, prop = oid.strip(), prop.strip() or "presentValue"
    else:
        oid, prop = raw, "presentValue"
    if not oid:
        raise ValueError("object_id required in read spec")
    return oid, prop


def _commissioning_report_modulation_rows(doc: dict) -> list[dict[str, str]]:
    """Flatten thermal modulation-related entries for CSV."""
    rows: list[dict[str, str]] = []
    entries = doc.get("entries")
    if not isinstance(entries, list):
        return rows
    for ent in entries:
        if not isinstance(ent, dict):
            continue
        kind = str(ent.get("kind", "")).strip()
        if kind == "thermal_modulation_sweep":
            base_ts = str(ent.get("ts", ""))
            ctrl = str(ent.get("controller_label", ""))
            step_id = str(ent.get("step_id", ""))
            report_ref = str(ent.get("report_ref", ""))
            tech = str(ent.get("technician_name", ""))
            cmd_oid = str(ent.get("command_object_id", ""))
            cmd_pct = str(ent.get("command_percent", ""))
            dwell = str(ent.get("dwell_seconds", ""))
            for r in ent.get("readings", []) if isinstance(ent.get("readings"), list) else []:
                if not isinstance(r, dict):
                    continue
                rows.append(
                    {
                        "entry_ts": base_ts,
                        "kind": kind,
                        "controller_label": ctrl,
                        "step_id": step_id,
                        "report_ref": report_ref,
                        "technician_name": tech,
                        "command_object_id": cmd_oid,
                        "command_percent": cmd_pct,
                        "dwell_seconds": dwell,
                        "object_id": str(r.get("logical_object_id", "")),
                        "property": "presentValue",
                        "status": str(r.get("status", "")),
                        "value_str": str(r.get("value_str", "")),
                        "read_source": str(r.get("source", "")),
                    }
                )
            continue
        if kind not in {"thermal_modulation_sample", "thermal_modulation_batch"}:
            continue
        base_ts = str(ent.get("ts", ""))
        ctrl = str(ent.get("controller_label", ""))
        step_id = str(ent.get("step_id", ""))
        report_ref = str(ent.get("report_ref", ""))
        tech = str(ent.get("technician_name", ""))
        if kind == "thermal_modulation_batch":
            for sub in ent.get("samples", []) if isinstance(ent.get("samples"), list) else []:
                if not isinstance(sub, dict):
                    continue
                sub_ts = str(sub.get("ts", base_ts))
                for r in sub.get("readings", []) if isinstance(sub.get("readings"), list) else []:
                    if not isinstance(r, dict):
                        continue
                    rows.append(
                        {
                            "entry_ts": sub_ts,
                            "kind": "thermal_modulation_sample",
                            "controller_label": str(sub.get("controller_label", ctrl)),
                            "step_id": str(sub.get("step_id", step_id)),
                            "report_ref": str(sub.get("report_ref", report_ref)),
                            "technician_name": str(sub.get("technician_name", tech)),
                            "command_object_id": "",
                            "command_percent": "",
                            "dwell_seconds": "",
                            "object_id": str(r.get("object_id", "")),
                            "property": str(r.get("property", "")),
                            "status": str(r.get("status", "")),
                            "value_str": str(r.get("value_str", "")),
                            "read_source": "",
                        }
                    )
            continue
        for r in ent.get("readings", []) if isinstance(ent.get("readings"), list) else []:
            if not isinstance(r, dict):
                continue
            rows.append(
                {
                    "entry_ts": base_ts,
                    "kind": kind,
                    "controller_label": ctrl,
                    "step_id": step_id,
                    "report_ref": report_ref,
                    "technician_name": tech,
                    "command_object_id": "",
                    "command_percent": "",
                    "dwell_seconds": "",
                    "object_id": str(r.get("object_id", "")),
                    "property": str(r.get("property", "")),
                    "status": str(r.get("status", "")),
                    "value_str": str(r.get("value_str", "")),
                    "read_source": "",
                }
            )
    return rows


# Customer-facing heat/cool table: modulation sweep/sample/batch rows only (narrow columns).
COMMISSIONING_REPORT_CUSTOMER_MODULATION_FIELDNAMES: tuple[str, ...] = (
    "entry_ts",
    "kind",
    "controller_label",
    "step_id",
    "report_ref",
    "technician_name",
    "command_object_id",
    "command_percent",
    "dwell_seconds",
    "object_id",
    "property",
    "status",
    "value_str",
    "read_source",
)


COMMISSIONING_REPORT_UNIFIED_FIELDNAMES: tuple[str, ...] = (
    "entry_ts",
    "kind",
    "controller_label",
    "step_id",
    "step_status",
    "report_ref",
    "technician_name",
    "note",
    "all_read_ok",
    "artifact_json",
    "command_object_id",
    "command_percent",
    "dwell_seconds",
    "sweep_index",
    "sweep_count",
    "trigger",
    "object_id",
    "property",
    "status",
    "value_str",
    "read_source",
    "measurement_branch_id",
    "measured_flow_L_s",
    "measurement_tool",
    "design_flow_L_s",
    "session_key",
    "target_flow_ratio_of_design",
    "design_supply_airflow_L_s",
    "prompt_id",
)


def _customer_modulation_rows_to_html(
    job_id: str, schema_version: str, rows: list[dict[str, str]]
) -> str:
    """Printable HTML for customer-facing thermal modulation table (narrow columns)."""
    title = html.escape(f"Commissioning — thermal modulation — job {job_id}")
    esc_job = html.escape(job_id)
    esc_schema = html.escape(schema_version)
    cols = list(COMMISSIONING_REPORT_CUSTOMER_MODULATION_FIELDNAMES)
    th = "".join(f"<th>{html.escape(h)}</th>" for h in cols)
    body_rows: list[str] = []
    for row in rows:
        cells = "".join(
            f"<td>{html.escape(str(row.get(k, '') or ''))}</td>" for k in cols
        )
        body_rows.append(f"<tr>{cells}</tr>")
    colspan = str(len(cols))
    tbody = (
        "\n".join(body_rows)
        if body_rows
        else f"<tr><td colspan=\"{colspan}\">(no modulation entries)</td></tr>"
    )
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>{title}</title>
<style>
body {{ font-family: system-ui, sans-serif; margin: 1rem; }}
h1 {{ font-size: 1.15rem; }}
.sub {{ color: #444; margin-bottom: 0.5rem; font-size: 0.95rem; }}
table {{ border-collapse: collapse; width: 100%; font-size: 0.9rem; }}
th, td {{ border: 1px solid #bbb; padding: 0.4rem 0.55rem; text-align: left; vertical-align: top; }}
th {{ background: #e8eef5; }}
tr:nth-child(even) td {{ background: #fafafa; }}
@media print {{ body {{ margin: 0; }} table {{ font-size: 0.82rem; }} }}
</style>
</head>
<body>
<h1>{title}</h1>
<p class="sub">Heat/cool modulation reads (sweep, sample, and batch rows). schema_version: {esc_schema}</p>
<table>
<thead><tr>{th}</tr></thead>
<tbody>
{tbody}
</tbody>
</table>
<p class="sub">Print to PDF from the browser (Ctrl+P) when sharing with customers.</p>
</body>
</html>
"""


def _commissioning_report_modulation_notes_for_pdf(doc: dict, *, max_lines: int = 24) -> list[str]:
    """Short lines from modulation-related report entries that carry operator notes."""
    lines: list[str] = []
    for ent in doc.get("entries", []) if isinstance(doc.get("entries"), list) else []:
        if not isinstance(ent, dict):
            continue
        kind = str(ent.get("kind", "")).strip()
        if not kind.startswith("thermal_modulation"):
            continue
        note = str(ent.get("note", "")).strip()
        if not note:
            continue
        ts = str(ent.get("ts", "")).strip()
        ctrl = str(ent.get("controller_label", "")).strip()
        head = f"{ts} {ctrl}".strip()
        lines.append(_pdf_cell_safe(f"{head} — {note}", 120))
        if len(lines) >= max_lines:
            break
    return lines


def _write_customer_modulation_pdf(
    path: Path,
    doc: dict,
    *,
    logo_image_path: Path | None = None,
) -> None:
    """Landscape PDF: cover page, modulation table, optional notes (A1 polish)."""
    try:
        from fpdf import FPDF
        from fpdf.enums import XPos, YPos
    except ImportError as err:  # pragma: no cover
        raise RuntimeError(
            "fpdf2 is required for --output-customer-pdf; install requirements.txt"
        ) from err

    rows = _commissioning_report_modulation_rows(doc)
    headers = list(COMMISSIONING_REPORT_CUSTOMER_MODULATION_FIELDNAMES)
    job_id = str(doc.get("job_id", "") or "unknown-job")
    schema_v = str(doc.get("schema_version", "") or "")
    gen_ts = _utc_timestamp()

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=10)
    margin_l = pdf.l_margin
    usable_w = pdf.w - margin_l - pdf.r_margin

    # --- Cover (portrait-style title block on first landscape page)
    pdf.add_page()
    y0 = _pdf_draw_commissioning_logo_strip(
        pdf,
        logo_image_path=logo_image_path,
        x=margin_l,
        y=pdf.get_y(),
    )
    pdf.set_y(y0 + 4)
    pdf.set_font("Helvetica", style="B", size=14)
    pdf.set_text_color(30, 30, 30)
    pdf.cell(0, 8, _pdf_cell_safe("Commissioning — thermal modulation summary", 80), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.set_font("Helvetica", size=10)
    pdf.cell(0, 6, _pdf_cell_safe(f"job_id: {job_id}", 100), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.cell(0, 6, _pdf_cell_safe(f"schema_version: {schema_v}", 100), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.cell(0, 6, _pdf_cell_safe(f"generated_utc: {gen_ts}", 100), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
    pdf.ln(6)
    pdf.set_font("Helvetica", size=9)
    pdf.multi_cell(0, 5, _pdf_cell_safe(
        "Following pages list sweep/sample/batch modulation rows exported from the commissioning report. "
        "Use the integrator unified export for full audit columns.",
        240,
    ))

    # --- Data table
    pdf.add_page()
    y_after_logo = _pdf_draw_commissioning_logo_strip(
        pdf,
        logo_image_path=logo_image_path,
        x=margin_l,
        y=pdf.get_y(),
    )
    pdf.set_y(y_after_logo)
    pdf.set_font("Helvetica", size=10)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(
        0,
        6,
        _pdf_cell_safe(
            f"Thermal modulation (heat/cool)  job_id={job_id}  schema={schema_v}", 220
        ),
        new_x=XPos.LMARGIN,
        new_y=YPos.NEXT,
    )
    pdf.ln(1)

    col_w = usable_w / len(headers)
    row_h = 5
    y_bottom = pdf.h - 10

    def draw_header() -> None:
        pdf.set_font("Helvetica", style="B", size=6)
        for h in headers:
            pdf.cell(col_w, 6, _pdf_cell_safe(h, 28), border=1)
        pdf.ln()

    draw_header()
    pdf.set_font("Helvetica", size=6)
    for row in rows:
        if pdf.get_y() + row_h > y_bottom:
            pdf.add_page()
            draw_header()
            pdf.set_font("Helvetica", size=6)
        for h in headers:
            pdf.cell(
                col_w,
                row_h,
                _pdf_cell_safe(str(row.get(h, "") or ""), 36),
                border=1,
            )
        pdf.ln()

    if not rows:
        pdf.set_font("Helvetica", size=9)
        pdf.cell(0, 6, "(no modulation entries)", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    note_lines = _commissioning_report_modulation_notes_for_pdf(doc)
    if note_lines:
        pdf.add_page()
        pdf.set_font("Helvetica", style="B", size=11)
        pdf.cell(0, 7, "Operator notes (modulation entries)", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        pdf.ln(2)
        pdf.set_font("Helvetica", size=8)
        for line in note_lines:
            pdf.multi_cell(0, 4, line)

    pdf.output(str(path))


def _write_commissioning_report_unified_xlsx(
    path: Path, doc: dict, *, include_modulation_sheet: bool = False
) -> None:
    """Write unified commissioning rows to an ``.xlsx`` file (requires openpyxl)."""
    try:
        from openpyxl import Workbook
    except ImportError as err:  # pragma: no cover — exercised when dep missing
        raise RuntimeError(
            "openpyxl is required for --output-xlsx; install requirements.txt"
        ) from err

    path.parent.mkdir(parents=True, exist_ok=True)
    rows = _commissioning_report_unified_csv_rows(doc)
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "commissioning"
    ws.append(list(COMMISSIONING_REPORT_UNIFIED_FIELDNAMES))
    for row in rows:
        ws.append([str(row.get(k, "") or "") for k in COMMISSIONING_REPORT_UNIFIED_FIELDNAMES])
    if include_modulation_sheet:
        mod_rows = _commissioning_report_modulation_rows(doc)
        ws2 = wb.create_sheet(title="modulation")
        ws2.append(list(COMMISSIONING_REPORT_CUSTOMER_MODULATION_FIELDNAMES))
        for r in mod_rows:
            ws2.append(
                [str(r.get(k, "") or "") for k in COMMISSIONING_REPORT_CUSTOMER_MODULATION_FIELDNAMES]
            )
    wb.save(path)


def _pdf_cell_safe(text: str, max_len: int) -> str:
    """ASCII-safe string for fpdf2 core fonts (Helvetica = Latin-1 subset)."""
    t = str(text).replace("\r", " ").replace("\n", " ")
    if len(t) > max_len:
        t = t[: max(0, max_len - 3)] + "..."
    return t.encode("ascii", errors="replace").decode("ascii")


def _commissioning_pdf_bundled_placeholder_logo() -> Path:
    """Neutral placeholder shipped in-repo (not a customer trademark)."""
    return ROOT / "docs" / "examples" / "branding" / "commissioning-logo-placeholder.png"


def _ensure_default_branding_placeholder(artifacts_dir: Path) -> Path | None:
    """Copy bundled neutral logo to ``artifacts/branding/logo.png`` when missing."""
    src = _commissioning_pdf_bundled_placeholder_logo()
    if not src.is_file():
        return None
    dest_dir = artifacts_dir / "branding"
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / "logo.png"
    if dest.is_file():
        return dest
    shutil.copy2(src, dest)
    return dest


def _resolve_commissioning_pdf_logo_path(
    run_dir: Path, cli_path: Path | None
) -> Path | None:
    """Resolve logo: ``--pdf-logo-image``, then ``artifacts/branding/logo.png``, then bundled placeholder."""
    if cli_path is not None:
        p = Path(cli_path)
        if p.is_file():
            return p
    customer = run_dir / "artifacts" / "branding" / "logo.png"
    if customer.is_file():
        return customer
    bundled = _commissioning_pdf_bundled_placeholder_logo()
    return bundled if bundled.is_file() else None


def _pdf_draw_logo_placeholder_vector(pdf, x: float, y: float) -> float:
    """Neutral vector block when no logo file is available (no customer marks)."""
    gray = (120, 120, 120)
    pdf.set_draw_color(*gray)
    pdf.set_text_color(*gray)
    pdf.rect(x, y, 75, 14, style="D")
    pdf.set_font("Helvetica", size=9)
    pdf.text(x + 3, y + 5.5, "Site logo (add PNG under artifacts/branding/logo.png)")
    pdf.set_font("Helvetica", size=6)
    pdf.text(x + 3, y + 10.5, "or pass --pdf-logo-image")
    return y + 16.0


def _pdf_draw_commissioning_logo_strip(
    pdf,
    *,
    logo_image_path: Path | None,
    x: float,
    y: float,
) -> float:
    """Draw logo image (PNG/JPEG) or vector fallback. Returns Y coordinate below strip."""
    if logo_image_path is not None and logo_image_path.is_file():
        suffix = logo_image_path.suffix.lower()
        if suffix in {".png", ".jpg", ".jpeg", ".gif"}:
            try:
                target_h = 14.0
                pdf.image(
                    str(logo_image_path),
                    x=x,
                    y=y,
                    w=0,
                    h=target_h,
                    keep_aspect_ratio=True,
                )
                return y + target_h + 2.0
            except Exception:
                pass
    return _pdf_draw_logo_placeholder_vector(pdf, x, y)


def _write_commissioning_report_unified_pdf(
    path: Path,
    doc: dict,
    *,
    logo_image_path: Path | None = None,
) -> None:
    """Write unified commissioning rows to PDF (landscape A4; requires fpdf2)."""
    try:
        from fpdf import FPDF
        from fpdf.enums import XPos, YPos
    except ImportError as err:  # pragma: no cover
        raise RuntimeError(
            "fpdf2 is required for --output-pdf; install requirements.txt"
        ) from err

    path.parent.mkdir(parents=True, exist_ok=True)
    rows = _commissioning_report_unified_csv_rows(doc)
    headers = list(COMMISSIONING_REPORT_UNIFIED_FIELDNAMES)
    job_id = str(doc.get("job_id", "") or "unknown-job")
    schema_v = str(doc.get("schema_version", "") or "")

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page()
    margin_l = pdf.l_margin
    usable_w = pdf.w - margin_l - pdf.r_margin
    y_after_logo = _pdf_draw_commissioning_logo_strip(
        pdf,
        logo_image_path=logo_image_path,
        x=margin_l,
        y=pdf.get_y(),
    )
    pdf.set_y(y_after_logo)

    pdf.set_font("Helvetica", size=9)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(
        0,
        6,
        _pdf_cell_safe(f"Commissioning report  job_id={job_id}  schema={schema_v}", 200),
        new_x=XPos.LMARGIN,
        new_y=YPos.NEXT,
    )
    pdf.ln(1)

    col_w = usable_w / len(headers)
    row_h = 4
    y_bottom = pdf.h - 10

    def draw_header() -> None:
        pdf.set_font("Helvetica", style="B", size=5)
        for h in headers:
            pdf.cell(col_w, 5, _pdf_cell_safe(h, 24), border=1)
        pdf.ln()

    draw_header()
    pdf.set_font("Helvetica", size=5)
    for row in rows:
        if pdf.get_y() + row_h > y_bottom:
            pdf.add_page()
            draw_header()
            pdf.set_font("Helvetica", size=5)
        for h in headers:
            pdf.cell(
                col_w,
                row_h,
                _pdf_cell_safe(str(row.get(h, "") or ""), 28),
                border=1,
            )
        pdf.ln()

    if not rows:
        pdf.set_font("Helvetica", size=8)
        pdf.cell(0, 6, "(no entries)", new_x=XPos.LMARGIN, new_y=YPos.NEXT)

    pdf.output(str(path))


def _commissioning_report_unified_csv_rows(doc: dict) -> list[dict[str, str]]:
    """Flatten all commissioning_report entry kinds into one CSV-shaped row list."""
    rows: list[dict[str, str]] = []
    entries = doc.get("entries")
    if not isinstance(entries, list):
        return rows
    for ent in entries:
        if not isinstance(ent, dict):
            continue
        kind = str(ent.get("kind", "")).strip()
        base_ts = str(ent.get("ts", ""))
        ctrl = str(ent.get("controller_label", ""))
        step_id = str(ent.get("step_id", ""))
        report_ref = str(ent.get("report_ref", ""))
        tech = str(ent.get("technician_name", ""))
        note = str(ent.get("note", ""))

        if kind == "point_checkout_after_step":
            step_status = str(ent.get("step_status", ""))
            all_ok = str(ent.get("all_read_ok", "")).lower()
            artifact_json = str(ent.get("artifact_json", ""))
            summary = ent.get("read_summary")
            if isinstance(summary, list) and summary:
                for item in summary:
                    if not isinstance(item, dict):
                        continue
                    rows.append(
                        {
                            "entry_ts": base_ts,
                            "kind": kind,
                            "controller_label": ctrl,
                            "step_id": step_id,
                            "step_status": step_status,
                            "report_ref": report_ref,
                            "technician_name": tech,
                            "note": "",
                            "all_read_ok": all_ok,
                            "artifact_json": artifact_json,
                            "command_object_id": "",
                            "command_percent": "",
                            "dwell_seconds": "",
                            "sweep_index": "",
                            "sweep_count": "",
                            "trigger": "",
                            "object_id": str(item.get("object_id", "")),
                            "property": str(item.get("property", "")),
                            "status": str(item.get("status", "")),
                            "value_str": "",
                            "read_source": "",
                        }
                    )
            else:
                rows.append(
                    {
                        "entry_ts": base_ts,
                        "kind": kind,
                        "controller_label": ctrl,
                        "step_id": step_id,
                        "step_status": step_status,
                        "report_ref": report_ref,
                        "technician_name": tech,
                        "note": "",
                        "all_read_ok": all_ok,
                        "artifact_json": artifact_json,
                        "command_object_id": "",
                        "command_percent": "",
                        "dwell_seconds": "",
                        "sweep_index": "",
                        "sweep_count": "",
                        "trigger": "",
                        "object_id": "",
                        "property": "",
                        "status": "",
                        "value_str": "",
                        "read_source": "",
                    }
                )
            continue

        if kind == "thermal_modulation_sweep":
            cmd_oid = str(ent.get("command_object_id", ""))
            cmd_pct = str(ent.get("command_percent", ""))
            dwell = str(ent.get("dwell_seconds", ""))
            sweep_ix = str(ent.get("sweep_index", ""))
            sweep_ct = str(ent.get("sweep_count", ""))
            trigger = str(ent.get("trigger", ""))
            artifact_json = ""
            all_ok = ""
            step_status = ""
            for r in ent.get("readings", []) if isinstance(ent.get("readings"), list) else []:
                if not isinstance(r, dict):
                    continue
                rows.append(
                    {
                        "entry_ts": base_ts,
                        "kind": kind,
                        "controller_label": ctrl,
                        "step_id": step_id,
                        "step_status": step_status,
                        "report_ref": report_ref,
                        "technician_name": tech,
                        "note": note,
                        "all_read_ok": all_ok,
                        "artifact_json": artifact_json,
                        "command_object_id": cmd_oid,
                        "command_percent": cmd_pct,
                        "dwell_seconds": dwell,
                        "sweep_index": sweep_ix,
                        "sweep_count": sweep_ct,
                        "trigger": trigger,
                        "object_id": str(r.get("logical_object_id", "")),
                        "property": "presentValue",
                        "status": str(r.get("status", "")),
                        "value_str": str(r.get("value_str", "")),
                        "read_source": str(r.get("source", "")),
                    }
                )
            continue

        if kind == "thermal_modulation_sample":
            artifact_json = ""
            all_ok = ""
            step_status = ""
            for r in ent.get("readings", []) if isinstance(ent.get("readings"), list) else []:
                if not isinstance(r, dict):
                    continue
                rows.append(
                    {
                        "entry_ts": base_ts,
                        "kind": kind,
                        "controller_label": ctrl,
                        "step_id": step_id,
                        "step_status": step_status,
                        "report_ref": report_ref,
                        "technician_name": tech,
                        "note": note,
                        "all_read_ok": all_ok,
                        "artifact_json": artifact_json,
                        "command_object_id": "",
                        "command_percent": "",
                        "dwell_seconds": "",
                        "sweep_index": "",
                        "sweep_count": "",
                        "trigger": "",
                        "object_id": str(r.get("object_id", "")),
                        "property": str(r.get("property", "")),
                        "status": str(r.get("status", "")),
                        "value_str": str(r.get("value_str", "")),
                        "read_source": "",
                    }
                )
            continue

        if kind == "thermal_modulation_batch":
            artifact_json = ""
            all_ok = ""
            step_status = ""
            for sub in ent.get("samples", []) if isinstance(ent.get("samples"), list) else []:
                if not isinstance(sub, dict):
                    continue
                sub_ts = str(sub.get("ts", base_ts))
                s_ctrl = str(sub.get("controller_label", ctrl))
                s_step = str(sub.get("step_id", step_id))
                s_ref = str(sub.get("report_ref", report_ref))
                s_tech = str(sub.get("technician_name", tech))
                s_note = str(sub.get("note", ""))
                for r in sub.get("readings", []) if isinstance(sub.get("readings"), list) else []:
                    if not isinstance(r, dict):
                        continue
                    rows.append(
                        {
                            "entry_ts": sub_ts,
                            "kind": "thermal_modulation_sample",
                            "controller_label": s_ctrl,
                            "step_id": s_step,
                            "step_status": step_status,
                            "report_ref": s_ref,
                            "technician_name": s_tech,
                            "note": s_note,
                            "all_read_ok": all_ok,
                            "artifact_json": artifact_json,
                            "command_object_id": "",
                            "command_percent": "",
                            "dwell_seconds": "",
                            "sweep_index": "",
                            "sweep_count": "",
                            "trigger": "thermal_modulation_batch",
                            "object_id": str(r.get("object_id", "")),
                            "property": str(r.get("property", "")),
                            "status": str(r.get("status", "")),
                            "value_str": str(r.get("value_str", "")),
                            "read_source": "",
                        }
                    )
            continue

        if kind == "tachometer_reference_confirmation":
            rows.append(
                {
                    "entry_ts": base_ts,
                    "kind": kind,
                    "controller_label": ctrl,
                    "step_id": step_id,
                    "step_status": "",
                    "report_ref": report_ref,
                    "technician_name": tech,
                    "note": note,
                    "all_read_ok": "",
                    "artifact_json": "",
                    "command_object_id": "",
                    "command_percent": "",
                    "dwell_seconds": "",
                    "sweep_index": "",
                    "sweep_count": "",
                    "trigger": "",
                    "object_id": str(ent.get("read_object_id", "")),
                    "property": "presentValue",
                    "status": "read_ok",
                    "value_str": str(ent.get("reading_value_str", "")),
                    "read_source": "bacnet",
                    "measurement_branch_id": "",
                    "measured_flow_L_s": "",
                    "measurement_tool": "",
                    "design_flow_L_s": "",
                    "session_key": str(ent.get("session_key", "")),
                    "target_flow_ratio_of_design": "",
                    "design_supply_airflow_L_s": "",
                    "prompt_id": "",
                }
            )
            continue

        if kind == "valve_prompt_confirmation":
            cmd_oid = str(ent.get("command_object_id", "ao_chw_valve"))
            cmd_pct = str(ent.get("command_percent", ""))
            rows.append(
                {
                    "entry_ts": base_ts,
                    "kind": kind,
                    "controller_label": ctrl,
                    "step_id": step_id,
                    "step_status": "",
                    "report_ref": report_ref,
                    "technician_name": tech,
                    "note": note,
                    "all_read_ok": "",
                    "artifact_json": "",
                    "command_object_id": cmd_oid,
                    "command_percent": cmd_pct,
                    "dwell_seconds": "",
                    "sweep_index": "",
                    "sweep_count": "",
                    "trigger": "",
                    "object_id": cmd_oid,
                    "property": "presentValue",
                    "status": "write_ok",
                    "value_str": "confirmed",
                    "read_source": "bacnet",
                    "measurement_branch_id": "",
                    "measured_flow_L_s": "",
                    "measurement_tool": "",
                    "design_flow_L_s": "",
                    "session_key": str(ent.get("session_key", "")),
                    "target_flow_ratio_of_design": "",
                    "design_supply_airflow_L_s": "",
                    "prompt_id": str(ent.get("prompt_id", "")),
                }
            )
            continue

        if kind == "airflow_adjust_command":
            try:
                ratio_v = float(ent.get("target_flow_ratio_of_design", ""))
                ratio_s = str(ratio_v)
            except (TypeError, ValueError):
                ratio_s = str(ent.get("target_flow_ratio_of_design", ""))
            dsa = ent.get("design_supply_airflow_L_s", "")
            rows.append(
                {
                    "entry_ts": base_ts,
                    "kind": kind,
                    "controller_label": ctrl,
                    "step_id": step_id,
                    "step_status": "",
                    "report_ref": report_ref,
                    "technician_name": tech,
                    "note": note,
                    "all_read_ok": "",
                    "artifact_json": "",
                    "command_object_id": str(ent.get("actuator_object_id", "")),
                    "command_percent": str(ent.get("command_percent", "")),
                    "dwell_seconds": "",
                    "sweep_index": "",
                    "sweep_count": "",
                    "trigger": "",
                    "object_id": "",
                    "property": "",
                    "status": "",
                    "value_str": "",
                    "read_source": "",
                    "measurement_branch_id": "",
                    "measured_flow_L_s": "",
                    "measurement_tool": "",
                    "design_flow_L_s": "",
                    "session_key": "",
                    "target_flow_ratio_of_design": ratio_s,
                    "design_supply_airflow_L_s": str(dsa) if dsa != "" else "",
                    "prompt_id": "",
                }
            )
            continue

        if kind == "manual_airflow_measurement":
            rows.append(
                {
                    "entry_ts": base_ts,
                    "kind": kind,
                    "controller_label": ctrl,
                    "step_id": step_id,
                    "step_status": "",
                    "report_ref": report_ref,
                    "technician_name": tech,
                    "note": note,
                    "all_read_ok": "",
                    "artifact_json": "",
                    "command_object_id": "",
                    "command_percent": "",
                    "dwell_seconds": "",
                    "sweep_index": "",
                    "sweep_count": "",
                    "trigger": "",
                    "object_id": "",
                    "property": "",
                    "status": "",
                    "value_str": "",
                    "read_source": "",
                    "measurement_branch_id": str(ent.get("branch_id", "")),
                    "measured_flow_L_s": str(ent.get("measured_flow_L_s", "")),
                    "measurement_tool": str(ent.get("measurement_tool", "")),
                    "design_flow_L_s": str(ent.get("design_flow_L_s", "")),
                    "session_key": str(ent.get("session_key", "")),
                    "target_flow_ratio_of_design": "",
                    "design_supply_airflow_L_s": "",
                    "prompt_id": "",
                }
            )
            continue

    for row in rows:
        for k in COMMISSIONING_REPORT_UNIFIED_FIELDNAMES:
            row.setdefault(k, "")
    return rows


def _modulation_sweep_cmd_sat_series(
    mod_rows: list[dict[str, str]],
) -> dict[str, list[tuple[float, float]]]:
    """Per controller: ordered (command_percent, ai_sat_degC) from thermal_modulation_sweep rows."""
    from collections import defaultdict

    groups: dict[tuple[str, str, str], dict[str, Any]] = {}
    for r in mod_rows:
        if str(r.get("kind", "")).strip() != "thermal_modulation_sweep":
            continue
        ctrl = str(r.get("controller_label", "")).strip()
        ts = str(r.get("entry_ts", "")).strip()
        cmd_s = str(r.get("command_percent", "")).strip()
        if not ctrl or not ts or not cmd_s:
            continue
        try:
            cmd = float(cmd_s)
        except ValueError:
            continue
        key = (ctrl, ts, cmd_s)
        g = groups.setdefault(key, {"cmd": cmd, "sat": None})
        oid = str(r.get("object_id", "")).strip()
        if oid == "ai_sat" and str(r.get("status", "")).strip() == "read_ok":
            vs = str(r.get("value_str", "")).strip()
            try:
                g["sat"] = float(vs)
            except ValueError:
                pass
    by_ctrl: dict[str, list[tuple[float, float]]] = defaultdict(list)
    for (ctrl, ts, _cmd_s) in sorted(groups.keys(), key=lambda k: (k[0], k[1])):
        entry = groups[(ctrl, ts, _cmd_s)]
        sat = entry.get("sat")
        if sat is None:
            continue
        by_ctrl[ctrl].append((float(entry["cmd"]), float(sat)))
    return dict(by_ctrl)


def _svg_sparkline_cmd_vs_sat(
    points: list[tuple[float, float]],
    *,
    width: int = 360,
    height: int = 120,
    pad: int = 12,
) -> str:
    """Inline SVG: X = command %, Y = SAT °C (simple polyline)."""
    if len(points) < 2:
        return ""
    cmds = [p[0] for p in points]
    sats = [p[1] for p in points]
    cmin, cmax = min(cmds), max(cmds)
    smin, smax = min(sats), max(sats)
    if cmax <= cmin:
        cmax = cmin + 1.0
    if smax <= smin:
        smax = smin + 0.1
    inner_w = width - 2 * pad
    inner_h = height - 2 * pad

    def cx(x: float) -> float:
        return pad + (x - cmin) / (cmax - cmin) * inner_w

    def cy(y: float) -> float:
        return pad + (1.0 - (y - smin) / (smax - smin)) * inner_h

    pts = " ".join(f"{cx(c):.1f},{cy(s):.1f}" for c, s in points)
    title = html.escape(f"command% {cmin:.0f}-{cmax:.0f} vs SAT °C {smin:.1f}-{smax:.1f}")
    return (
        f'<svg class="commissioning-mod-chart" xmlns="http://www.w3.org/2000/svg" '
        f'role="img" aria-label="Modulation SAT vs command" width="{width}" height="{height}">'
        f'<title>{title}</title>'
        f'<rect x="0" y="0" width="{width}" height="{height}" fill="#fafafa" stroke="#ccc"/>'
        f'<polyline fill="none" stroke="#1a5fb4" stroke-width="2" points="{pts}"/>'
        f'<text x="{pad}" y="{height - 4}" font-size="10" fill="#333">'
        f'X: command % &nbsp; Y: ai_sat °C</text></svg>'
    )


def _html_modulation_charts_section(doc: dict) -> str:
    mod_rows = _commissioning_report_modulation_rows(doc)
    series = _modulation_sweep_cmd_sat_series(mod_rows)
    if not series:
        return ""
    parts: list[str] = [
        '<section class="mod-charts"><h2>Modulation (command % vs SAT)</h2>',
        '<p class="meta">From <code>thermal_modulation_sweep</code> rows with '
        "<code>ai_sat</code> read_ok.</p>",
    ]
    for ctrl in sorted(series):
        pts = series[ctrl]
        svg = _svg_sparkline_cmd_vs_sat(pts)
        if not svg:
            continue
        parts.append(f'<div class="mod-chart-block"><h3>{html.escape(ctrl)}</h3>{svg}</div>')
    parts.append("</section>")
    return "\n".join(parts)


def _commissioning_report_unified_rows_to_html(
    job_id: str, schema_version: str, rows: list[dict[str, str]], *, doc: dict | None = None
) -> str:
    """Minimal printable HTML table from unified commissioning report rows."""
    title = html.escape(f"Commissioning report — job {job_id}")
    esc_job = html.escape(job_id)
    esc_schema = html.escape(schema_version)
    cols = list(COMMISSIONING_REPORT_UNIFIED_FIELDNAMES)
    th = "".join(f"<th>{html.escape(h)}</th>" for h in cols)
    body_rows: list[str] = []
    for row in rows:
        cells = "".join(
            f"<td>{html.escape(str(row.get(k, '') or ''))}</td>" for k in cols
        )
        body_rows.append(f"<tr>{cells}</tr>")
    colspan = str(len(cols))
    tbody = (
        "\n".join(body_rows)
        if body_rows
        else f"<tr><td colspan=\"{colspan}\">(no entries)</td></tr>"
    )
    chart_block = ""
    if doc is not None:
        sec = _html_modulation_charts_section(doc)
        if sec:
            chart_block = f"\n{sec}\n"
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1"/>
<title>{title}</title>
<style>
body {{ font-family: system-ui, sans-serif; margin: 1rem; }}
h1 {{ font-size: 1.1rem; }}
.meta {{ color: #444; margin-bottom: 0.75rem; font-size: 0.9rem; }}
.mod-charts h2 {{ font-size: 1rem; margin-top: 1.25rem; }}
.mod-charts h3 {{ font-size: 0.95rem; margin: 0.5rem 0 0.25rem; }}
.mod-chart-block {{ margin-bottom: 1rem; }}
table {{ border-collapse: collapse; width: 100%; font-size: 0.85rem; }}
th, td {{ border: 1px solid #ccc; padding: 0.35rem 0.5rem; text-align: left; vertical-align: top; }}
th {{ background: #f0f0f0; }}
tr:nth-child(even) td {{ background: #fafafa; }}
@media print {{ body {{ margin: 0; }} table {{ font-size: 0.75rem; }} }}
</style>
</head>
<body>
<h1>{title}</h1>
<p class="meta">schema_version: {esc_schema} &middot; job_id: {esc_job} &middot; Print to PDF from browser (Ctrl+P).</p>
{chart_block}
<table>
<thead><tr>{th}</tr></thead>
<tbody>
{tbody}
</tbody>
</table>
</body>
</html>
"""


def cmd_append_commissioning_modulation_sample(args: argparse.Namespace) -> int:
    """BACnet-read allowlisted points and append one thermal_modulation_sample report row."""
    run_dir = args.run_dir
    logs_path = run_dir / "logs" / "events.jsonl"
    runtime_job_path = run_dir / "state" / "runtime-job.json"
    runtime_job = json.loads(runtime_job_path.read_text(encoding="utf-8"))
    target = None
    for controller in runtime_job.get("controllers", []):
        if controller.get("controller_label") == args.controller_label:
            target = controller
            break
    if target is None:
        print(f"error: controller not found in runtime job: {args.controller_label}")
        return 2

    read_specs: list[tuple[str, str]] = []
    for item in getattr(args, "read", None) or []:
        try:
            read_specs.append(_parse_read_spec(str(item)))
        except ValueError as err:
            print(f"error: invalid --read value {item!r}: {err}")
            return 2
    if not read_specs:
        print("error: supply at least one --read object_id or object_id:property")
        return 2

    try:
        bind_port = int(getattr(args, "bacnet_bind_port", 0) or 0)
    except ValueError:
        bind_port = 0
    try:
        _bacnet_adapter().commissioning_apdu_timeout_seconds(args.apdu_timeout)
    except (TypeError, ValueError) as err:
        print(f"error: invalid --apdu-timeout: {err}")
        return 2

    readings: list[dict] = []
    for oid, prop in read_specs:
        one = _bacnet_read_one(
            controller_label=args.controller_label,
            target=target,
            object_id=oid,
            property_name=prop,
            timeout_seconds=float(args.timeout_seconds),
            retries=int(args.retries),
            bacnet_bind_port=bind_port,
            apdu_timeout_override=args.apdu_timeout,
        )
        vr = ""
        if one.get("status") == "read_ok":
            vr = str(one.get("read", {}).get("value_str", ""))
        readings.append(
            {
                "object_id": oid,
                "property": prop,
                "status": str(one.get("status", "")),
                "value_str": vr,
            }
        )

    entry = {
        "ts": _utc_timestamp(),
        "kind": "thermal_modulation_sample",
        "controller_label": args.controller_label,
        "step_id": str(getattr(args, "step_id", "") or "").strip() or None,
        "report_ref": str(getattr(args, "report_ref", "") or "").strip() or None,
        "technician_name": str(args.technician_name).strip(),
        "note": str(getattr(args, "note", "") or ""),
        "readings": readings,
    }
    if entry["step_id"] is None:
        del entry["step_id"]
    if entry["report_ref"] is None:
        del entry["report_ref"]

    report_path = _append_commissioning_report_entry(run_dir, entry)
    _append_event(
        logs_path,
        "commissioning_modulation_sample_appended",
        {
            "controller_label": args.controller_label,
            "read_count": len(readings),
            "commissioning_report_json": str(report_path.resolve()),
        },
    )
    out = {"appended": True, "commissioning_report_json": str(report_path.resolve()), "readings": readings}
    print(json.dumps(out, indent=2, sort_keys=True))
    return 0 if all(r.get("status") == "read_ok" for r in readings) else 2


def _resolve_profile_object_bacnet(
    target: dict, object_id: str
) -> tuple[int, int] | None:
    """Return (object_type_int, instance) for logical ``object_id`` or None."""
    oid = str(object_id).strip()
    objs = target.get("objects_by_id", {})
    if not isinstance(objs, dict) or oid not in objs:
        return None
    meta = objs.get(oid)
    if not isinstance(meta, dict):
        return None
    bacnet = meta.get("bacnet", {})
    if not isinstance(bacnet, dict):
        return None
    type_name = str(bacnet.get("object_type", "")).strip()
    try:
        inst = int(bacnet.get("instance"))
    except (TypeError, ValueError):
        return None
    ad = _bacnet_adapter()
    ot = ad.object_type_name_to_int(type_name)
    if ot is None:
        return None
    return ot, inst


def _parse_modulation_command_percents(args: argparse.Namespace) -> list[float]:
    """Resolve command percent list from CLI (comma list overrides single percent)."""
    raw_list = str(getattr(args, "command_percents", "") or "").strip()
    if raw_list:
        out: list[float] = []
        for part in raw_list.split(","):
            text = part.strip()
            if not text:
                continue
            out.append(float(text))
        return out
    pct = getattr(args, "command_percent", None)
    if pct is None:
        return []
    return [float(pct)]


def _session_rat_reading(session_key: str, session_values: dict[str, str]) -> dict:
    key = str(session_key).strip()
    raw = str(session_values.get(key, "")).strip()
    ok = bool(raw)
    return {
        "logical_object_id": key,
        "status": "read_ok" if ok else "read_missing",
        "value_str": raw,
        "source": "session",
    }


def _execute_modulation_sweep_sequence(
    *,
    run_dir: Path,
    logs_path: Path,
    controller_label: str,
    step_id: str,
    step: dict,
    target: dict,
    action: dict,
    command_percents: list[float],
    dwell_seconds: float,
    technician_name: str,
    note: str,
    timeout_seconds: float,
    retries: int,
    bind_port: int,
    apdu_timeout_override: float | None,
    report_ref_override: str,
    trigger: str,
) -> dict:
    """Run one or more write/dwell/read cycles; append ``thermal_modulation_sweep`` entries."""
    cmd_oid = str(action.get("command_object_id", "")).strip()
    sat_oid = str(action.get("result_supply_temperature_object_id", "")).strip()
    rat_raw = action.get("result_return_temperature_object_id")
    rat_oid = str(rat_raw).strip() if rat_raw is not None else ""
    session_rat_key = str(action.get("session_return_air_temperature_key", "") or "").strip()

    if not cmd_oid or not sat_oid:
        return {
            "ok": False,
            "message": "modulate action missing command_object_id or result_supply_temperature_object_id",
        }

    cmd_res = _resolve_profile_object_bacnet(target, cmd_oid)
    sat_res = _resolve_profile_object_bacnet(target, sat_oid)
    if cmd_res is None or sat_res is None:
        return {
            "ok": False,
            "message": "could not resolve BACnet object for command or SAT from profile",
        }
    cmd_ot, cmd_oi = cmd_res
    sat_ot, sat_oi = sat_res

    rat_ot = rat_oi = None
    rat_bacnet_oid = ""
    if rat_oid:
        rat_res = _resolve_profile_object_bacnet(target, rat_oid)
        if rat_res is not None:
            rat_ot, rat_oi = rat_res
            rat_bacnet_oid = rat_oid
        elif not session_rat_key:
            return {
                "ok": False,
                "message": f"could not resolve BACnet object for RAT {rat_oid!r} (no session_return_air_temperature_key)",
            }

    session_state = _load_session_state(run_dir, controller_label)
    session_values = _session_values_map(session_state or {})

    extra_reads: list[tuple[str, int, int]] = []
    raw_ctx = action.get("optional_context_object_ids")
    if isinstance(raw_ctx, list):
        for oid in raw_ctx:
            text = str(oid).strip()
            if not text:
                continue
            res = _resolve_profile_object_bacnet(target, text)
            if res is None:
                return {
                    "ok": False,
                    "message": f"could not resolve optional_context_object_ids entry {text!r}",
                }
            extra_reads.append((text, res[0], res[1]))

    objects_by_id = target.get("objects_by_id", {})
    if not isinstance(objects_by_id, dict) or cmd_oid not in objects_by_id:
        return {"ok": False, "message": "command object not in objects_by_id"}
    if not bool(objects_by_id[cmd_oid].get("writable")):
        return {"ok": False, "message": f"command object {cmd_oid!r} is not writable in profile"}

    addr = target.get("bacnet", {})
    host = str(addr.get("host", "")).strip()
    port = int(addr.get("port", 0))
    expected_instance = int(addr.get("device_instance", 0))
    bacnet_ad = _bacnet_adapter()
    target_addr = bacnet_ad.format_ipv4_target(host, port)

    try:
        apdu_t = bacnet_ad.commissioning_apdu_timeout_seconds(apdu_timeout_override)
    except (TypeError, ValueError) as err:
        return {"ok": False, "message": f"invalid apdu_timeout: {err}"}
    who_t = bacnet_ad.effective_who_is_timeout(float(timeout_seconds), int(retries))

    if dwell_seconds < 0:
        return {"ok": False, "message": "dwell_seconds must be >= 0"}

    report_ref = str(step.get("report_ref", "")).strip() or str(report_ref_override or "").strip()

    def _read_one(oid: str, ot: int, oi: int) -> dict:
        r = bacnet_ad.read_present_value(
            bind_port=bind_port,
            target_address=target_addr,
            expected_device_instance=expected_instance,
            object_type=ot,
            object_instance=oi,
            property_name="presentValue",
            who_is_timeout=who_t,
            apdu_timeout=apdu_t,
        )
        vr = ""
        if r.get("status") == "read_ok":
            vr = str(r.get("value_str", ""))
        return {
            "logical_object_id": oid,
            "status": str(r.get("status", "")),
            "value_str": vr,
            "source": "bacnet",
        }

    sweep_rows: list[dict] = []
    last_report_path: Path | None = None
    overall_reads_ok = True

    for sweep_index, pct in enumerate(command_percents):
        write_res = bacnet_ad.write_present_value(
            bind_port=bind_port,
            target_address=target_addr,
            expected_device_instance=expected_instance,
            object_type=cmd_ot,
            object_instance=cmd_oi,
            value=float(pct),
            who_is_timeout=who_t,
            apdu_timeout=apdu_t,
        )
        if write_res.get("status") != "write_ok":
            return {
                "ok": False,
                "message": "write_failed",
                "write": write_res,
                "sweep_index": sweep_index,
                "command_percent": float(pct),
            }

        if dwell_seconds > 0:
            time.sleep(dwell_seconds)

        readings: list[dict] = []
        readings.append(_read_one(sat_oid, sat_ot, sat_oi))
        if rat_ot is not None:
            readings.append(_read_one(rat_bacnet_oid, rat_ot, rat_oi))
        elif session_rat_key:
            readings.append(_session_rat_reading(session_rat_key, session_values))
        for logical, ot, oi in extra_reads:
            readings.append(_read_one(logical, ot, oi))

        step_reads_ok = all(r.get("status") == "read_ok" for r in readings)
        overall_reads_ok = overall_reads_ok and step_reads_ok

        entry = {
            "ts": _utc_timestamp(),
            "kind": "thermal_modulation_sweep",
            "controller_label": controller_label,
            "step_id": step_id,
            "command_object_id": cmd_oid,
            "command_percent": float(pct),
            "dwell_seconds": float(dwell_seconds),
            "readings": readings,
            "write": {"status": write_res.get("status")},
            "technician_name": str(technician_name).strip(),
            "note": str(note or ""),
            "trigger": trigger,
            "sweep_index": sweep_index,
            "sweep_count": len(command_percents),
        }
        if report_ref:
            entry["report_ref"] = report_ref

        last_report_path = _append_commissioning_report_entry(run_dir, entry)
        sweep_rows.append(
            {
                "command_percent": float(pct),
                "readings": readings,
                "all_read_ok": step_reads_ok,
            }
        )

    if last_report_path is not None:
        _append_event(
            logs_path,
            "bacnet_modulation_sweep_completed",
            {
                "controller_label": controller_label,
                "step_id": step_id,
                "all_read_ok": bool(overall_reads_ok),
                "sweep_steps": len(command_percents),
                "trigger": trigger,
                "commissioning_report_json": str(last_report_path.resolve()),
            },
        )

    return {
        "ok": bool(overall_reads_ok),
        "partial_reads": not overall_reads_ok,
        "sweep_rows": sweep_rows,
        "commissioning_report_json": str(last_report_path.resolve())
        if last_report_path
        else "",
    }


def cmd_bacnet_modulation_sweep(args: argparse.Namespace) -> int:
    """Write command percent(s), dwell, read SAT/RAT/context points; append commissioning_report."""
    run_dir = args.run_dir
    logs_path = run_dir / "logs" / "events.jsonl"
    runtime_job_path = run_dir / "state" / "runtime-job.json"
    runtime_job = json.loads(runtime_job_path.read_text(encoding="utf-8"))
    target = None
    for controller in runtime_job.get("controllers", []):
        if controller.get("controller_label") == args.controller_label:
            target = controller
            break
    if target is None:
        print(f"error: controller not found in runtime job: {args.controller_label}")
        return 2

    flow_path = _flow_state_path(run_dir, args.controller_label)
    if not flow_path.is_file():
        print(f"error: flow state missing; run init-flow first ({flow_path})")
        return 2
    flow_state = json.loads(flow_path.read_text(encoding="utf-8"))
    step = None
    for item in flow_state.get("steps", []):
        if item.get("step_id") == args.step_id:
            step = item
            break
    if step is None:
        print(f"error: step_id not found in flow state: {args.step_id}")
        return 2

    action = _find_modulate_actuator_action(step)
    if action is None:
        print(
            "error: step has no modulate_actuator_log_sat_for_report action; "
            "compile-import with profile commissioning_flow.actions"
        )
        return 2

    try:
        bind_port = int(getattr(args, "bacnet_bind_port", 0) or 0)
    except ValueError:
        bind_port = 0

    percents = _parse_modulation_command_percents(args)
    if not percents:
        print("error: provide --command-percent and/or non-empty --command-percents")
        return 2

    dwell = float(args.dwell_seconds)

    result = _execute_modulation_sweep_sequence(
        run_dir=run_dir,
        logs_path=logs_path,
        controller_label=args.controller_label,
        step_id=args.step_id,
        step=step,
        target=target,
        action=action,
        command_percents=percents,
        dwell_seconds=dwell,
        technician_name=str(args.technician_name).strip(),
        note=str(getattr(args, "note", "") or ""),
        timeout_seconds=float(args.timeout_seconds),
        retries=int(args.retries),
        bind_port=bind_port,
        apdu_timeout_override=args.apdu_timeout,
        report_ref_override=str(getattr(args, "report_ref", "") or "").strip(),
        trigger="bacnet_modulation_sweep_cli",
    )

    if not result.get("ok") and result.get("message") == "write_failed":
        print(
            json.dumps(
                {"status": "write_failed", "write": result.get("write")},
                indent=2,
                sort_keys=True,
            )
        )
        return 2
    if result.get("message") and not result.get("sweep_rows"):
        print(f"error: {result.get('message')}")
        return 2

    sweep_rows = result.get("sweep_rows") or []
    last_reads = sweep_rows[-1]["readings"] if sweep_rows else []
    print(
        json.dumps(
            {
                "status": "sweep_ok" if result.get("ok") else "sweep_partial_reads",
                "commissioning_report_json": result.get("commissioning_report_json", ""),
                "sweep_steps": len(sweep_rows),
                "readings": last_reads,
            },
            indent=2,
            sort_keys=True,
        )
    )
    return 0 if result.get("ok") else 2


def cmd_append_commissioning_modulation_batch(args: argparse.Namespace) -> int:
    """Append multiple thermal_modulation_sample entries from a JSON file."""
    run_dir = args.run_dir
    logs_path = run_dir / "logs" / "events.jsonl"
    raw = json.loads(args.input_json.read_text(encoding="utf-8"))
    if isinstance(raw, dict) and isinstance(raw.get("samples"), list):
        samples_in = raw["samples"]
    elif isinstance(raw, list):
        samples_in = raw
    else:
        print("error: JSON must be a list of samples or {\"samples\": [...]}")
        return 2

    runtime_job_path = run_dir / "state" / "runtime-job.json"
    runtime_job = json.loads(runtime_job_path.read_text(encoding="utf-8"))
    by_label = {
        str(c.get("controller_label", "")).strip(): c
        for c in runtime_job.get("controllers", [])
        if str(c.get("controller_label", "")).strip()
    }

    try:
        bind_port = int(getattr(args, "bacnet_bind_port", 0) or 0)
    except ValueError:
        bind_port = 0
    try:
        _bacnet_adapter().commissioning_apdu_timeout_seconds(args.apdu_timeout)
    except (TypeError, ValueError) as err:
        print(f"error: invalid --apdu-timeout: {err}")
        return 2

    batch_readings: list[dict] = []
    all_ok = True
    for idx, sample in enumerate(samples_in):
        if not isinstance(sample, dict):
            print(f"error: sample[{idx}] must be an object")
            return 2
        label = str(sample.get("controller_label", "")).strip()
        if not label or label not in by_label:
            print(f"error: sample[{idx}] missing or unknown controller_label {label!r}")
            return 2
        target = by_label[label]
        reads_in = sample.get("reads") or sample.get("read") or []
        if isinstance(reads_in, str):
            reads_in = [reads_in]
        if not isinstance(reads_in, list) or not reads_in:
            print(f"error: sample[{idx}] needs non-empty reads array")
            return 2
        specs: list[tuple[str, str]] = []
        for r in reads_in:
            if isinstance(r, dict):
                oid = str(r.get("object_id", "")).strip()
                prop = str(r.get("property", "presentValue")).strip() or "presentValue"
                if not oid:
                    print(f"error: sample[{idx}] read object missing object_id")
                    return 2
                specs.append((oid, prop))
            else:
                try:
                    specs.append(_parse_read_spec(str(r)))
                except ValueError as err:
                    print(f"error: sample[{idx}] invalid read {r!r}: {err}")
                    return 2
        readings: list[dict] = []
        for oid, prop in specs:
            one = _bacnet_read_one(
                controller_label=label,
                target=target,
                object_id=oid,
                property_name=prop,
                timeout_seconds=float(
                    sample.get("timeout_seconds", getattr(args, "timeout_seconds", 0.5))
                ),
                retries=int(sample.get("retries", getattr(args, "retries", 1))),
                bacnet_bind_port=bind_port,
                apdu_timeout_override=args.apdu_timeout,
            )
            vr = ""
            if one.get("status") == "read_ok":
                vr = str(one.get("read", {}).get("value_str", ""))
            readings.append(
                {
                    "object_id": oid,
                    "property": prop,
                    "status": str(one.get("status", "")),
                    "value_str": vr,
                }
            )
            if one.get("status") != "read_ok":
                all_ok = False
        sub = {
            "ts": str(sample.get("ts", "")).strip() or _utc_timestamp(),
            "controller_label": label,
            "step_id": str(sample.get("step_id", "")).strip() or None,
            "report_ref": str(sample.get("report_ref", "")).strip() or None,
            "technician_name": (
                str(sample.get("technician_name", "")).strip()
                or str(getattr(args, "default_technician", "") or "").strip()
                or "unknown"
            ),
            "note": str(sample.get("note", "")),
            "readings": readings,
        }
        if sub["step_id"] is None:
            del sub["step_id"]
        if sub["report_ref"] is None:
            del sub["report_ref"]
        batch_readings.append(sub)

    entry = {
        "ts": _utc_timestamp(),
        "kind": "thermal_modulation_batch",
        "sample_count": len(batch_readings),
        "samples": batch_readings,
    }
    report_path = _append_commissioning_report_entry(run_dir, entry)
    _append_event(
        logs_path,
        "commissioning_modulation_batch_appended",
        {
            "sample_count": len(batch_readings),
            "commissioning_report_json": str(report_path.resolve()),
        },
    )
    print(
        json.dumps(
            {
                "appended": True,
                "all_read_ok": bool(all_ok),
                "commissioning_report_json": str(report_path.resolve()),
                "sample_count": len(batch_readings),
            },
            indent=2,
            sort_keys=True,
        )
    )
    return 0 if all_ok else 2


def cmd_bacnet_point_checkout(args: argparse.Namespace) -> int:
    run_dir = args.run_dir
    logs_path = run_dir / "logs" / "events.jsonl"
    runtime_job_path = run_dir / "state" / "runtime-job.json"
    runtime_job = json.loads(runtime_job_path.read_text(encoding="utf-8"))

    target = None
    for controller in runtime_job.get("controllers", []):
        if controller.get("controller_label") == args.controller_label:
            target = controller
            break
    if target is None:
        print(f"error: controller not found in runtime job: {args.controller_label}")
        return 2

    checkout = target.get("point_checkout", [])
    if not isinstance(checkout, list) or not checkout:
        print(
            "error: profile has no point_checkout list; "
            "add point_checkout: [{object_id, property}, ...] to the unit profile JSON"
        )
        return 2

    try:
        bind_port = int(getattr(args, "bacnet_bind_port", 0) or 0)
    except ValueError:
        bind_port = 0

    try:
        _bacnet_adapter().commissioning_apdu_timeout_seconds(args.apdu_timeout)
    except (TypeError, ValueError) as err:
        print(f"error: invalid --apdu-timeout: {err}")
        return 2

    strict = bool(getattr(args, "strict", False))
    rows, all_ok = _run_point_checkout_reads(
        controller_label=args.controller_label,
        target=target,
        timeout_seconds=args.timeout_seconds,
        retries=args.retries,
        bacnet_bind_port=bind_port,
        apdu_timeout_override=args.apdu_timeout,
        strict=strict,
    )

    payload = {
        "controller_label": args.controller_label,
        "strict": strict,
        "point_count": len(rows),
        "all_read_ok": bool(all_ok),
        "reads": rows,
    }

    out_dir = run_dir / "artifacts" / "bacnet_point_checkout"
    out_dir.mkdir(parents=True, exist_ok=True)
    artifact = out_dir / f"{args.controller_label}.json"
    artifact.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")

    _append_event(
        logs_path,
        "bacnet_point_checkout_completed",
        {
            "controller_label": args.controller_label,
            "all_read_ok": bool(all_ok),
            "artifact_json": str(artifact.resolve()),
        },
    )
    print(json.dumps(payload, sort_keys=True))
    return 0 if all_ok else 2


def cmd_record_step(args: argparse.Namespace) -> int:
    run_dir = args.run_dir
    logs_path = run_dir / "logs" / "events.jsonl"
    flow_state_path = _flow_state_path(run_dir, args.controller_label)
    flow_state = json.loads(flow_state_path.read_text(encoding="utf-8"))
    runtime_job_path = run_dir / "state" / "runtime-job.json"

    if flow_state.get("controller_label") != args.controller_label:
        print(
            f"error: flow state is for {flow_state.get('controller_label')}, "
            f"not {args.controller_label}"
        )
        return 2

    step = None
    for item in flow_state.get("steps", []):
        if item.get("step_id") == args.step_id:
            step = item
            break
    if step is None:
        print(f"error: step_id not found in flow state: {args.step_id}")
        return 2

    steps = flow_state.get("steps", [])
    session_state = _load_session_state(run_dir, args.controller_label)
    session_vals = _session_values_map(session_state or {})
    transition_error = _validate_step_transition(
        steps=steps,
        step=step,
        requested_status=args.status,
        session_values=session_vals,
    )
    previous_status = str(step.get("status", "pending"))
    if transition_error is not None:
        normalized_reason = _normalize_rejection_reason(
            transition_error.get("reason_code", "")
        )
        rejection_message = str(
            transition_error.get("message", "invalid step transition")
        )
        history = step.setdefault("history", [])
        history.append(
            {
                "ts": _utc_timestamp(),
                "previous_status": previous_status,
                "attempted_status": args.status,
                "new_status": previous_status,
                "reason_code": normalized_reason,
                "rejection_reason_code": normalized_reason,
                "rejected": True,
                "message": rejection_message,
            }
        )
        flow_state_path.write_text(json.dumps(flow_state, indent=2), encoding="utf-8")
        _append_event(
            logs_path,
            "flow_step_rejected",
            {
                "controller_label": args.controller_label,
                "step_id": args.step_id,
                "previous_status": previous_status,
                "attempted_status": args.status,
                "reason_code": normalized_reason,
                "rejection_reason_code": normalized_reason,
                "rejection_message": rejection_message,
                "technician_name": args.technician_name,
                "flow_state_json": str(flow_state_path.resolve()),
            },
        )
        print(f"error: invalid step transition: {rejection_message}")
        return 2

    run_bacnet_pc = str(args.status).strip() in {"passed", "manual_passed"} and (
        step.get("run_point_checkout_on_pass") is True
        or str(step.get("step_type", "")).strip() == "bacnet_point_checkout"
    )
    checkout_payload: dict | None = None
    if run_bacnet_pc:
        runtime_job = json.loads(runtime_job_path.read_text(encoding="utf-8"))
        target = None
        for controller in runtime_job.get("controllers", []):
            if controller.get("controller_label") == args.controller_label:
                target = controller
                break
        if target is None:
            print(f"error: controller not found in runtime job: {args.controller_label}")
            return 2
        checkout = target.get("point_checkout", [])
        if not isinstance(checkout, list) or not checkout:
            print(
                "error: point checkout step requires profile point_checkout list "
                "(compile-import after profile edit)"
            )
            return 2
        try:
            bind_port = int(getattr(args, "bacnet_bind_port", 0) or 0)
        except ValueError:
            bind_port = 0
        try:
            _bacnet_adapter().commissioning_apdu_timeout_seconds(args.apdu_timeout)
        except (TypeError, ValueError) as err:
            print(f"error: invalid --apdu-timeout: {err}")
            return 2
        rows, all_ok = _run_point_checkout_reads(
            controller_label=args.controller_label,
            target=target,
            timeout_seconds=float(args.bacnet_timeout_seconds),
            retries=int(args.bacnet_retries),
            bacnet_bind_port=bind_port,
            apdu_timeout_override=args.apdu_timeout,
            strict=bool(args.bacnet_checkout_strict),
        )
        out_dir = run_dir / "artifacts" / "bacnet_point_checkout"
        out_dir.mkdir(parents=True, exist_ok=True)
        stamp = _utc_timestamp().replace(":", "-")
        artifact = out_dir / f"{args.controller_label}-{args.step_id}-{stamp}.json"
        checkout_payload = {
            "controller_label": args.controller_label,
            "step_id": args.step_id,
            "trigger": "record_step",
            "strict": bool(args.bacnet_checkout_strict),
            "point_count": len(rows),
            "all_read_ok": bool(all_ok),
            "reads": rows,
        }
        artifact.write_text(
            json.dumps(checkout_payload, indent=2, sort_keys=True), encoding="utf-8"
        )
        artifact_json = str(artifact.resolve())
        if not all_ok:
            print(
                "error: BACnet point checkout failed for this step "
                f"(artifact={artifact_json})"
            )
            return 2
        report_ref = str(step.get("report_ref", "")).strip()
        report_entry: dict = {
            "ts": _utc_timestamp(),
            "kind": "point_checkout_after_step",
            "controller_label": args.controller_label,
            "step_id": args.step_id,
            "step_status": args.status,
            "all_read_ok": bool(all_ok),
            "read_summary": [
                {
                    "object_id": r.get("profile_object_id"),
                    "status": r.get("status"),
                    "property": r.get("property"),
                }
                for r in rows
            ],
            "artifact_json": artifact_json,
        }
        if report_ref:
            report_entry["report_ref"] = report_ref
        report_path = _append_commissioning_report_entry(run_dir, report_entry)
        _append_event(
            logs_path,
            "flow_step_point_checkout",
            {
                "controller_label": args.controller_label,
                "step_id": args.step_id,
                "all_read_ok": bool(all_ok),
                "artifact_json": artifact_json,
                "commissioning_report_json": str(report_path.resolve()),
            },
        )
        checkout_payload["artifact_json"] = artifact_json

    modulation_summary: dict | None = None
    run_modulation = str(args.status).strip() in {"passed", "manual_passed"} and bool(
        getattr(args, "run_modulation_on_pass", True)
    )
    if run_modulation:
        mod_action = _find_modulate_actuator_action(step)
        if mod_action is not None:
            raw_pcts = str(getattr(args, "modulation_command_percents", "") or "").strip()
            if not raw_pcts:
                print(
                    "error: step has modulate_actuator_log_sat_for_report; "
                    "pass --modulation-command-percents (comma list, e.g. 0,50,100) "
                    "or disable with --no-run-modulation-on-pass"
                )
                return 2
            percents: list[float] = []
            for part in raw_pcts.split(","):
                t = part.strip()
                if not t:
                    continue
                percents.append(float(t))
            if not percents:
                print("error: --modulation-command-percents must list at least one value")
                return 2
            try:
                bind_port_m = int(getattr(args, "modulation_bacnet_bind_port", 0) or 0)
            except ValueError:
                bind_port_m = 0
            try:
                _bacnet_adapter().commissioning_apdu_timeout_seconds(
                    getattr(args, "modulation_apdu_timeout", None)
                )
            except (TypeError, ValueError) as err:
                print(f"error: invalid --modulation-apdu-timeout: {err}")
                return 2
            runtime_job = json.loads(runtime_job_path.read_text(encoding="utf-8"))
            target_m = None
            for controller in runtime_job.get("controllers", []):
                if controller.get("controller_label") == args.controller_label:
                    target_m = controller
                    break
            if target_m is None:
                print(f"error: controller not found in runtime job: {args.controller_label}")
                return 2
            mod_result = _execute_modulation_sweep_sequence(
                run_dir=run_dir,
                logs_path=logs_path,
                controller_label=args.controller_label,
                step_id=args.step_id,
                step=step,
                target=target_m,
                action=mod_action,
                command_percents=percents,
                dwell_seconds=float(getattr(args, "modulation_dwell_seconds", 0.0)),
                technician_name=str(args.technician_name).strip(),
                note=str(args.note or ""),
                timeout_seconds=float(getattr(args, "modulation_timeout_seconds", 0.5)),
                retries=int(getattr(args, "modulation_retries", 1)),
                bind_port=bind_port_m,
                apdu_timeout_override=getattr(args, "modulation_apdu_timeout", None),
                report_ref_override="",
                trigger="record_step",
            )
            if mod_result.get("message") == "write_failed":
                print(
                    "error: modulation sweep write failed: "
                    f"{json.dumps(mod_result.get('write'), sort_keys=True)}"
                )
                return 2
            if mod_result.get("message") and not mod_result.get("sweep_rows"):
                print(f"error: modulation sweep failed: {mod_result.get('message')}")
                return 2
            if not mod_result.get("ok"):
                print(
                    "error: modulation sweep completed with failed reads "
                    f"(report={mod_result.get('commissioning_report_json', '')})"
                )
                return 2
            modulation_summary = {
                "sweep_steps": len(mod_result.get("sweep_rows") or []),
                "commissioning_report_json": str(
                    mod_result.get("commissioning_report_json", "")
                ),
            }
            _append_event(
                logs_path,
                "flow_step_modulation_sweep",
                {
                    "controller_label": args.controller_label,
                    "step_id": args.step_id,
                    "sweep_steps": modulation_summary["sweep_steps"],
                    "commissioning_report_json": modulation_summary[
                        "commissioning_report_json"
                    ],
                },
            )

    record = {
        "ts": _utc_timestamp(),
        "status": args.status,
        "technician_name": args.technician_name,
        "note": args.note,
    }
    if checkout_payload is not None:
        record["point_checkout"] = {
            "all_read_ok": bool(checkout_payload["all_read_ok"]),
            "artifact_json": str(checkout_payload.get("artifact_json", "")),
            "point_count": int(checkout_payload["point_count"]),
        }
    if modulation_summary is not None:
        record["modulation_sweep"] = modulation_summary
    records = step.setdefault("records", [])
    records.append(record)
    history = step.setdefault("history", [])
    history.append(
        {
            "ts": _utc_timestamp(),
            "previous_status": previous_status,
            "attempted_status": args.status,
            "new_status": args.status,
            "reason_code": "status_update",
            "rejected": False,
        }
    )
    step["status"] = args.status
    step["technician_name"] = args.technician_name
    step["note"] = args.note
    if checkout_payload is not None:
        step["last_point_checkout"] = {
            "ts": _utc_timestamp(),
            "all_read_ok": bool(checkout_payload["all_read_ok"]),
            "artifact_json": str(checkout_payload.get("artifact_json", "")),
        }
    if modulation_summary is not None:
        step["last_modulation_sweep"] = {
            "ts": _utc_timestamp(),
            "sweep_steps": int(modulation_summary["sweep_steps"]),
            "commissioning_report_json": modulation_summary["commissioning_report_json"],
        }

    flow_state_path.write_text(json.dumps(flow_state, indent=2), encoding="utf-8")
    _append_event(
        logs_path,
        "flow_step_recorded",
        {
            "controller_label": args.controller_label,
            "step_id": args.step_id,
            "status": args.status,
            "previous_status": previous_status,
            "new_status": args.status,
            "reason_code": "status_update",
            "technician_name": args.technician_name,
            "flow_state_json": str(flow_state_path.resolve()),
        },
    )
    print(
        f"step_recorded=true controller_label={args.controller_label} "
        f"step_id={args.step_id} status={args.status}"
    )
    return 0


def cmd_probe_bip(args: argparse.Namespace) -> int:
    run_dir = args.run_dir
    logs_path = run_dir / "logs" / "events.jsonl"
    artifacts_path = run_dir / "artifacts" / "bip"

    runtime_job_path = run_dir / "state" / "runtime-job.json"
    runtime_job = json.loads(runtime_job_path.read_text(encoding="utf-8"))
    target = None
    for controller in runtime_job.get("controllers", []):
        if controller.get("controller_label") == args.controller_label:
            target = controller
            break
    if target is None:
        print(f"error: controller not found in runtime job: {args.controller_label}")
        return 2

    bacnet = target.get("bacnet", {})
    host = str(bacnet.get("host", "")).strip()
    port = int(bacnet.get("port"))
    expected_instance = int(bacnet.get("device_instance"))

    result = _probe_bip(
        host=host,
        port=port,
        expected_device_instance=expected_instance,
        timeout_seconds=args.timeout_seconds,
        retries=args.retries,
    )
    result["controller_label"] = args.controller_label

    artifact_file = artifacts_path / f"{args.controller_label}.json"
    artifacts_path.mkdir(parents=True, exist_ok=True)
    artifact_file.write_text(json.dumps(result, sort_keys=True), encoding="utf-8")

    _append_event(
        logs_path,
        "bip_probed",
        {
            "controller_label": args.controller_label,
            "host": host,
            "port": port,
            "expected_device_instance": expected_instance,
            "status": result.get("status"),
            "artifact_json": str(artifact_file.resolve()),
        },
    )
    print(json.dumps(result, sort_keys=True))
    return 0 if result.get("status") == "reachable_verified" else 2


def cmd_verify_bip_list(args: argparse.Namespace) -> int:
    run_dir = args.run_dir
    logs_path = run_dir / "logs" / "events.jsonl"
    runtime_job = json.loads((run_dir / "state" / "runtime-job.json").read_text(encoding="utf-8"))

    bip_artifacts_dir = run_dir / "artifacts" / "bip"
    bip_artifacts_dir.mkdir(parents=True, exist_ok=True)

    status_counts: dict[str, int] = {}
    rows: list[dict] = []
    total = 0
    unresolved = 0
    strict_pass = True

    allow_labels = {label.strip() for label in args.allow_known_unavailable}
    if args.known_unavailable_file:
        data = json.loads(args.known_unavailable_file.read_text(encoding="utf-8"))
        if bool(data.get("allow_known_unavailable", False)):
            for label in data.get("controller_labels", []):
                text = str(label).strip()
                if text:
                    allow_labels.add(text)

    for controller in runtime_job.get("controllers", []):
        total += 1
        controller_label = str(controller.get("controller_label", "")).strip()
        bacnet = controller.get("bacnet", {})
        host = str(bacnet.get("host", "")).strip()
        port = int(bacnet.get("port", 0))
        expected_instance = int(bacnet.get("device_instance", 0))

        probe = _probe_bip(
            host=host,
            port=port,
            expected_device_instance=expected_instance,
            timeout_seconds=args.timeout_seconds,
            retries=args.retries,
        )
        probe["controller_label"] = controller_label
        if controller_label in allow_labels and probe.get("status") == "unreachable_timeout":
            probe["status"] = "known_unavailable"
            probe["allow_known_unavailable"] = True

        status = str(probe.get("status"))
        status_counts[status] = status_counts.get(status, 0) + 1
        if status != "reachable_verified":
            unresolved += 1

        if args.strict:
            if status != "reachable_verified":
                strict_pass = False
        else:
            if status == "known_unavailable":
                if not bool(probe.get("allow_known_unavailable") is True):
                    strict_pass = False
            elif status != "reachable_verified":
                strict_pass = False

        artifact_file = bip_artifacts_dir / f"{controller_label}.json"
        artifact_file.write_text(json.dumps(probe, sort_keys=True), encoding="utf-8")
        rows.append(probe)

    summary = {
        "total": total,
        "found": total - unresolved,
        "unresolved": unresolved,
        "strict_mode": bool(args.strict),
        "strict_pass": bool(strict_pass),
        "status_counts": status_counts,
        "rows": rows,
    }

    summary_path = bip_artifacts_dir / "list-summary.json"
    summary_path.write_text(json.dumps(summary, sort_keys=True), encoding="utf-8")

    _append_event(
        logs_path,
        "bip_list_verified",
        {
            "total": total,
            "unresolved": unresolved,
            "strict_mode": bool(args.strict),
            "strict_pass": bool(strict_pass),
            "summary_json": str(summary_path.resolve()),
        },
    )
    print(json.dumps(summary, sort_keys=True))
    return 0 if strict_pass else 2


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Runtime skeleton CLI for commissioning app flows."
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    init_run = subparsers.add_parser("init-run", help="Initialize a run directory.")
    init_run.add_argument("--run-dir", required=True, type=Path)
    init_run.add_argument("--job-id", required=True)
    init_run.add_argument("--controllers-csv", required=True, type=Path)
    init_run.add_argument("--profiles-dir", required=True, type=Path)
    init_run.add_argument("--scenarios-dir", required=True, type=Path)
    init_run.set_defaults(handler=cmd_init_run)

    compile_import = subparsers.add_parser(
        "compile-import", help="Compile import model using run config."
    )
    compile_import.add_argument("--run-dir", required=True, type=Path)
    compile_import.set_defaults(handler=cmd_compile_import)

    validate_import = subparsers.add_parser(
        "validate-import",
        help="Dry-run compile to artifacts/import-validation (does not overwrite state/runtime-job.json).",
    )
    validate_import.add_argument("--run-dir", required=True, type=Path)
    validate_import.add_argument(
        "--output-dir",
        type=Path,
        help="Destination directory (default: <run-dir>/artifacts/import-validation).",
    )
    validate_import.set_defaults(handler=cmd_validate_import)

    op_gui = subparsers.add_parser(
        "operator-gui",
        help=(
            "Local browser UI (127.0.0.1) for common commissioning commands; "
            "runs tools/runtime/app.py subprocesses from tools/operator_gui_server.py."
        ),
    )
    op_gui.add_argument("--run-dir", required=True, type=Path)
    op_gui.add_argument(
        "--gui-host",
        default="127.0.0.1",
        help="Bind address (default 127.0.0.1; do not expose to untrusted networks).",
    )
    op_gui.add_argument("--gui-port", type=int, default=8765, help="HTTP port (default 8765).")
    op_gui.set_defaults(handler=cmd_operator_gui)

    print_graph = subparsers.add_parser(
        "print-job-graph",
        help="Print human-readable summary of controllers and commissioning flow sizes.",
    )
    print_graph.add_argument("--run-dir", required=True, type=Path)
    print_graph.set_defaults(handler=cmd_print_job_graph)

    export_summary = subparsers.add_parser(
        "export-run-summary",
        help="Write aggregated run summary JSON (controllers, flow hints, import/BIP flags).",
    )
    export_summary.add_argument("--run-dir", required=True, type=Path)
    export_summary.add_argument(
        "--output-json",
        type=Path,
        help="Destination JSON (default: <run-dir>/artifacts/run-summary.json).",
    )
    export_summary.add_argument(
        "--embed-import-report",
        action="store_true",
        help="Include full import-report.json object under key import_report when present.",
    )
    export_summary.add_argument(
        "--embed-bip-list-summary",
        action="store_true",
        help="Include full list-summary.json object under key bip_list_summary when present.",
    )
    export_summary.add_argument(
        "--output-csv",
        type=Path,
        help="Also write controller rollup as CSV (same rows as summary controllers).",
    )
    export_summary.set_defaults(handler=cmd_export_run_summary)

    export_cr = subparsers.add_parser(
        "export-commissioning-report",
        help="Print or copy artifacts/commissioning_report.json (point checkout / future rows).",
    )
    export_cr.add_argument("--run-dir", required=True, type=Path)
    export_cr.add_argument(
        "--output-json",
        type=Path,
        help="Optional copy destination; default prints to stdout.",
    )
    export_cr.add_argument(
        "--allow-empty",
        action="store_true",
        help=(
            "If no report exists yet: with --output-json write empty stub JSON; "
            "with --output-csv / --output-csv-unified / --output-html / --output-xlsx "
            "/ --output-pdf / --output-customer-html / --output-customer-pdf write "
            "headers-only (or empty HTML / empty sheet / empty PDF). "
            "At least one output path flag is required."
        ),
    )
    export_cr.add_argument(
        "--output-csv",
        type=Path,
        help="Also write thermal modulation rows (thermal_modulation_*) to CSV.",
    )
    export_cr.add_argument(
        "--output-csv-unified",
        type=Path,
        help=(
            "Also write one CSV with point checkout + modulation rows "
            "(shared columns; unused fields empty per kind)."
        ),
    )
    export_cr.add_argument(
        "--output-html",
        type=Path,
        help=(
            "Also write a simple HTML table (same rows as --output-csv-unified); "
            "open in a browser and print to PDF (no extra Python deps)."
        ),
    )
    export_cr.add_argument(
        "--output-xlsx",
        type=Path,
        help=(
            "Also write unified rows to an Excel workbook (.xlsx; requires openpyxl)."
        ),
    )
    export_cr.add_argument(
        "--xlsx-include-modulation",
        action="store_true",
        help=(
            "With --output-xlsx: add a second sheet ``modulation`` with the same rows as "
            "the customer modulation CSV (thermal_modulation_* only)."
        ),
    )
    export_cr.add_argument(
        "--output-pdf",
        type=Path,
        help=(
            "Also write unified rows to a landscape PDF table (.pdf; requires fpdf2)."
        ),
    )
    export_cr.add_argument(
        "--output-customer-html",
        type=Path,
        help=(
            "Also write a customer-facing HTML table: thermal modulation rows only "
            "(narrow columns; browser print-to-PDF)."
        ),
    )
    export_cr.add_argument(
        "--output-customer-pdf",
        type=Path,
        help=(
            "Also write a customer-facing landscape PDF: thermal modulation rows only "
            "(wider cells than --output-pdf; requires fpdf2)."
        ),
    )
    export_cr.add_argument(
        "--pdf-logo-image",
        type=Path,
        default=None,
        help=(
            "PNG/JPEG logo for PDF header. If omitted: "
            "<run-dir>/artifacts/branding/logo.png when present, else the "
            "neutral placeholder shipped under docs/examples/branding/, else a "
            "simple vector box."
        ),
    )
    export_cr.set_defaults(handler=cmd_export_commissioning_report)

    mod_sample = subparsers.add_parser(
        "append-commissioning-modulation-sample",
        help="Read allowlisted BACnet points and append thermal_modulation_sample to commissioning_report.json.",
    )
    mod_sample.add_argument("--run-dir", required=True, type=Path)
    mod_sample.add_argument("--controller-label", required=True)
    mod_sample.add_argument(
        "--read",
        action="append",
        required=True,
        help="Logical object id or object_id:property (repeat for multiple reads).",
    )
    mod_sample.add_argument("--technician-name", required=True)
    mod_sample.add_argument("--note", default="")
    mod_sample.add_argument("--step-id", default="")
    mod_sample.add_argument("--report-ref", default="")
    mod_sample.add_argument("--timeout-seconds", type=float, default=0.5)
    mod_sample.add_argument("--retries", type=int, default=1)
    mod_sample.add_argument("--bacnet-bind-port", type=int, default=0)
    mod_sample.add_argument("--apdu-timeout", type=float, default=None)
    mod_sample.set_defaults(handler=cmd_append_commissioning_modulation_sample)

    mod_batch = subparsers.add_parser(
        "append-commissioning-modulation-batch",
        help="Append thermal_modulation_batch from JSON (list of samples or {samples:[]}).",
    )
    mod_batch.add_argument("--run-dir", required=True, type=Path)
    mod_batch.add_argument("--input-json", required=True, type=Path)
    mod_batch.add_argument(
        "--default-technician",
        default="",
        help="Fallback technician_name when a sample omits it.",
    )
    mod_batch.add_argument("--timeout-seconds", type=float, default=0.5)
    mod_batch.add_argument("--retries", type=int, default=1)
    mod_batch.add_argument("--bacnet-bind-port", type=int, default=0)
    mod_batch.add_argument("--apdu-timeout", type=float, default=None)
    mod_batch.set_defaults(handler=cmd_append_commissioning_modulation_batch)

    mod_sweep = subparsers.add_parser(
        "bacnet-modulation-sweep",
        help=(
            "Write command_object_id percent, dwell, read SAT/RAT/context from a flow step's "
            "modulate_actuator_log_sat_for_report action; append thermal_modulation_sweep to report."
        ),
    )
    mod_sweep.add_argument("--run-dir", required=True, type=Path)
    mod_sweep.add_argument("--controller-label", required=True)
    mod_sweep.add_argument("--step-id", required=True)
    mod_sweep.add_argument(
        "--command-percent",
        type=float,
        default=None,
        help="Present-value to write to command_object_id (e.g. 50 for 50%%).",
    )
    mod_sweep.add_argument(
        "--command-percents",
        default="",
        help=(
            "Comma-separated sequence (e.g. 0,50,100). When non-empty, runs one sweep entry "
            "per value after each dwell; omit --command-percent or combine for a single step."
        ),
    )
    mod_sweep.add_argument(
        "--dwell-seconds",
        type=float,
        default=0.2,
        help="Sleep after each write before reads (default 0.2).",
    )
    mod_sweep.add_argument("--technician-name", required=True)
    mod_sweep.add_argument("--note", default="")
    mod_sweep.add_argument(
        "--report-ref",
        default="",
        help="Override report_ref when flow step has none.",
    )
    mod_sweep.add_argument("--timeout-seconds", type=float, default=0.5)
    mod_sweep.add_argument("--retries", type=int, default=1)
    mod_sweep.add_argument("--bacnet-bind-port", type=int, default=0)
    mod_sweep.add_argument("--apdu-timeout", type=float, default=None)
    mod_sweep.set_defaults(handler=cmd_bacnet_modulation_sweep)

    verify_sim = subparsers.add_parser(
        "verify-simulator", help="Run simulator verification for one scenario."
    )
    verify_sim.add_argument("--run-dir", required=True, type=Path)
    verify_sim.add_argument("--profile", required=True, choices=["ci", "lab", "multisubnet"])
    verify_sim.add_argument("--scenario", required=True)
    verify_sim.add_argument(
        "--strict",
        action="store_true",
        help="Pass strict flag through to simulator verification.",
    )
    verify_sim.set_defaults(handler=cmd_verify_simulator)

    probe_bip = subparsers.add_parser(
        "probe-bip", help="Probe one BACnet/IP endpoint and classify identity."
    )
    probe_bip.add_argument("--run-dir", required=True, type=Path)
    probe_bip.add_argument("--controller-label", required=True)
    probe_bip.add_argument("--timeout-seconds", type=float, default=0.5)
    probe_bip.add_argument("--retries", type=int, default=1)
    probe_bip.set_defaults(handler=cmd_probe_bip)

    verify_bip_list = subparsers.add_parser(
        "verify-bip-list", help="Probe all controllers in runtime job via BACnet/IP."
    )
    verify_bip_list.add_argument("--run-dir", required=True, type=Path)
    verify_bip_list.add_argument(
        "--strict",
        action="store_true",
        help="Require every controller to be reachable_verified.",
    )
    verify_bip_list.add_argument("--timeout-seconds", type=float, default=0.5)
    verify_bip_list.add_argument("--retries", type=int, default=1)
    verify_bip_list.add_argument(
        "--allow-known-unavailable",
        action="append",
        default=[],
        help="Controller labels allowed to classify as known_unavailable in non-strict mode.",
    )
    verify_bip_list.add_argument(
        "--known-unavailable-file",
        type=Path,
        help="Optional JSON file with {controller_labels:[], allow_known_unavailable:true}.",
    )
    verify_bip_list.set_defaults(handler=cmd_verify_bip_list)

    dry_write = subparsers.add_parser(
        "dry-run-bacnet-write",
        help="Validate allowlisted WriteProperty intent (default dry-run; no frame sent).",
    )
    dry_write.add_argument("--run-dir", required=True, type=Path)
    dry_write.add_argument("--controller-label", required=True)
    dry_write.add_argument(
        "--object-id",
        required=True,
        help="Profile object id (e.g. msv_test_mode).",
    )
    dry_write.add_argument(
        "--value",
        required=True,
        type=int,
        help="Integer present-value to write (e.g. MSV state number).",
    )
    dry_write.add_argument("--technician-name", required=True)
    dry_write.add_argument("--note", default="")
    dry_write.add_argument("--timeout-seconds", type=float, default=0.5)
    dry_write.add_argument("--retries", type=int, default=1)
    dry_write.add_argument(
        "--bacnet-bind-port",
        type=int,
        default=0,
        help="Local UDP bind port for BACpypes3 client (0 = OS-assigned).",
    )
    dry_write.add_argument(
        "--apdu-timeout",
        type=float,
        default=None,
        help="BACpypes3 confirmed service timeout in seconds for --execute (default: adapter default, typically 8).",
    )
    dry_write.add_argument(
        "--execute",
        action="store_true",
        help="Send WriteProperty via BACpypes3 (requires pip install -r requirements.txt).",
    )
    dry_write.set_defaults(handler=cmd_dry_run_bacnet_write)

    bacnet_read = subparsers.add_parser(
        "bacnet-read",
        help="ReadProperty via BACpypes3 for allowlisted profile object (requires bacpypes3).",
    )
    bacnet_read.add_argument("--run-dir", required=True, type=Path)
    bacnet_read.add_argument("--controller-label", required=True)
    bacnet_read.add_argument(
        "--object-id",
        required=True,
        help="Profile object id (must be in commissioning_read_allowlist).",
    )
    bacnet_read.add_argument(
        "--property",
        default="presentValue",
        help="BACnet property name (default: presentValue).",
    )
    bacnet_read.add_argument("--timeout-seconds", type=float, default=0.5)
    bacnet_read.add_argument("--retries", type=int, default=1)
    bacnet_read.add_argument(
        "--bacnet-bind-port",
        type=int,
        default=0,
        help="Local UDP bind port for BACpypes3 client (0 = OS-assigned).",
    )
    bacnet_read.add_argument(
        "--apdu-timeout",
        type=float,
        default=None,
        help="BACpypes3 ReadProperty timeout in seconds (default: adapter default, typically 8).",
    )
    bacnet_read.set_defaults(handler=cmd_bacnet_read)

    point_checkout = subparsers.add_parser(
        "bacnet-point-checkout",
        help="Read profile point_checkout list in order (BACpypes3; requires bacpypes3).",
    )
    point_checkout.add_argument("--run-dir", required=True, type=Path)
    point_checkout.add_argument("--controller-label", required=True)
    point_checkout.add_argument("--timeout-seconds", type=float, default=0.5)
    point_checkout.add_argument("--retries", type=int, default=1)
    point_checkout.add_argument(
        "--bacnet-bind-port",
        type=int,
        default=0,
        help="Local UDP bind port for BACpypes3 client (0 = OS-assigned).",
    )
    point_checkout.add_argument(
        "--apdu-timeout",
        type=float,
        default=None,
        help="BACpypes3 ReadProperty timeout in seconds for each checkout read (default: adapter default).",
    )
    point_checkout.add_argument(
        "--strict",
        action="store_true",
        help="Stop after first failed read instead of continuing.",
    )
    point_checkout.set_defaults(handler=cmd_bacnet_point_checkout)

    init_flow = subparsers.add_parser(
        "init-flow", help="Initialize commissioning flow state for one controller."
    )
    init_flow.add_argument("--run-dir", required=True, type=Path)
    init_flow.add_argument("--controller-label", required=True)
    init_flow.add_argument(
        "--force",
        action="store_true",
        help="Replace existing flow state; requires reset audit fields and backs up prior file.",
    )
    init_flow.add_argument(
        "--reset-technician-name",
        default="",
        help="With --force, who authorized replacing existing flow state.",
    )
    init_flow.add_argument(
        "--reset-reason",
        default="",
        help="With --force, why the prior flow state is being discarded.",
    )
    init_flow.set_defaults(handler=cmd_init_flow)

    list_flows = subparsers.add_parser(
        "list-flows",
        help="List commissioning flow state files for this run (summary JSON).",
    )
    list_flows.add_argument("--run-dir", required=True, type=Path)
    list_flows.set_defaults(handler=cmd_list_flows)

    show_flow = subparsers.add_parser(
        "show-flow",
        help="Print full commissioning flow JSON for one controller.",
    )
    show_flow.add_argument("--run-dir", required=True, type=Path)
    show_flow.add_argument("--controller-label", required=True)
    show_flow.set_defaults(handler=cmd_show_flow)

    guided_next = subparsers.add_parser(
        "commissioning-guided-next",
        help=(
            "Print compact commissioning flow guidance JSON: next open step, "
            "per-step status labels, session key list (after init-flow)."
        ),
    )
    guided_next.add_argument("--run-dir", required=True, type=Path)
    guided_next.add_argument("--controller-label", required=True)
    guided_next.set_defaults(handler=cmd_commissioning_guided_next)

    set_session = subparsers.add_parser(
        "set-session-value",
        help="Store operator-entered session value for a controller (e.g. manual RAT).",
    )
    set_session.add_argument("--run-dir", required=True, type=Path)
    set_session.add_argument("--controller-label", required=True)
    set_session.add_argument(
        "--key",
        required=True,
        help="Session field key (e.g. rat_degC).",
    )
    set_session.add_argument(
        "--value",
        required=True,
        help="Value to store (string; caller may pass numeric text).",
    )
    set_session.add_argument("--technician-name", required=True)
    set_session.add_argument("--note", default="")
    set_session.set_defaults(handler=cmd_set_session_value)

    confirm_prompt = subparsers.add_parser(
        "commissioning-confirm-prompt",
        help=(
            "CHW valve stroke (no CHW): re-write ao_chw_valve for a profile prompt_id, "
            "then set session prompt_confirm.<id> (required before record-step passed)."
        ),
    )
    confirm_prompt.add_argument("--run-dir", required=True, type=Path)
    confirm_prompt.add_argument("--controller-label", required=True)
    confirm_prompt.add_argument("--step-id", required=True)
    confirm_prompt.add_argument(
        "--prompt-id",
        required=True,
        help="Profile operator_prompt_confirm prompt_id (e.g. chw_valve_at_100).",
    )
    confirm_prompt.add_argument("--technician-name", required=True)
    confirm_prompt.add_argument("--note", default="")
    confirm_prompt.add_argument(
        "--bacnet-timeout-seconds",
        type=float,
        default=0.5,
        help="Who-Is timeout base for BACnet read/write in this command.",
    )
    confirm_prompt.add_argument(
        "--bacnet-retries",
        type=int,
        default=1,
        help="Who-Is retries for BACnet I/O.",
    )
    confirm_prompt.add_argument(
        "--bacnet-bind-port",
        type=int,
        default=0,
        help="Local UDP bind port (0 = OS-assigned).",
    )
    confirm_prompt.add_argument(
        "--apdu-timeout",
        type=float,
        default=None,
        help="BACpypes3 APDU timeout override (default: adapter default).",
    )
    confirm_prompt.set_defaults(handler=cmd_commissioning_confirm_prompt)

    confirm_tacho = subparsers.add_parser(
        "commissioning-confirm-tachometer-reference",
        help=(
            "Airflow checkpoint: BACnet-read tachometer from profile step "
            "operator_confirm_tachometer_reference and store session_key (required before record-step pass)."
        ),
    )
    confirm_tacho.add_argument("--run-dir", required=True, type=Path)
    confirm_tacho.add_argument("--controller-label", required=True)
    confirm_tacho.add_argument("--step-id", required=True)
    confirm_tacho.add_argument("--technician-name", required=True)
    confirm_tacho.add_argument("--note", default="")
    confirm_tacho.add_argument(
        "--bacnet-timeout-seconds",
        type=float,
        default=0.5,
        help="Who-Is timeout base for BACnet read.",
    )
    confirm_tacho.add_argument(
        "--bacnet-retries",
        type=int,
        default=1,
        help="Who-Is retries for BACnet I/O.",
    )
    confirm_tacho.add_argument(
        "--bacnet-bind-port",
        type=int,
        default=0,
        help="Local UDP bind port (0 = OS-assigned).",
    )
    confirm_tacho.add_argument(
        "--apdu-timeout",
        type=float,
        default=None,
        help="BACpypes3 APDU timeout override (default: adapter default).",
    )
    confirm_tacho.set_defaults(handler=cmd_commissioning_confirm_tachometer_reference)

    airflow_adj = subparsers.add_parser(
        "commissioning-airflow-adjust-write",
        help=(
            "Airflow checkpoint: WriteProperty presentValue on actuator_object_id from profile step "
            "automatic_airflow_adjustment (e.g. fan command %). When step arms airflow_verify, "
            "msv_test_mode must be state 3 first."
        ),
    )
    airflow_adj.add_argument("--run-dir", required=True, type=Path)
    airflow_adj.add_argument("--controller-label", required=True)
    airflow_adj.add_argument("--step-id", required=True)
    airflow_adj.add_argument(
        "--fan-command-percent",
        required=True,
        type=float,
        help="Actuator command 0–100 (written to profile automatic_airflow_adjustment actuator_object_id).",
    )
    airflow_adj.add_argument("--technician-name", required=True)
    airflow_adj.add_argument("--note", default="")
    airflow_adj.add_argument(
        "--bacnet-timeout-seconds",
        type=float,
        default=0.5,
        help="Who-Is timeout base for BACnet read/write.",
    )
    airflow_adj.add_argument(
        "--bacnet-retries",
        type=int,
        default=1,
        help="Who-Is retries for BACnet I/O.",
    )
    airflow_adj.add_argument(
        "--bacnet-bind-port",
        type=int,
        default=0,
        help="Local UDP bind port (0 = OS-assigned).",
    )
    airflow_adj.add_argument(
        "--apdu-timeout",
        type=float,
        default=None,
        help="BACpypes3 APDU timeout override (default: adapter default).",
    )
    airflow_adj.set_defaults(handler=cmd_commissioning_airflow_adjust_write)

    airflow_cl = subparsers.add_parser(
        "commissioning-airflow-closed-loop-iterate",
        help=(
            "Iteratively adjust automatic_airflow_adjustment actuator toward target L/s using "
            "BACnet flow_read_object_id feedback (profile closed_loop block; see docs)."
        ),
    )
    airflow_cl.add_argument("--run-dir", required=True, type=Path)
    airflow_cl.add_argument("--controller-label", required=True)
    airflow_cl.add_argument("--step-id", required=True)
    airflow_cl.add_argument(
        "--initial-fan-command-percent",
        type=float,
        default=None,
        help="Override profile closed_loop.initial_command_percent for this run.",
    )
    airflow_cl.add_argument("--technician-name", default="operator")
    airflow_cl.add_argument("--note", default="closed-loop iterate")
    airflow_cl.add_argument(
        "--bacnet-timeout-seconds",
        type=float,
        default=0.5,
        help="Who-Is timeout base for BACnet read/write.",
    )
    airflow_cl.add_argument(
        "--bacnet-retries",
        type=int,
        default=1,
        help="Who-Is retries for BACnet I/O.",
    )
    airflow_cl.add_argument(
        "--bacnet-bind-port",
        type=int,
        default=0,
        help="Local UDP bind port (0 = OS-assigned).",
    )
    airflow_cl.add_argument(
        "--apdu-timeout",
        type=float,
        default=None,
        help="BACpypes3 APDU timeout override (default: adapter default).",
    )
    airflow_cl.set_defaults(handler=cmd_commissioning_airflow_closed_loop_iterate)

    manual_air = subparsers.add_parser(
        "commissioning-record-manual-airflow",
        help=(
            "Record measured airflow (L/s) for manual_airflow_verification_assisted "
            "(session key per branch; required before record-step pass on that step)."
        ),
    )
    manual_air.add_argument("--run-dir", required=True, type=Path)
    manual_air.add_argument("--controller-label", required=True)
    manual_air.add_argument("--step-id", required=True)
    manual_air.add_argument(
        "--branch-id",
        required=True,
        help="Branch id from profile airflow_verification.branches[].id (must appear in step branch_ids).",
    )
    manual_air.add_argument(
        "--measured-flow-L-s",
        required=True,
        dest="measured_flow_L_s",
        help="Measured airflow in L/s (> 0).",
    )
    manual_air.add_argument(
        "--measurement-tool",
        required=True,
        help="Tool used (must be in profile branch measurement.allowed_tools when defined).",
    )
    manual_air.add_argument("--technician-name", required=True)
    manual_air.add_argument("--note", default="")
    manual_air.add_argument(
        "--bacnet-timeout-seconds",
        type=float,
        default=0.5,
        help="Who-Is timeout base when verifying MSV for airflow_verify.",
    )
    manual_air.add_argument(
        "--bacnet-retries",
        type=int,
        default=1,
        help="Who-Is retries for BACnet read.",
    )
    manual_air.add_argument(
        "--bacnet-bind-port",
        type=int,
        default=0,
        help="Local UDP bind port (0 = OS-assigned).",
    )
    manual_air.add_argument(
        "--apdu-timeout",
        type=float,
        default=None,
        help="BACpypes3 APDU timeout override (default: adapter default).",
    )
    manual_air.set_defaults(handler=cmd_commissioning_record_manual_airflow)

    show_session = subparsers.add_parser(
        "show-session",
        help="Print session values JSON for one controller.",
    )
    show_session.add_argument("--run-dir", required=True, type=Path)
    show_session.add_argument("--controller-label", required=True)
    show_session.set_defaults(handler=cmd_show_session)

    record_step = subparsers.add_parser(
        "record-step", help="Record technician signoff for a commissioning step."
    )
    record_step.add_argument("--run-dir", required=True, type=Path)
    record_step.add_argument("--controller-label", required=True)
    record_step.add_argument("--step-id", required=True)
    record_step.add_argument(
        "--status",
        required=True,
        choices=["pending", "passed", "failed", "skipped", "manual_passed"],
    )
    record_step.add_argument("--technician-name", required=True)
    record_step.add_argument("--note", default="")
    record_step.add_argument(
        "--bacnet-timeout-seconds",
        type=float,
        default=0.5,
        help="Who-Is / probe timeout base for automatic point checkout after passed/manual_passed.",
    )
    record_step.add_argument(
        "--bacnet-retries",
        type=int,
        default=1,
        help="Retries for probe-derived Who-Is timeout when running automatic point checkout.",
    )
    record_step.add_argument(
        "--bacnet-bind-port",
        type=int,
        default=0,
        help="Local UDP bind port for BACpypes3 during automatic point checkout (0 = OS-assigned).",
    )
    record_step.add_argument(
        "--apdu-timeout",
        type=float,
        default=None,
        help="BACpypes3 ReadProperty timeout for automatic point checkout (default: adapter default).",
    )
    record_step.add_argument(
        "--bacnet-checkout-strict",
        action="store_true",
        help="Stop point checkout after first failed read (default: run all points).",
    )
    record_step.add_argument(
        "--no-run-modulation-on-pass",
        dest="run_modulation_on_pass",
        action="store_false",
        help="Skip automatic modulation sweep on pass even if the step defines the action.",
    )
    record_step.add_argument(
        "--modulation-command-percents",
        default="",
        help="Comma list of command %% values for record-step modulation (required when step has modulation action).",
    )
    record_step.add_argument(
        "--modulation-dwell-seconds",
        type=float,
        default=0.2,
        help="Dwell after each modulation write before BACnet reads.",
    )
    record_step.add_argument(
        "--modulation-timeout-seconds",
        type=float,
        default=0.5,
        help="Who-Is timeout base for modulation sweep BACnet I/O.",
    )
    record_step.add_argument(
        "--modulation-retries",
        type=int,
        default=1,
        help="Who-Is retries for modulation sweep.",
    )
    record_step.add_argument(
        "--modulation-bacnet-bind-port",
        type=int,
        default=0,
        help="Local UDP bind port for modulation sweep (0 = OS-assigned).",
    )
    record_step.add_argument(
        "--modulation-apdu-timeout",
        type=float,
        default=None,
        help="BACpypes3 APDU timeout override for modulation sweep (default: adapter default).",
    )
    record_step.set_defaults(handler=cmd_record_step, run_modulation_on_pass=True)

    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    try:
        return args.handler(args)
    except (OSError, json.JSONDecodeError, KeyError) as err:
        print(f"error: {err}")
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
